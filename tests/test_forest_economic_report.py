import sys
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
WEB_APPS_DIR = ROOT_DIR / "web_apps"
if str(WEB_APPS_DIR) not in sys.path:
    sys.path.insert(0, str(WEB_APPS_DIR))

from forest_economic_report import write_forest_economic_report
from forest_integration import EcosystemUserInput, calculate_forest_valuation_bundle_from_outputs


class ForestEconomicReportTests(unittest.TestCase):
    def test_write_forest_economic_report_creates_master_component_and_audit_sheets(self):
        outputs = {
            "DETAIL_VOLUME": pd.DataFrame(
                [
                    {
                        "sheet_name": "comp_internal_a",
                        "block_type": "Tree",
                        "row_no": 3,
                        "Plot": "P1",
                        "TQ": 1,
                        "volume_m3": 2.0,
                        "thai_standard": "ประดู่",
                        "Species_norm": "ประดู่",
                        "Species_raw": "ประดู่",
                    },
                    {
                        "sheet_name": "comp_internal_b",
                        "block_type": "Tree",
                        "row_no": 4,
                        "Plot": "P2",
                        "TQ": 2,
                        "volume_m3": 3.0,
                        "thai_standard": "สัก",
                        "Species_norm": "สัก",
                        "Species_raw": "สัก",
                    },
                ]
            ),
            "DETAIL_TREE_BIOMASS": pd.DataFrame(
                [
                    {
                        "sheet_name": "comp_internal_a",
                        "row_no": 3,
                        "Plot": "P1",
                        "DBH_cm": 20.0,
                        "forest_type_clean": "ป่าเบญจพรรณ",
                        "forest_type_raw": "เบญจพรรณ",
                    },
                    {
                        "sheet_name": "comp_internal_b",
                        "row_no": 4,
                        "Plot": "P2",
                        "DBH_cm": 25.0,
                        "forest_type_clean": "ป่าดิบแห้ง",
                        "forest_type_raw": "ดิบแห้ง",
                    },
                ]
            ),
            "SUMMARY_ALL": pd.DataFrame(
                [
                    {
                        "sheet_name": "comp_internal_a",
                        "n_tree": 12,
                        "n_sapling": 4,
                        "total_seedling_number": 10,
                        "sapling_per_rai": 4.0,
                        "seedling_per_rai": 10.0,
                    },
                    {
                        "sheet_name": "comp_internal_b",
                        "n_tree": 8,
                        "n_sapling": 2,
                        "total_seedling_number": 5,
                        "sapling_per_rai": 2.0,
                        "seedling_per_rai": 5.0,
                    },
                ]
            ),
            "__meta__": {
                "plot_area_ha": 0.1,
                "rai_per_hectare": 6.25,
                "sheet_groups": [
                    {"internal_name": "comp_internal_a", "name": "component 1", "sheet_names": ["S1"]},
                    {"internal_name": "comp_internal_b", "name": "component 2", "sheet_names": ["S2"]},
                ],
            },
        }
        bundle = calculate_forest_valuation_bundle_from_outputs(
            outputs=outputs,
            component_area_inputs={"component 1": 10, "component 2": 20},
            ecosystem_user_inputs=[
                EcosystemUserInput(
                    component_name="component 1",
                    component_area_rai=10,
                    canopy_cover_percent=60,
                    canopy_layer_count=2,
                    soil_depth_m=0.3,
                    annual_rainfall_mm=1320,
                    topography_score=16,
                ),
                EcosystemUserInput(
                    component_name="component 2",
                    component_area_rai=20,
                    canopy_cover_percent=55,
                    canopy_layer_count=2,
                    soil_depth_m=0.25,
                    annual_rainfall_mm=1280,
                    topography_score=14,
                ),
            ],
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            report_path = Path(temp_dir) / "forest_economic_report.xlsx"
            write_forest_economic_report(report_path, outputs, bundle)
            workbook = load_workbook(report_path, data_only=True)

            self.assertIn("MASTER_SUMMARY", workbook.sheetnames)
            self.assertIn("MASTER_WARNINGS", workbook.sheetnames)
            self.assertTrue(any(name.startswith("COMP_01_component 1") for name in workbook.sheetnames))
            self.assertTrue(any(name.startswith("COMP_02_component 2") for name in workbook.sheetnames))
            self.assertIn("AUDIT_FOREST_ECON", workbook.sheetnames)
            self.assertIn("AUDIT_SPECIES", workbook.sheetnames)
            self.assertIn("AUDIT_FUTURE_VALUE", workbook.sheetnames)
            self.assertIn("AUDIT_ECOSYSTEM", workbook.sheetnames)

            master = workbook["MASTER_SUMMARY"]
            self.assertEqual(master["A1"].value, "สรุปการประเมินมูลค่าทรัพยากรป่าไม้และระบบนิเวศ")
            self.assertEqual(master["C2"].value, "component 1")
            self.assertEqual(master["D2"].value, "component 2")
            self.assertAlmostEqual(master["C6"].value, 16.0)
            self.assertAlmostEqual(master["D6"].value, 32.0)
            self.assertAlmostEqual(master["E6"].value, 48.0)
            self.assertAlmostEqual(master["C7"].value, 40.0)
            self.assertAlmostEqual(master["D7"].value, 40.0)
            self.assertAlmostEqual(master["E7"].value, 80.0)
            self.assertAlmostEqual(master["C8"].value, 100.0)
            self.assertAlmostEqual(master["D8"].value, 100.0)
            self.assertAlmostEqual(master["E8"].value, 200.0)
            component_sheet = workbook[next(name for name in workbook.sheetnames if name.startswith("COMP_01_component 1"))]
            self.assertEqual(component_sheet["A1"].value, "รายงานองค์ประกอบ: component 1")
            self.assertIn("พื้นที่โครงการ", str(component_sheet["A2"].value))


if __name__ == "__main__":
    unittest.main()
