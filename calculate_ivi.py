from __future__ import annotations

import argparse
import math
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


PLOT_AREA_HA = 0.4
RAI_PER_HECTARE = 6.25
PI = 3.14


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Calculate IVI from the tree section in an Excel workbook."
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        default=r"C:\tong\work\cal_Biomass\test_IVI.xlsx",
        help="Path to the source workbook.",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Path to the output workbook. Defaults to '<input>_calculated_IVI.xlsx'.",
    )
    return parser.parse_args()


def is_number(value: object) -> bool:
    if value is None or isinstance(value, bool):
        return False
    try:
        float(value)
    except (TypeError, ValueError):
        return False
    return True


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def safe_console_text(value: object) -> str:
    text = normalize_text(value)
    return text.encode("unicode_escape").decode() if text else ""


def find_tree_header_row(worksheet) -> int:
    for row_idx in range(1, min(worksheet.max_row, 10) + 1):
        values = [normalize_text(worksheet.cell(row_idx, col).value) for col in range(1, 10)]
        if "Species" in values and "DBH (cm)" in values and "Plot" in values:
            return row_idx
    raise ValueError(f"Could not find the tree header row in sheet '{worksheet.title}'.")


def extract_tree_data(worksheet) -> pd.DataFrame:
    header_row = find_tree_header_row(worksheet)
    records: list[dict[str, object]] = []

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        no_value = worksheet.cell(row_idx, 1).value
        species = normalize_text(worksheet.cell(row_idx, 2).value)
        dbh_value = worksheet.cell(row_idx, 3).value
        girth_value = worksheet.cell(row_idx, 4).value
        plot = normalize_text(worksheet.cell(row_idx, 6).value)

        if not is_number(no_value) or not species or not plot:
            continue

        dbh_cm = float(dbh_value) if is_number(dbh_value) else None
        if dbh_cm is None and is_number(girth_value):
            dbh_cm = float(girth_value) / PI

        if dbh_cm is None or dbh_cm <= 0:
            continue

        records.append(
            {
                "Species": species,
                "DBH_cm": dbh_cm,
                "Plot": plot,
            }
        )

    if not records:
        raise ValueError(f"No usable tree records were found in sheet '{worksheet.title}'.")

    return pd.DataFrame(records)


def build_ivi_table(tree_df: pd.DataFrame) -> pd.DataFrame:
    total_plots = tree_df["Plot"].nunique()
    if total_plots == 0:
        raise ValueError("No plots were found in the tree data.")

    summary = (
        tree_df.groupby("Species", sort=True)
        .agg(
            **{
                "Number of tree": ("Species", "size"),
                "Plot": ("Plot", "nunique"),
                "DBH_sum_cm": ("DBH_cm", "sum"),
            }
        )
        .reset_index()
    )

    summary["Density (tree/ha)"] = summary["Number of tree"] / (summary["Plot"] * PLOT_AREA_HA)
    summary["Density (tree/rai)"] = summary["Density (tree/ha)"] / RAI_PER_HECTARE
    summary["Frequency"] = summary["Plot"] / summary["Plot"].sum() * 100
    summary["DBH (m)"] = summary["DBH_sum_cm"] / 100
    summary["BA (m2)"] = ((summary["DBH (m)"] ** 2) / 4) * PI
    summary["Dominance"] = summary["BA (m2)"] / total_plots

    summary["RDensity"] = summary["Density (tree/rai)"] / summary["Density (tree/rai)"].sum() * 100
    summary["RFrequency"] = summary["Frequency"] / summary["Frequency"].sum() * 100
    summary["RDorminance"] = summary["Dominance"] / summary["Dominance"].sum() * 100
    summary["IVI"] = (
        summary["RDensity"] + summary["RFrequency"] + summary["RDorminance"]
    )

    total_trees = summary["Number of tree"].sum()
    summary["Pi"] = summary["Number of tree"] / total_trees
    summary["ln Pi"] = summary["Pi"].apply(math.log)
    summary["Pi (ln Pi)"] = summary["Pi"] * summary["ln Pi"]
    summary["Shannon contribution"] = -summary["Pi (ln Pi)"]

    summary = summary[
        [
            "Species",
            "Number of tree",
            "Density (tree/ha)",
            "Density (tree/rai)",
            "Plot",
            "Frequency",
            "DBH (m)",
            "BA (m2)",
            "Dominance",
            "RDensity",
            "RFrequency",
            "RDorminance",
            "IVI",
            "Pi",
            "ln Pi",
            "Pi (ln Pi)",
            "Shannon contribution",
        ]
    ]

    totals = {
        "Species": "SUM",
        "Number of tree": summary["Number of tree"].sum(),
        "Density (tree/ha)": summary["Density (tree/ha)"].sum(),
        "Density (tree/rai)": summary["Density (tree/rai)"].sum(),
        "Plot": summary["Plot"].sum(),
        "Frequency": summary["Frequency"].sum(),
        "DBH (m)": summary["DBH (m)"].sum(),
        "BA (m2)": summary["BA (m2)"].sum(),
        "Dominance": summary["Dominance"].sum(),
        "RDensity": summary["RDensity"].sum(),
        "RFrequency": summary["RFrequency"].sum(),
        "RDorminance": summary["RDorminance"].sum(),
        "IVI": summary["IVI"].sum(),
        "Pi": summary["Pi"].sum(),
        "ln Pi": summary["ln Pi"].sum(),
        "Pi (ln Pi)": summary["Pi (ln Pi)"].sum(),
        "Shannon contribution": summary["Shannon contribution"].sum(),
    }

    return pd.concat([summary, pd.DataFrame([totals])], ignore_index=True)


def autofit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 24)


def style_ivi_sheet(worksheet) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    total_font = Font(bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    last_row = worksheet.max_row
    for cell in worksheet[last_row]:
        cell.fill = total_fill
        cell.font = total_font

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.column != 2:
                cell.number_format = "0.000000"
    worksheet.freeze_panes = "A2"
    autofit_columns(worksheet)


def write_results(input_file: Path, output_file: Path, results: dict[str, pd.DataFrame]) -> None:
    workbook = load_workbook(input_file)

    for source_sheet_name, table in results.items():
        ivi_sheet_name = f"IVI_{source_sheet_name}"[:31]
        if ivi_sheet_name in workbook.sheetnames:
            del workbook[ivi_sheet_name]

        ws = workbook.create_sheet(ivi_sheet_name)
        ws.append(list(table.columns))
        for row in table.itertuples(index=False, name=None):
            ws.append(list(row))
        style_ivi_sheet(ws)

    workbook.save(output_file)


def main() -> None:
    args = parse_args()
    input_file = Path(args.input_file)
    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")

    output_file = (
        Path(args.output)
        if args.output
        else input_file.with_name(f"{input_file.stem}_calculated_IVI.xlsx")
    )

    wb_values = load_workbook(input_file, data_only=True)
    source_sheet_names = [name for name in wb_values.sheetnames if not name.startswith("IVI_")]
    if not source_sheet_names:
        raise ValueError("No source data sheets were found. Expected at least one non-IVI sheet.")

    results: dict[str, pd.DataFrame] = {}
    for sheet_name in source_sheet_names:
        tree_df = extract_tree_data(wb_values[sheet_name])
        results[sheet_name] = build_ivi_table(tree_df)

    write_results(input_file, output_file, results)

    print(f"Calculated IVI for {len(results)} sheet(s).")
    print(f"Output saved to: {output_file}")
    for sheet_name, table in results.items():
        shannon_index = table.iloc[:-1]["Shannon contribution"].sum()
        print(
            f"- {safe_console_text(sheet_name)}: {len(table) - 1} species, "
            f"Shannon index = {shannon_index:.6f}"
        )


if __name__ == "__main__":
    main()
