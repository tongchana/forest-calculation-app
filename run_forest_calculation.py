from __future__ import annotations

import argparse
import logging
import math
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PLOT_AREA_HA = 0.4
RAI_PER_HECTARE = 6.25
PI = math.pi

SKIP_SHEET_PREFIXES = ("OUT_", "SUMMARY", "DETAIL_", "IVI_", "CHECK_")
SKIP_SHEET_NAMES = {
    "alias_map",
    "taxon_master",
    "species_master",
    "species_reference",
    "manual_review",
    "readme",
}

TREE_COLUMNS = {
    "No.": 0,
    "Species": 1,
    "DBH (cm)": 2,
    "Girth (cm)": 3,
    "Height (m)": 4,
    "Plot": 5,
    "TQ": 6,
    "forest type": 7,
}
SAPLING_COLUMNS = {
    "No.": 8,
    "Species": 9,
    "Girth (cm)": 10,
    "Height (m)": 11,
    "Number": 12,
    "Plot": 13,
}
SEEDLING_COLUMNS = {
    "No.": 14,
    "Species": 15,
    "Number": 16,
    "Plot": 17,
}
BAMBOO_COLUMNS = {
    "No.": 18,
    "Species": 19,
    "Culm": 20,
}

HEADER_ALIASES = {
    "No.": {"no", "no.", "ลำดับ"},
    "Species": {"species", "ชนิด", "ชนิดไม้"},
    "DBH (cm)": {"dbh", "dbh(cm)", "dbhcm", "dbh_cm"},
    "Girth (cm)": {"girth", "girth(cm)", "girthcm", "girth_cm", "grith"},
    "Height (m)": {"height", "height(m)", "heightm", "ht", "h", "height_m"},
    "Plot": {"plot", "plotid", "plot_id"},
    "TQ": {"tq"},
    "forest type": {"foresttype", "forest_type", "ชนิดป่า"},
    "Number": {"number", "no.tree", "no tree", "จำนวน"},
    "Culm": {"culm", "จำนวนลำ"},
}

FOREST_MAP = {
    "ดิบแล้ง": "ป่าดิบแล้ง",
    "ป่าดิบแล้ง": "ป่าดิบแล้ง",
    "ดิบเขา": "ป่าดิบเขา",
    "ป่าดิบเขา": "ป่าดิบเขา",
    "ดิบชื้น": "ป่าดิบชื้น",
    "ป่าดิบชื้น": "ป่าดิบชื้น",
    "เบญจพรรณ": "ป่าเบญจพรรณ",
    "ป่าเบญจพรรณ": "ป่าเบญจพรรณ",
    "เต็งรัง": "ป่าเต็งรัง",
    "ป่าเต็งรัง": "ป่าเต็งรัง",
    "สนสองใบ": "ป่าสนเขา(สนสองใบ)",
    "ป่าสนเขา(สนสองใบ)": "ป่าสนเขา(สนสองใบ)",
    "ป่าสนเขาสนสองใบ": "ป่าสนเขา(สนสองใบ)",
    "สนสามใบ": "ป่าสนเขา(สนสามใบ)",
    "ป่าสนเขา(สนสามใบ)": "ป่าสนเขา(สนสามใบ)",
    "ป่าสนเขาสนสามใบ": "ป่าสนเขา(สนสามใบ)",
}
VALID_FORESTS = {
    "ป่าดิบแล้ง",
    "ป่าดิบเขา",
    "ป่าดิบชื้น",
    "ป่าเบญจพรรณ",
    "ป่าเต็งรัง",
    "ป่าสนเขา(สนสองใบ)",
    "ป่าสนเขา(สนสามใบ)",
}
JUNK_FOREST_VALUES = {"", "nan", "none", "null", "ชนิดป่า", "ชิดป่า", "ประเภทป่า"}

EQUATIONS = {
    1: {"a": 2.372083, "b": 2.443847},
    2: {"a": 2.134494, "b": 2.363034},
    3: {"a": 1.880578, "b": 2.053321},
    4: {"a": 1.789563, "b": 2.025666},
    5: {"a": 2.037096, "b": 2.299618},
    6: {"a": 2.119907, "b": 2.296511},
    7: {"a": 2.250111, "b": 2.414209},
}

OUTPUT_SHEETS = [
    "DETAIL_TREE_BIOMASS",
    "SUMMARY_BIOMASS",
    "DETAIL_VOLUME",
    "SUMMARY_VOLUME",
    "CHECK_UNMATCHED_SPECIES",
    "DETAIL_IVI",
    "SUMMARY_SHANNON",
    "DETAIL_SEEDLING",
    "SUMMARY_SEEDLING",
    "DETAIL_BAMBOO",
    "SUMMARY_BAMBOO",
    "SUMMARY_ALL",
]
DETAIL_WORKBOOK_SHEETS = [
    "DETAIL_TREE_BIOMASS",
    "DETAIL_VOLUME",
    "DETAIL_IVI",
    "DETAIL_SEEDLING",
    "DETAIL_BAMBOO",
    "CHECK_UNMATCHED_SPECIES",
]

SUMMARY_SECTION_SPECS = [
    ("Overall Summary", "SUMMARY_ALL"),
    ("Biomass Summary", "SUMMARY_BIOMASS"),
    ("Volume Summary", "SUMMARY_VOLUME"),
    ("Shannon Summary", "SUMMARY_SHANNON"),
    ("IVI Summary", "DETAIL_IVI"),
    ("General Tree Stand Summary", None),
    ("DBH Class Summary", None),
    ("TQ Volume Summary", None),
]

LOG = logging.getLogger("forest_calculation")

SUMMARY_SECTION_COLORS = {
    "Overall Summary": "FFF2CC",
    "Biomass Summary": "E2F0D9",
    "Volume Summary": "FCE4D6",
    "Shannon Summary": "D9EAF7",
    "IVI Summary": "D9EAF7",
    "General Tree Stand Summary": "FFF2CC",
    "DBH Class Summary": "E2F0D9",
    "TQ Volume Summary": "FCE4D6",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run biomass, volume, IVI/Shannon, seedling, and bamboo summaries from the forest template workbook."
    )
    parser.add_argument("input_file", help="Path to template.xlsx")
    parser.add_argument("--master", required=True, help="Path to species_reference_master_v1.xlsx")
    parser.add_argument("-o", "--output", default=None, help="Path to the output workbook")
    parser.add_argument("--plot-area-ha", type=float, default=PLOT_AREA_HA, help="Plot area in hectares")
    parser.add_argument(
        "--rai-per-hectare",
        type=float,
        default=RAI_PER_HECTARE,
        help="Rai per hectare conversion factor",
    )
    return parser.parse_args()


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if pd.isna(value):
        return ""
    return str(value).strip()


def display_or_dash(value: object) -> object:
    if value is None or pd.isna(value):
        return "-"
    return value


def safe_divide(numerator: float, denominator: float) -> float | str:
    if denominator in (None, 0) or pd.isna(denominator):
        return "-"
    return numerator / denominator


def normalize_header(value: object) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"\s+", "", text)
    text = text.replace("_", "")
    return text


def normalize_species_name(value: object) -> str | None:
    text = normalize_text(value).lower()
    if not text:
        return None
    text = re.sub(r"\s+", " ", text)
    for ch in [".", ",", ";", "(", ")", "[", "]", "{", "}"]:
        text = text.replace(ch, "")
    replacements = {
        "ไม้ขี้เหล็ก": "ขี้เหล็ก",
        "ไม้ประดู่": "ประดู่",
        "ไม้สัก": "สัก",
        "ไม้แดง": "แดง",
        "ไม้รัง": "รัง",
    }
    return replacements.get(text, text)


def normalize_forest_type(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    compact = re.sub(r"\s+", "", text).lower()
    if compact in JUNK_FOREST_VALUES:
        return ""
    return FOREST_MAP.get(compact, "")


def girth_to_dbh_cm(girth_cm: object) -> float:
    if pd.isna(girth_cm) or girth_cm is None:
        return np.nan
    try:
        girth = float(girth_cm)
    except (TypeError, ValueError):
        return np.nan
    if girth <= 0:
        return np.nan
    return girth / PI


def get_dbh_cm(dbh_cm: object, girth_cm: object) -> float:
    if pd.notna(dbh_cm):
        try:
            dbh = float(dbh_cm)
        except (TypeError, ValueError):
            dbh = np.nan
        if pd.notna(dbh) and dbh > 0:
            return dbh
    return girth_to_dbh_cm(girth_cm)


def calculate_tree_biomass(dbh_cm: float, height_m: float, forest_type: str) -> dict[str, float]:
    if pd.isna(dbh_cm) or pd.isna(height_m) or not forest_type:
        return {"Ws": np.nan, "Wb": np.nan, "Wl": np.nan, "Wr": np.nan, "biomass_total": np.nan}

    d = float(dbh_cm)
    h = float(height_m)
    ws = wb = wl = wr = np.nan

    if forest_type in {"ป่าดิบแล้ง", "ป่าดิบเขา"}:
        ws = 0.0509 * (d**2) * (h**0.919)
        wb = 0.00893 * (d**2) * (h**0.977)
        wl = 0.0140 * (d**2) * (h**0.669)
        wr = 0.0313 * (d**2) * (h**0.805)
    elif forest_type in {"ป่าเบญจพรรณ", "ป่าเต็งรัง"}:
        ws = 0.0396 * (d**2) * (h**0.9326)
        wb = 0.003487 * (d**2) * (h**1.0270)
        wtc = ws + wb
        wl = (28 / wtc + 0.0250) ** (-1)
    elif forest_type == "ป่าดิบชื้น":
        ws = 0.0396 * (d**2) * (h**0.9326)
        wb = 0.006002 * (d**2) * (h**1.0270)
        wtc = ws + wb
        wl = (18 / wtc + 0.0250) ** (-1)
        wr = 0.0264 * (d**2) * (h**0.7750)
    elif forest_type == "ป่าสนเขา(สนสองใบ)":
        ws = 0.02141 * (d**2) * (h**0.9814)
        wb = 0.00002 * (d**2) * (h**1.4561)
        wl = 0.00072 * (d**2) * (h**1.0138)
    elif forest_type == "ป่าสนเขา(สนสามใบ)":
        ws = 0.02698 * (d**2) * (h**0.9460)
        wb = 0.00018 * (d**2) * (h**1.4550)
        wl = 0.00072 * (d**2) * (h**1.0940)

    values = [ws, wb, wl, wr]
    total = np.nansum(values) if any(pd.notna(v) for v in values) else np.nan
    return {"Ws": ws, "Wb": wb, "Wl": wl, "Wr": wr, "biomass_total": total}


def calculate_volume_from_dbh(dbh_cm: object, group_id: object) -> float:
    if pd.isna(dbh_cm) or pd.isna(group_id):
        return np.nan
    try:
        dbh = float(dbh_cm)
        group = int(group_id)
    except (TypeError, ValueError):
        return np.nan
    if dbh <= 0:
        return np.nan
    if group not in EQUATIONS:
        group = 7
    a = EQUATIONS[group]["a"]
    b = EQUATIONS[group]["b"]
    return math.exp(a + b * math.log(dbh / 100.0))


def should_skip_sheet(sheet_name: str) -> bool:
    lower_name = sheet_name.strip().lower()
    if lower_name in SKIP_SHEET_NAMES:
        return True
    return any(sheet_name.startswith(prefix) for prefix in SKIP_SHEET_PREFIXES)


def safe_sheet_name(name: str, existing_names: set[str]) -> str:
    cleaned = re.sub(r'[:\\/?*\[\]]', "", normalize_text(name))
    cleaned = cleaned or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 1
    while candidate in existing_names:
        suffix_text = f"_{suffix}"
        candidate = f"{cleaned[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    existing_names.add(candidate)
    return candidate


def format_tq_label(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return "-"
    numeric = pd.to_numeric(pd.Series([text]), errors="coerce").iloc[0]
    if pd.notna(numeric):
        return f"TQ {numeric:g}"
    return text if text.lower().startswith("tq") else f"TQ {text}"


def tq_sort_key(label: str) -> tuple[int, float | str]:
    match = re.match(r"^TQ\s+(.+)$", label.strip(), re.IGNORECASE)
    if match:
        numeric = pd.to_numeric(pd.Series([match.group(1)]), errors="coerce").iloc[0]
        if pd.notna(numeric):
            return (0, float(numeric))
    return (1, label)


def validate_block_headers(sheet_name: str, row_headers: list[object], block_name: str, expected_columns: dict[str, int]) -> bool:
    missing: list[str] = []
    for expected_header, idx in expected_columns.items():
        actual = row_headers[idx] if idx < len(row_headers) else None
        actual_normalized = normalize_header(actual)
        allowed = HEADER_ALIASES[expected_header]
        if actual_normalized not in allowed:
            missing.append(f"{expected_header} (found '{normalize_text(actual)}')")
    if missing:
        LOG.warning("Sheet '%s' has invalid %s headers: %s", sheet_name, block_name, ", ".join(missing))
        return False
    return True


def read_sheet_frame(input_file: Path, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(input_file, sheet_name=sheet_name, header=1, dtype=object)


def list_processable_sheet_names(input_file: Path) -> list[str]:
    workbook_headers = pd.read_excel(input_file, sheet_name=None, header=None, nrows=2)
    return [
        sheet_name
        for sheet_name, header_df in workbook_headers.items()
        if not should_skip_sheet(sheet_name) and header_df.shape[0] >= 2
    ]


def prepare_block_frame(
    source_df: pd.DataFrame,
    sheet_name: str,
    block_type: str,
    expected_columns: dict[str, int],
    numeric_columns: list[str],
) -> pd.DataFrame:
    records: list[dict[str, object]] = []
    for excel_row_no, row in enumerate(source_df.itertuples(index=False, name=None), start=3):
        record: dict[str, object] = {
            "sheet_name": sheet_name,
            "source_sheet_name": sheet_name,
            "block_type": block_type,
            "row_no": excel_row_no,
        }
        for header, col_idx in expected_columns.items():
            value = row[col_idx] if col_idx < len(row) else None
            if header in {"Species", "Plot", "forest type"}:
                record[header] = normalize_text(value)
            else:
                record[header] = value
        if not record.get("No.") and not record.get("Species"):
            continue
        if normalize_text(record.get("No.")) == "" and normalize_text(record.get("Species")) == "":
            continue
        if normalize_text(record.get("Species")) == "":
            continue
        no_numeric = pd.to_numeric(pd.Series([record.get("No.")]), errors="coerce").iloc[0]
        if pd.isna(no_numeric):
            continue
        if block_type in {"Tree", "Sapling", "Seedling"} and normalize_text(record.get("Plot")) == "":
            continue
        records.append(record)

    frame = pd.DataFrame(records)
    if frame.empty:
        return frame

    for column in numeric_columns:
        if column in frame.columns:
            frame[column] = pd.to_numeric(frame[column], errors="coerce")
    return frame


def normalize_sheet_groups(
    sheet_groups: list[dict[str, object]] | None,
    available_sheet_names: list[str],
) -> list[dict[str, list[str]]]:
    if not sheet_groups:
        return []

    available_set = set(available_sheet_names)
    normalized: list[dict[str, list[str]]] = []
    seen_names: set[str] = set()

    for raw_group in sheet_groups:
        name = normalize_text(raw_group.get("name"))
        if not name:
            raise ValueError("Each sheet group must have a name.")
        if name in seen_names:
            raise ValueError(f"Duplicate sheet group name: {name}")
        if name in available_set:
            raise ValueError(f"Sheet group name '{name}' conflicts with an existing worksheet name.")

        raw_sheets = raw_group.get("sheet_names") or []
        group_sheet_names: list[str] = []
        seen_sheet_names: set[str] = set()
        for sheet_name in raw_sheets:
            normalized_sheet_name = normalize_text(sheet_name)
            if not normalized_sheet_name:
                continue
            if normalized_sheet_name not in available_set:
                raise ValueError(f"Worksheet '{normalized_sheet_name}' was not found in the uploaded workbook.")
            if normalized_sheet_name in seen_sheet_names:
                continue
            seen_sheet_names.add(normalized_sheet_name)
            group_sheet_names.append(normalized_sheet_name)

        if not group_sheet_names:
            continue

        normalized.append({"name": name, "sheet_names": group_sheet_names})
        seen_names.add(name)

    return normalized


def append_grouped_records(frame: pd.DataFrame, sheet_groups: list[dict[str, list[str]]]) -> pd.DataFrame:
    if frame.empty or not sheet_groups or "sheet_name" not in frame.columns:
        return frame

    grouped_frames = [frame]
    for group in sheet_groups:
        group_rows = frame[frame["sheet_name"].isin(group["sheet_names"])].copy()
        if group_rows.empty:
            continue
        group_rows["sheet_name"] = group["name"]
        grouped_frames.append(group_rows)

    if len(grouped_frames) == 1:
        return frame
    return pd.concat(grouped_frames, ignore_index=True)


def get_component_sheet_names(sheets: dict[str, pd.DataFrame]) -> set[str]:
    meta = sheets.get("__meta__", {})
    raw_groups = meta.get("sheet_groups", []) if isinstance(meta, dict) else []
    return {
        normalize_text(group.get("name"))
        for group in raw_groups
        if normalize_text(group.get("name"))
    }


def filter_out_component_rows(frame: pd.DataFrame, component_names: set[str]) -> pd.DataFrame:
    if frame.empty or not component_names or "sheet_name" not in frame.columns:
        return frame
    return frame[~frame["sheet_name"].astype(str).isin(component_names)].copy()


def load_master_reference(master_file: Path) -> dict[str, dict[str, object]]:
    xls = pd.ExcelFile(master_file)
    taxon_df = pd.read_excel(master_file, sheet_name="taxon_master") if "taxon_master" in xls.sheet_names else pd.DataFrame()
    alias_df = pd.read_excel(master_file, sheet_name="alias_map") if "alias_map" in xls.sheet_names else pd.DataFrame()

    if taxon_df.empty and alias_df.empty:
        raise ValueError("Master workbook must contain alias_map and/or taxon_master sheets.")

    taxon_by_key: dict[str, dict[str, object]] = {}
    if not taxon_df.empty:
        for _, row in taxon_df.iterrows():
            standard_key = normalize_text(row.get("standard_key"))
            if standard_key:
                taxon_by_key[standard_key] = {
                    "thai_standard": row.get("thai_standard"),
                    "scientific_name": row.get("scientific_name"),
                    "group_id": row.get("group_id"),
                }

    records: list[dict[str, object]] = []

    if not alias_df.empty:
        for _, row in alias_df.iterrows():
            raw_name = row.get("raw_name")
            normalized_name = row.get("normalized_name")
            standard_key = normalize_text(row.get("standard_key"))
            taxon_info = taxon_by_key.get(standard_key, {})
            thai_standard = taxon_info.get("thai_standard") or row.get("thai_standard") or raw_name
            scientific_name = taxon_info.get("scientific_name") or row.get("scientific_name")
            group_id = row.get("group_id")
            if pd.isna(group_id):
                group_id = taxon_info.get("group_id")

            for candidate in [raw_name, normalized_name]:
                lookup_name = normalize_species_name(candidate)
                if not lookup_name:
                    continue
                records.append(
                    {
                        "lookup_name": lookup_name,
                        "thai_standard": thai_standard,
                        "scientific_name": scientific_name,
                        "group_id": group_id,
                    }
                )

    if not taxon_df.empty:
        for _, row in taxon_df.iterrows():
            thai_standard = row.get("thai_standard")
            scientific_name = row.get("scientific_name")
            english_common = row.get("english_common")
            group_id = row.get("group_id")
            for candidate in [thai_standard, scientific_name, english_common]:
                if pd.isna(candidate) or normalize_text(candidate) == "":
                    continue
                records.append(
                    {
                        "lookup_name": normalize_species_name(candidate),
                        "thai_standard": thai_standard,
                        "scientific_name": scientific_name,
                        "group_id": group_id,
                    }
                )

    ref_df = pd.DataFrame(records).dropna(subset=["lookup_name"]).drop_duplicates(subset=["lookup_name"], keep="first")
    if ref_df.empty:
        raise ValueError("No usable species mappings were found in the master workbook.")

    ref_map: dict[str, dict[str, object]] = {}
    for _, row in ref_df.iterrows():
        group_id = pd.to_numeric(pd.Series([row.get("group_id")]), errors="coerce").iloc[0]
        ref_map[row["lookup_name"]] = {
            "thai_standard": row.get("thai_standard"),
            "scientific_name": row.get("scientific_name"),
            "group_id": int(group_id) if pd.notna(group_id) else 7,
        }
    return ref_map


def build_biomass_outputs(tree_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if tree_df.empty:
        detail_columns = [
            "sheet_name",
            "row_no",
            "No.",
            "Species",
            "DBH_cm",
            "Girth_cm",
            "Height_m",
            "Plot",
            "TQ",
            "forest_type_raw",
            "forest_type_clean",
            "Ws",
            "Wb",
            "Wl",
            "Wr",
            "biomass_total",
        ]
        summary_columns = [
            "sheet_name",
            "forest_type",
            "n_tree",
            "Ws_sum",
            "Wb_sum",
            "Wl_sum",
            "Wr_sum",
            "biomass_total_sum",
        ]
        return pd.DataFrame(columns=detail_columns), pd.DataFrame(columns=summary_columns)

    biomass_df = tree_df.copy()
    biomass_df["DBH_cm"] = biomass_df.apply(lambda row: get_dbh_cm(row["DBH (cm)"], row["Girth (cm)"]), axis=1)
    biomass_df["forest_type_raw"] = biomass_df["forest type"].fillna("").astype(str)
    biomass_df["forest_type_clean"] = biomass_df["forest type"].apply(normalize_forest_type)
    biomass_df["Height_m"] = biomass_df["Height (m)"]
    biomass_df["Girth_cm"] = biomass_df["Girth (cm)"]

    eligible = biomass_df[
        biomass_df["DBH_cm"].notna()
        & biomass_df["Height_m"].notna()
        & biomass_df["forest_type_clean"].isin(VALID_FORESTS)
    ].copy()

    if eligible.empty:
        detail = pd.DataFrame(
            columns=[
                "sheet_name",
                "row_no",
                "No.",
                "Species",
                "DBH_cm",
                "Girth_cm",
                "Height_m",
                "Plot",
                "TQ",
                "forest_type_raw",
                "forest_type_clean",
                "Ws",
                "Wb",
                "Wl",
                "Wr",
                "biomass_total",
            ]
        )
        summary = pd.DataFrame(
            columns=["sheet_name", "forest_type", "n_tree", "Ws_sum", "Wb_sum", "Wl_sum", "Wr_sum", "biomass_total_sum"]
        )
        return detail, summary

    biomass_values = eligible.apply(
        lambda row: pd.Series(calculate_tree_biomass(row["DBH_cm"], row["Height_m"], row["forest_type_clean"])),
        axis=1,
    )
    eligible = pd.concat([eligible, biomass_values], axis=1)
    detail = eligible[
        [
            "sheet_name",
            "row_no",
            "No.",
            "Species",
            "DBH_cm",
            "Girth_cm",
            "Height_m",
            "Plot",
            "TQ",
            "forest_type_raw",
            "forest_type_clean",
            "Ws",
            "Wb",
            "Wl",
            "Wr",
            "biomass_total",
        ]
    ].copy()

    summary = (
        detail.groupby(["sheet_name", "forest_type_clean"], dropna=False)
        .agg(
            n_tree=("Species", "size"),
            Ws_sum=("Ws", "sum"),
            Wb_sum=("Wb", "sum"),
            Wl_sum=("Wl", "sum"),
            Wr_sum=("Wr", "sum"),
            biomass_total_sum=("biomass_total", "sum"),
        )
        .reset_index()
        .rename(columns={"forest_type_clean": "forest_type"})
    )
    return detail, summary


def build_volume_outputs(tree_df: pd.DataFrame, sapling_df: pd.DataFrame, ref_map: dict[str, dict[str, object]]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    detail_records: list[dict[str, object]] = []

    def process_frame(frame: pd.DataFrame, block_type: str) -> None:
        if frame.empty:
            return
        for _, row in frame.iterrows():
            species_raw = row["Species"]
            species_norm = normalize_species_name(species_raw)
            ref = ref_map.get(species_norm) if species_norm else None
            dbh_cm = get_dbh_cm(row.get("DBH (cm)"), row.get("Girth (cm)"))
            group_id = ref["group_id"] if ref else np.nan
            volume_m3 = calculate_volume_from_dbh(dbh_cm, group_id) if ref else np.nan
            detail_records.append(
                {
                    "sheet_name": row["sheet_name"],
                    "block_type": block_type,
                    "row_no": row["row_no"],
                    "No.": row.get("No."),
                    "Species_raw": species_raw,
                    "Species_norm": species_norm,
                    "thai_standard": ref["thai_standard"] if ref else None,
                    "scientific_name": ref["scientific_name"] if ref else None,
                    "DBH_cm": dbh_cm,
                    "Girth_cm": row.get("Girth (cm)"),
                    "Height_m": row.get("Height (m)"),
                    "Number": row.get("Number"),
                    "Plot": row.get("Plot"),
                    "TQ": row.get("TQ"),
                    "group_id": group_id,
                    "volume_m3": volume_m3,
                    "matched": bool(ref),
                }
            )

    process_frame(tree_df, "Tree")
    process_frame(sapling_df, "Sapling")

    detail = pd.DataFrame(detail_records)
    if detail.empty:
        detail = pd.DataFrame(
            columns=[
                "sheet_name",
                "block_type",
                "row_no",
                "No.",
                "Species_raw",
                "Species_norm",
                "thai_standard",
                "scientific_name",
                "DBH_cm",
                "Girth_cm",
                "Height_m",
                "Number",
                "Plot",
                "TQ",
                "group_id",
                "volume_m3",
                "matched",
            ]
        )

    unmatched = detail[~detail["matched"]].copy() if not detail.empty else pd.DataFrame()
    if unmatched.empty:
        unmatched = pd.DataFrame(
            columns=["sheet_name", "block_type", "row_no", "Species_raw", "Species_norm", "Girth_cm", "DBH_cm"]
        )
    else:
        unmatched = unmatched[
            ["sheet_name", "block_type", "row_no", "Species_raw", "Species_norm", "Girth_cm", "DBH_cm"]
        ].copy()

    if detail.empty:
        summary = pd.DataFrame(columns=["sheet_name", "block_type", "n_records", "n_matched", "n_unmatched", "total_volume_m3"])
    else:
        summary = (
            detail.groupby(["sheet_name", "block_type"], dropna=False)
            .agg(
                n_records=("Species_raw", "size"),
                n_matched=("matched", lambda s: int(pd.Series(s).sum())),
                n_unmatched=("matched", lambda s: int((~pd.Series(s)).sum())),
                total_volume_m3=("volume_m3", "sum"),
            )
            .reset_index()
        )
    return detail, summary, unmatched


def build_ivi_outputs(tree_df: pd.DataFrame, plot_area_ha: float, rai_per_hectare: float) -> tuple[pd.DataFrame, pd.DataFrame]:
    detail_columns = [
        "sheet_name",
        "Species",
        "Number of tree",
        "Density (tree/ha)",
        "Density (tree/rai)",
        "Plot",
        "Frequency",
        "DBH (m)",
        "BA (m2)",
        "Dominance",
        "RDensity",
        "RFrequency",
        "RDominance",
        "IVI",
        "Pi",
        "ln Pi",
        "Pi (ln Pi)",
        "Shannon contribution",
    ]
    summary_columns = ["sheet_name", "Shannon_index", "n_species", "n_tree"]

    if tree_df.empty:
        return pd.DataFrame(columns=detail_columns), pd.DataFrame(columns=summary_columns)

    working = tree_df.copy()
    working["DBH_cm"] = working.apply(lambda row: get_dbh_cm(row["DBH (cm)"], row["Girth (cm)"]), axis=1)
    working["Plot"] = working["Plot"].astype(str).str.strip()
    working = working[working["Species"].astype(str).str.strip().ne("") & working["Plot"].ne("") & working["DBH_cm"].notna()].copy()
    if working.empty:
        return pd.DataFrame(columns=detail_columns), pd.DataFrame(columns=summary_columns)

    detail_frames: list[pd.DataFrame] = []
    summary_rows: list[dict[str, object]] = []
    for sheet_name, group in working.groupby("sheet_name", dropna=False):
        total_plots = group["Plot"].nunique()
        if total_plots == 0:
            LOG.warning("Sheet '%s' has no valid plot values for IVI calculation.", sheet_name)
            continue

        species_summary = (
            group.groupby("Species", sort=True)
            .agg(
                **{
                    "Number of tree": ("Species", "size"),
                    "Plot": ("Plot", "nunique"),
                    "DBH_sum_cm": ("DBH_cm", "sum"),
                }
            )
            .reset_index()
        )
        species_summary["sheet_name"] = sheet_name
        species_summary["Density (tree/ha)"] = species_summary["Number of tree"] / (species_summary["Plot"] * plot_area_ha)
        species_summary["Density (tree/rai)"] = species_summary["Density (tree/ha)"] / rai_per_hectare
        species_summary["Frequency"] = species_summary["Plot"] / species_summary["Plot"].sum() * 100
        species_summary["DBH (m)"] = species_summary["DBH_sum_cm"] / 100
        species_summary["BA (m2)"] = ((species_summary["DBH (m)"] ** 2) / 4) * PI
        species_summary["Dominance"] = species_summary["BA (m2)"] / total_plots
        species_summary["RDensity"] = species_summary["Density (tree/rai)"] / species_summary["Density (tree/rai)"].sum() * 100
        species_summary["RFrequency"] = species_summary["Frequency"] / species_summary["Frequency"].sum() * 100
        species_summary["RDominance"] = species_summary["Dominance"] / species_summary["Dominance"].sum() * 100
        species_summary["IVI"] = (
            species_summary["RDensity"] + species_summary["RFrequency"] + species_summary["RDominance"]
        ) / 3
        total_trees = species_summary["Number of tree"].sum()
        species_summary["Pi"] = species_summary["Number of tree"] / total_trees
        species_summary["ln Pi"] = species_summary["Pi"].apply(math.log)
        species_summary["Pi (ln Pi)"] = species_summary["Pi"] * species_summary["ln Pi"]
        species_summary["Shannon contribution"] = -species_summary["Pi (ln Pi)"]

        detail_frames.append(species_summary[detail_columns].copy())
        summary_rows.append(
            {
                "sheet_name": sheet_name,
                "Shannon_index": species_summary["Shannon contribution"].sum(),
                "n_species": int(species_summary["Species"].nunique()),
                "n_tree": int(total_trees),
            }
        )

    detail = pd.concat(detail_frames, ignore_index=True) if detail_frames else pd.DataFrame(columns=detail_columns)
    summary = pd.DataFrame(summary_rows, columns=summary_columns)
    return detail, summary


def build_seedling_outputs(seedling_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    detail_columns = ["sheet_name", "row_no", "No.", "Species", "Number", "Plot"]
    summary_columns = ["sheet_name", "Species", "total_number", "n_plots"]
    if seedling_df.empty:
        return pd.DataFrame(columns=detail_columns), pd.DataFrame(columns=summary_columns)
    detail = seedling_df[detail_columns].copy()
    summary = (
        detail.groupby(["sheet_name", "Species"], dropna=False)
        .agg(total_number=("Number", "sum"), n_plots=("Plot", "nunique"))
        .reset_index()
    )
    return detail, summary


def build_bamboo_outputs(bamboo_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    detail_columns = ["sheet_name", "row_no", "No.", "Species", "Culm"]
    summary_columns = ["sheet_name", "Species", "total_culm"]
    if bamboo_df.empty:
        return pd.DataFrame(columns=detail_columns), pd.DataFrame(columns=summary_columns)
    detail = bamboo_df[detail_columns].copy()
    summary = detail.groupby(["sheet_name", "Species"], dropna=False).agg(total_culm=("Culm", "sum")).reset_index()
    return detail, summary


def get_tree_detail_for_site(source_sheet_name: str, sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    detail = sheets["DETAIL_VOLUME"]
    if detail.empty:
        return pd.DataFrame(columns=detail.columns)
    return detail[(detail["sheet_name"] == source_sheet_name) & (detail["block_type"] == "Tree")].copy()


def build_ivi_summary(source_sheet_name: str, sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    detail_ivi = sheets["DETAIL_IVI"]
    if detail_ivi.empty:
        return pd.DataFrame(
            columns=[
                "Species",
                "Number of tree",
                "Density (tree/ha)",
                "Density (tree/rai)",
                "Plot",
                "Frequency",
                "Dominance",
                "RDensity",
                "RFrequency",
                "RDominance",
                "IVI",
                "Shannon contribution",
            ]
        )
    filtered = detail_ivi[detail_ivi["sheet_name"] == source_sheet_name].copy()
    if filtered.empty:
        return pd.DataFrame(
            columns=[
                "Species",
                "Number of tree",
                "Density (tree/ha)",
                "Density (tree/rai)",
                "Plot",
                "Frequency",
                "Dominance",
                "RDensity",
                "RFrequency",
                "RDominance",
                "IVI",
                "Shannon contribution",
            ]
        )
    filtered = filtered.sort_values("IVI", ascending=False)
    return filtered[
        [
            "Species",
            "Number of tree",
            "Density (tree/ha)",
            "Density (tree/rai)",
            "Plot",
            "Frequency",
            "Dominance",
            "RDensity",
            "RFrequency",
            "RDominance",
            "IVI",
            "Shannon contribution",
        ]
    ].reset_index(drop=True)


def build_general_tree_stand_summary(
    source_sheet_name: str,
    sheets: dict[str, pd.DataFrame],
    plot_area_ha: float,
    rai_per_hectare: float,
) -> pd.DataFrame:
    tree_detail = get_tree_detail_for_site(source_sheet_name, sheets)
    columns = ["Item", "Plot", "DBH (cm)", "Girth (cm)", "Height (m)", "TQ", "Log", "Volume (m3)"]
    if tree_detail.empty:
        return pd.DataFrame(columns=columns)

    n_tree = len(tree_detail)
    n_plots = tree_detail["Plot"].astype(str).str.strip().replace("", np.nan).dropna().nunique()
    total_area_ha = n_plots * plot_area_ha
    total_area_rai = total_area_ha * rai_per_hectare if pd.notna(total_area_ha) else np.nan
    species_series = tree_detail["Species_norm"].where(
        tree_detail["Species_norm"].notna() & tree_detail["Species_norm"].astype(str).str.strip().ne(""),
        tree_detail["Species_raw"],
    )
    n_species = species_series.astype(str).str.strip().replace("", np.nan).dropna().nunique()

    dbh_sum = tree_detail["DBH_cm"].sum()
    girth_sum = tree_detail["Girth_cm"].sum()
    height_sum = tree_detail["Height_m"].sum()
    volume_sum = tree_detail["volume_m3"].sum()

    rows = [
        {
            "Item": "Total",
            "Plot": "-",
            "DBH (cm)": display_or_dash(dbh_sum),
            "Girth (cm)": display_or_dash(girth_sum),
            "Height (m)": display_or_dash(height_sum),
            "TQ": "-",
            "Log": "-",
            "Volume (m3)": display_or_dash(volume_sum),
        },
        {
            "Item": "Average per tree",
            "Plot": "-",
            "DBH (cm)": display_or_dash(tree_detail["DBH_cm"].mean()),
            "Girth (cm)": display_or_dash(tree_detail["Girth_cm"].mean()),
            "Height (m)": display_or_dash(tree_detail["Height_m"].mean()),
            "TQ": "-",
            "Log": "-",
            "Volume (m3)": display_or_dash(tree_detail["volume_m3"].mean()),
        },
        {
            "Item": "Tree count",
            "Plot": "-",
            "DBH (cm)": "-",
            "Girth (cm)": "-",
            "Height (m)": "-",
            "TQ": "-",
            "Log": n_tree,
            "Volume (m3)": "-",
        },
        {
            "Item": "Average per plot",
            "Plot": "-",
            "DBH (cm)": safe_divide(n_tree, n_plots),
            "Girth (cm)": "-",
            "Height (m)": "-",
            "TQ": "-",
            "Log": "-",
            "Volume (m3)": safe_divide(volume_sum, n_plots),
        },
        {
            "Item": "Trees per hectare",
            "Plot": "-",
            "DBH (cm)": safe_divide(n_tree, total_area_ha),
            "Girth (cm)": "-",
            "Height (m)": "-",
            "TQ": "-",
            "Log": "-",
            "Volume (m3)": "-",
        },
        {
            "Item": "Volume per hectare",
            "Plot": "-",
            "DBH (cm)": "-",
            "Girth (cm)": "-",
            "Height (m)": "-",
            "TQ": "-",
            "Log": "-",
            "Volume (m3)": safe_divide(volume_sum, total_area_ha),
        },
        {
            "Item": "Species count",
            "Plot": "-",
            "DBH (cm)": n_species,
            "Girth (cm)": "-",
            "Height (m)": "-",
            "TQ": "-",
            "Log": "-",
            "Volume (m3)": "-",
        },
        {
            "Item": "Trees per rai",
            "Plot": "-",
            "DBH (cm)": "-",
            "Girth (cm)": "-",
            "Height (m)": "-",
            "TQ": "-",
            "Log": "-",
            "Volume (m3)": safe_divide(n_tree, total_area_rai),
        },
    ]
    return pd.DataFrame(rows, columns=columns)


def build_dbh_class_summary(
    source_sheet_name: str,
    sheets: dict[str, pd.DataFrame],
    plot_area_ha: float,
    rai_per_hectare: float,
) -> pd.DataFrame:
    tree_detail = get_tree_detail_for_site(source_sheet_name, sheets)
    columns = ["DBH Class", "Total", "Average per plot", "Density per hectare", "Trees per rai"]
    if tree_detail.empty:
        return pd.DataFrame(columns=columns)

    dbh = tree_detail["DBH_cm"]
    n_plots = tree_detail["Plot"].astype(str).str.strip().replace("", np.nan).dropna().nunique()
    total_area_ha = n_plots * plot_area_ha
    total_area_rai = total_area_ha * rai_per_hectare if pd.notna(total_area_ha) else np.nan

    class_specs = [
        ("dbh 10-30", (dbh >= 10) & (dbh < 30)),
        ("dbh 30-60", (dbh >= 30) & (dbh <= 60)),
        ("dbh > 60", dbh > 60),
    ]
    rows: list[dict[str, object]] = []
    total_count = 0
    total_avg_plot = 0.0
    total_density_ha = 0.0
    total_per_rai = 0.0
    for label, mask in class_specs:
        n_class = int(mask.fillna(False).sum())
        avg_plot = safe_divide(n_class, n_plots)
        density_ha = safe_divide(n_class, total_area_ha)
        per_rai = safe_divide(n_class, total_area_rai)
        rows.append(
            {
                "DBH Class": label,
                "Total": n_class,
                "Average per plot": avg_plot,
                "Density per hectare": density_ha,
                "Trees per rai": per_rai,
            }
        )
        total_count += n_class
        total_avg_plot += avg_plot if isinstance(avg_plot, (int, float)) else 0.0
        total_density_ha += density_ha if isinstance(density_ha, (int, float)) else 0.0
        total_per_rai += per_rai if isinstance(per_rai, (int, float)) else 0.0

    rows.append(
        {
            "DBH Class": "Total",
            "Total": total_count,
            "Average per plot": total_avg_plot if n_plots else "-",
            "Density per hectare": total_density_ha if total_area_ha else "-",
            "Trees per rai": total_per_rai if total_area_rai else "-",
        }
    )
    return pd.DataFrame(rows, columns=columns)


def build_tq_volume_summary(
    source_sheet_name: str,
    sheets: dict[str, pd.DataFrame],
    plot_area_ha: float,
    rai_per_hectare: float,
) -> pd.DataFrame:
    tree_detail = get_tree_detail_for_site(source_sheet_name, sheets)
    columns = ["Timber Quality Class (TQ)", "Volume", "Per plot", "Per hectare", "Per rai"]
    if tree_detail.empty:
        return pd.DataFrame(columns=columns)

    n_plots = tree_detail["Plot"].astype(str).str.strip().replace("", np.nan).dropna().nunique()
    total_area_ha = n_plots * plot_area_ha
    total_area_rai = total_area_ha * rai_per_hectare if pd.notna(total_area_ha) else np.nan

    working = tree_detail.copy()
    working["TQ_display"] = working["TQ"].apply(format_tq_label)
    grouped = (
        working.groupby("TQ_display", dropna=False)["volume_m3"]
        .sum()
        .reset_index()
        .rename(columns={"TQ_display": "Timber Quality Class (TQ)", "volume_m3": "Volume"})
    )
    grouped["Per plot"] = grouped["Volume"].apply(lambda x: safe_divide(x, n_plots))
    grouped["Per hectare"] = grouped["Volume"].apply(lambda x: safe_divide(x, total_area_ha))
    grouped["Per rai"] = grouped["Volume"].apply(lambda x: safe_divide(x, total_area_rai))
    grouped = grouped.sort_values(
        by="Timber Quality Class (TQ)",
        key=lambda s: s.map(tq_sort_key),
    ).reset_index(drop=True)

    total_volume = grouped["Volume"].sum()
    total_row = pd.DataFrame(
        [
            {
                "Timber Quality Class (TQ)": "Total",
                "Volume": total_volume,
                "Per plot": safe_divide(total_volume, n_plots),
                "Per hectare": safe_divide(total_volume, total_area_ha),
                "Per rai": safe_divide(total_volume, total_area_rai),
            }
        ]
    )
    return pd.concat([grouped, total_row], ignore_index=True)[columns]


def build_summary_all(
    tree_df: pd.DataFrame,
    sapling_df: pd.DataFrame,
    seedling_df: pd.DataFrame,
    bamboo_df: pd.DataFrame,
    biomass_summary: pd.DataFrame,
    volume_summary: pd.DataFrame,
    shannon_summary: pd.DataFrame,
    unmatched_df: pd.DataFrame,
) -> pd.DataFrame:
    sheet_names = sorted(
        {
            *tree_df.get("sheet_name", pd.Series(dtype=object)).dropna().astype(str).tolist(),
            *sapling_df.get("sheet_name", pd.Series(dtype=object)).dropna().astype(str).tolist(),
            *seedling_df.get("sheet_name", pd.Series(dtype=object)).dropna().astype(str).tolist(),
            *bamboo_df.get("sheet_name", pd.Series(dtype=object)).dropna().astype(str).tolist(),
        }
    )

    rows: list[dict[str, object]] = []
    for sheet_name in sheet_names:
        tree_volume = volume_summary[
            (volume_summary["sheet_name"] == sheet_name) & (volume_summary["block_type"] == "Tree")
        ]
        sapling_volume = volume_summary[
            (volume_summary["sheet_name"] == sheet_name) & (volume_summary["block_type"] == "Sapling")
        ]
        shannon_row = shannon_summary[shannon_summary["sheet_name"] == sheet_name]
        biomass_rows = biomass_summary[biomass_summary["sheet_name"] == sheet_name]

        rows.append(
            {
                "sheet_name": sheet_name,
                "n_tree": int(tree_volume["n_records"].sum()) if not tree_volume.empty else 0,
                "n_sapling": int(sapling_volume["n_records"].sum()) if not sapling_volume.empty else 0,
                "n_seedling_records": int((seedling_df["sheet_name"] == sheet_name).sum()) if not seedling_df.empty else 0,
                "n_bamboo_records": int((bamboo_df["sheet_name"] == sheet_name).sum()) if not bamboo_df.empty else 0,
                "total_tree_biomass": biomass_rows["biomass_total_sum"].sum() if not biomass_rows.empty else 0.0,
                "total_tree_volume_m3": tree_volume["total_volume_m3"].sum() if not tree_volume.empty else 0.0,
                "total_sapling_volume_m3": sapling_volume["total_volume_m3"].sum() if not sapling_volume.empty else 0.0,
                "total_seedling_number": seedling_df.loc[seedling_df["sheet_name"] == sheet_name, "Number"].sum()
                if not seedling_df.empty
                else 0.0,
                "total_bamboo_culm": bamboo_df.loc[bamboo_df["sheet_name"] == sheet_name, "Culm"].sum()
                if not bamboo_df.empty
                else 0.0,
                "shannon_index": shannon_row["Shannon_index"].iloc[0] if not shannon_row.empty else np.nan,
                "n_unmatched_tree_species": int(
                    ((unmatched_df["sheet_name"] == sheet_name) & (unmatched_df["block_type"] == "Tree")).sum()
                )
                if not unmatched_df.empty
                else 0,
                "n_unmatched_sapling_species": int(
                    ((unmatched_df["sheet_name"] == sheet_name) & (unmatched_df["block_type"] == "Sapling")).sum()
                )
                if not unmatched_df.empty
                else 0,
            }
        )

    return pd.DataFrame(rows)


def autofit_worksheet_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        col_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 40)


def apply_number_formats(worksheet) -> None:
    biomass_volume_columns = {
        "DBH_cm",
        "Girth_cm",
        "Height_m",
        "Ws",
        "Wb",
        "Wl",
        "Wr",
        "biomass_total",
        "Ws_sum",
        "Wb_sum",
        "Wl_sum",
        "Wr_sum",
        "biomass_total_sum",
        "volume_m3",
        "total_volume_m3",
        "total_tree_biomass",
        "total_tree_volume_m3",
        "total_sapling_volume_m3",
    }
    ivi_columns = {
        "Density (tree/ha)",
        "Density (tree/rai)",
        "Frequency",
        "DBH (m)",
        "BA (m2)",
        "Dominance",
        "RDensity",
        "RFrequency",
        "RDominance",
        "IVI",
        "Pi",
        "ln Pi",
        "Pi (ln Pi)",
        "Shannon contribution",
        "Shannon_index",
        "shannon_index",
    }
    header_to_column = {cell.column: normalize_text(cell.value) for cell in worksheet[1]}
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            header = header_to_column.get(cell.column, "")
            if header in biomass_volume_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.0000"
            elif header in ivi_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "0.000000"
            elif (
                isinstance(cell.value, (int, float))
                and float(cell.value).is_integer()
                and header
                not in biomass_volume_columns
                and header
                not in ivi_columns
            ):
                cell.number_format = "#,##0"


def format_detail_worksheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    apply_number_formats(worksheet)
    autofit_worksheet_columns(worksheet)


def apply_summary_cell_number_format(cell, header: str) -> None:
    if not isinstance(cell.value, (int, float)):
        return
    volume_headers = {"Volume (m3)", "Volume", "Per plot", "Per hectare", "Per rai", "total_volume_m3"}
    relative_headers = {
        "Frequency",
        "Dominance",
        "RDensity",
        "RFrequency",
        "RDominance",
        "IVI",
        "Shannon contribution",
        "Shannon_index",
        "shannon_index",
        "Pi",
        "ln Pi",
        "Pi (ln Pi)",
    }
    two_decimal_headers = {
        "DBH (cm)",
        "Girth (cm)",
        "Height (m)",
        "Density (tree/ha)",
        "Density (tree/rai)",
        "Tree count",
        "Total",
        "Average per plot",
        "Density per hectare",
        "Trees per rai",
        "n_tree",
        "n_sapling",
        "n_seedling_records",
        "n_bamboo_records",
        "Log",
        "DBH_cm",
        "Girth_cm",
        "Height_m",
        "Ws_sum",
        "Wb_sum",
        "Wl_sum",
        "Wr_sum",
        "biomass_total_sum",
        "total_tree_biomass",
        "total_tree_volume_m3",
        "total_sapling_volume_m3",
        "total_seedling_number",
        "total_bamboo_culm",
    }
    if header in volume_headers:
        cell.number_format = "#,##0.000"
    elif header in relative_headers:
        cell.number_format = "0.000000"
    else:
        cell.number_format = "#,##0.00" if header in two_decimal_headers else "#,##0.00"


def format_summary_site_worksheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    header_fill = "D9EAF7"
    section_titles = [title for title, _ in SUMMARY_SECTION_SPECS]
    row_idx = 1
    while row_idx <= worksheet.max_row:
        first_value = worksheet.cell(row_idx, 1).value
        if first_value in section_titles:
            title_cell = worksheet.cell(row_idx, 1)
            title_cell.font = Font(bold=True, size=13)
            title_cell.fill = PatternFill(
                fill_type="solid",
                fgColor=SUMMARY_SECTION_COLORS.get(str(first_value), "E2F0D9"),
            )

            header_row = row_idx + 1
            header_values = []
            if header_row <= worksheet.max_row:
                for cell in worksheet[header_row]:
                    if cell.value not in (None, ""):
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(fill_type="solid", fgColor=header_fill)
                    header_values.append(normalize_text(cell.value))

                data_row = header_row + 1
                while data_row <= worksheet.max_row:
                    row_values = [worksheet.cell(data_row, c).value for c in range(1, worksheet.max_column + 1)]
                    if all(value in (None, "") for value in row_values):
                        break
                    for col_idx, header in enumerate(header_values, start=1):
                        apply_summary_cell_number_format(worksheet.cell(data_row, col_idx), header)
                    data_row += 1
                row_idx = data_row
                continue
        row_idx += 1

    autofit_worksheet_columns(worksheet)


def ensure_output_frames(frames: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    normalized: dict[str, pd.DataFrame] = {}
    for sheet_name in OUTPUT_SHEETS:
        normalized[sheet_name] = frames.get(sheet_name, pd.DataFrame())
    return normalized


def process_workbook(
    input_file: Path,
    master_file: Path,
    plot_area_ha: float,
    rai_per_hectare: float,
    sheet_groups: list[dict[str, object]] | None = None,
) -> dict[str, pd.DataFrame]:
    ref_map = load_master_reference(master_file)
    LOG.info("Loaded %s species reference entries.", f"{len(ref_map):,}")

    workbook_headers = pd.read_excel(input_file, sheet_name=None, header=None, nrows=2)
    available_sheet_names = [
        sheet_name
        for sheet_name, header_df in workbook_headers.items()
        if not should_skip_sheet(sheet_name) and header_df.shape[0] >= 2
    ]
    normalized_sheet_groups = normalize_sheet_groups(sheet_groups, available_sheet_names)

    tree_frames: list[pd.DataFrame] = []
    sapling_frames: list[pd.DataFrame] = []
    seedling_frames: list[pd.DataFrame] = []
    bamboo_frames: list[pd.DataFrame] = []

    for sheet_name, header_df in workbook_headers.items():
        if should_skip_sheet(sheet_name):
            LOG.info("Skipping sheet '%s'.", sheet_name)
            continue

        if header_df.shape[0] < 2:
            LOG.warning("Sheet '%s' does not have the expected header rows. Skipping.", sheet_name)
            continue

        row_headers = header_df.iloc[1].tolist()
        try:
            sheet_df = read_sheet_frame(input_file, sheet_name)
        except Exception as exc:
            LOG.warning("Could not read sheet '%s': %s", sheet_name, exc)
            continue

        if validate_block_headers(sheet_name, row_headers, "Tree", TREE_COLUMNS):
            tree_frame = prepare_block_frame(
                sheet_df,
                sheet_name,
                "Tree",
                TREE_COLUMNS,
                ["DBH (cm)", "Girth (cm)", "Height (m)", "TQ"],
            )
            if not tree_frame.empty:
                tree_frames.append(tree_frame)
            else:
                LOG.warning("Sheet '%s' has no usable Tree rows.", sheet_name)

        if validate_block_headers(sheet_name, row_headers, "Sapling", SAPLING_COLUMNS):
            sapling_frame = prepare_block_frame(
                sheet_df,
                sheet_name,
                "Sapling",
                SAPLING_COLUMNS,
                ["Girth (cm)", "Height (m)", "Number"],
            )
            if not sapling_frame.empty:
                sapling_frames.append(sapling_frame)
            else:
                LOG.warning("Sheet '%s' has no usable Sapling rows.", sheet_name)

        if validate_block_headers(sheet_name, row_headers, "Seedling", SEEDLING_COLUMNS):
            seedling_frame = prepare_block_frame(
                sheet_df,
                sheet_name,
                "Seedling",
                SEEDLING_COLUMNS,
                ["Number"],
            )
            if not seedling_frame.empty:
                seedling_frames.append(seedling_frame)
            else:
                LOG.warning("Sheet '%s' has no usable Seedling rows.", sheet_name)

        if validate_block_headers(sheet_name, row_headers, "Bamboo", BAMBOO_COLUMNS):
            bamboo_frame = prepare_block_frame(
                sheet_df,
                sheet_name,
                "Bamboo",
                BAMBOO_COLUMNS,
                ["Culm"],
            )
            if not bamboo_frame.empty:
                bamboo_frames.append(bamboo_frame)
            else:
                LOG.warning("Sheet '%s' has no usable Bamboo rows.", sheet_name)

    tree_df = pd.concat(tree_frames, ignore_index=True) if tree_frames else pd.DataFrame(columns=["sheet_name"])
    sapling_df = pd.concat(sapling_frames, ignore_index=True) if sapling_frames else pd.DataFrame(columns=["sheet_name"])
    seedling_df = pd.concat(seedling_frames, ignore_index=True) if seedling_frames else pd.DataFrame(columns=["sheet_name"])
    bamboo_df = pd.concat(bamboo_frames, ignore_index=True) if bamboo_frames else pd.DataFrame(columns=["sheet_name"])

    tree_df = append_grouped_records(tree_df, normalized_sheet_groups)
    sapling_df = append_grouped_records(sapling_df, normalized_sheet_groups)
    seedling_df = append_grouped_records(seedling_df, normalized_sheet_groups)
    bamboo_df = append_grouped_records(bamboo_df, normalized_sheet_groups)

    biomass_detail, biomass_summary = build_biomass_outputs(tree_df)
    volume_detail, volume_summary, unmatched_species = build_volume_outputs(tree_df, sapling_df, ref_map)
    ivi_detail, shannon_summary = build_ivi_outputs(tree_df, plot_area_ha, rai_per_hectare)
    seedling_detail, seedling_summary = build_seedling_outputs(seedling_df)
    bamboo_detail, bamboo_summary = build_bamboo_outputs(bamboo_df)
    summary_all = build_summary_all(
        tree_df,
        sapling_df,
        seedling_df,
        bamboo_df,
        biomass_summary,
        volume_summary,
        shannon_summary,
        unmatched_species,
    )

    outputs = ensure_output_frames(
        {
            "DETAIL_TREE_BIOMASS": biomass_detail,
            "SUMMARY_BIOMASS": biomass_summary,
            "DETAIL_VOLUME": volume_detail,
            "SUMMARY_VOLUME": volume_summary,
            "CHECK_UNMATCHED_SPECIES": unmatched_species,
            "DETAIL_IVI": ivi_detail,
            "SUMMARY_SHANNON": shannon_summary,
            "DETAIL_SEEDLING": seedling_detail,
            "SUMMARY_SEEDLING": seedling_summary,
            "DETAIL_BAMBOO": bamboo_detail,
            "SUMMARY_BAMBOO": bamboo_summary,
            "SUMMARY_ALL": summary_all,
        }
    )
    outputs["__meta__"] = {
        "plot_area_ha": plot_area_ha,
        "rai_per_hectare": rai_per_hectare,
        "sheet_groups": normalized_sheet_groups,
    }
    return outputs


def write_section_table(worksheet, start_row: int, title: str, frame: pd.DataFrame) -> int:
    worksheet.cell(start_row, 1).value = title
    next_row = start_row + 1

    if frame.empty:
        worksheet.cell(next_row, 1).value = "No data."
        return next_row + 2

    for col_idx, column_name in enumerate(frame.columns, start=1):
        worksheet.cell(next_row, col_idx).value = column_name

    data_row = next_row + 1
    for row_values in frame.itertuples(index=False, name=None):
        for col_idx, value in enumerate(row_values, start=1):
            worksheet.cell(data_row, col_idx).value = value
        data_row += 1
    return data_row + 1


def write_summary_by_site_workbook(summary_file: Path, sheets: dict[str, pd.DataFrame]) -> None:
    source_names = sheets["SUMMARY_ALL"].get("sheet_name", pd.Series(dtype=object)).dropna().astype(str).tolist()
    component_names = get_component_sheet_names(sheets)
    unique_source_names = list(dict.fromkeys(source_names))
    source_names = [name for name in unique_source_names if name not in component_names] + [
        name for name in unique_source_names if name in component_names
    ]

    with pd.ExcelWriter(summary_file, engine="openpyxl") as writer:
        if not source_names:
            pd.DataFrame({"message": ["No summary data available."]}).to_excel(
                writer, sheet_name="SUMMARY", index=False
            )
        else:
            used_names: set[str] = set()
            for source_sheet_name in source_names:
                safe_name = safe_sheet_name(source_sheet_name, used_names)
                worksheet = writer.book.create_sheet(title=safe_name)
                writer.sheets[safe_name] = worksheet

                current_row = 1
                for section_title, frame_name in SUMMARY_SECTION_SPECS:
                    if section_title == "IVI Summary":
                        filtered = build_ivi_summary(source_sheet_name, sheets)
                    elif section_title == "General Tree Stand Summary":
                        filtered = build_general_tree_stand_summary(
                            source_sheet_name,
                            sheets,
                            plot_area_ha=sheets["__meta__"]["plot_area_ha"],
                            rai_per_hectare=sheets["__meta__"]["rai_per_hectare"],
                        )
                    elif section_title == "DBH Class Summary":
                        filtered = build_dbh_class_summary(
                            source_sheet_name,
                            sheets,
                            plot_area_ha=sheets["__meta__"]["plot_area_ha"],
                            rai_per_hectare=sheets["__meta__"]["rai_per_hectare"],
                        )
                    elif section_title == "TQ Volume Summary":
                        filtered = build_tq_volume_summary(
                            source_sheet_name,
                            sheets,
                            plot_area_ha=sheets["__meta__"]["plot_area_ha"],
                            rai_per_hectare=sheets["__meta__"]["rai_per_hectare"],
                        )
                    else:
                        frame = sheets[frame_name]
                        filtered = frame[frame["sheet_name"] == source_sheet_name].copy() if "sheet_name" in frame.columns else frame.copy()
                    current_row = write_section_table(worksheet, current_row, section_title, filtered)

                format_summary_site_worksheet(worksheet)

            if "Sheet" in writer.book.sheetnames and len(writer.book.sheetnames) > 1:
                del writer.book["Sheet"]


def write_detail_workbook(detail_file: Path, sheets: dict[str, pd.DataFrame]) -> None:
    component_names = get_component_sheet_names(sheets)
    with pd.ExcelWriter(detail_file, engine="openpyxl") as writer:
        for sheet_name in DETAIL_WORKBOOK_SHEETS:
            frame = filter_out_component_rows(sheets[sheet_name], component_names)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = load_workbook(detail_file)
    for sheet_name in DETAIL_WORKBOOK_SHEETS:
        format_detail_worksheet(workbook[sheet_name])
    workbook.save(detail_file)


def write_combined_output_workbook(output_file: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet_name in OUTPUT_SHEETS:
            sheets[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = load_workbook(output_file)
    for sheet_name in OUTPUT_SHEETS:
        worksheet = workbook[sheet_name]
        format_detail_worksheet(worksheet)
    workbook.save(output_file)


def resolve_output_paths(input_file: Path, output_arg: str | None) -> tuple[Path, Path]:
    if output_arg:
        base_path = Path(output_arg)
    else:
        base_path = input_file.with_name(input_file.stem)

    parent = base_path.parent
    stem = base_path.stem if output_arg else base_path.name
    summary_file = parent / f"{stem}_summary_by_site.xlsx"
    detail_file = parent / f"{stem}_details.xlsx"
    return summary_file, detail_file


def run_calculation(
    input_file: str | Path,
    master_file: str | Path,
    output_file: str | Path | None = None,
    plot_area_ha: float = PLOT_AREA_HA,
    rai_per_hectare: float = RAI_PER_HECTARE,
    sheet_groups: list[dict[str, object]] | None = None,
) -> tuple[Path | None, dict[str, pd.DataFrame]]:
    input_path = Path(input_file)
    master_path = Path(master_file)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    if not master_path.exists():
        raise FileNotFoundError(f"Master file not found: {master_path}")

    output_sheets = process_workbook(
        input_file=input_path,
        master_file=master_path,
        plot_area_ha=plot_area_ha,
        rai_per_hectare=rai_per_hectare,
        sheet_groups=sheet_groups,
    )

    output_path = Path(output_file) if output_file else None
    if output_path is not None:
        write_combined_output_workbook(output_path, output_sheets)

    return output_path, output_sheets


def run_calculation_split_outputs(
    input_file: str | Path,
    master_file: str | Path,
    output_base: str | Path | None = None,
    plot_area_ha: float = PLOT_AREA_HA,
    rai_per_hectare: float = RAI_PER_HECTARE,
    sheet_groups: list[dict[str, object]] | None = None,
) -> tuple[Path, Path, dict[str, pd.DataFrame]]:
    input_path = Path(input_file)
    master_path = Path(master_file)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    if not master_path.exists():
        raise FileNotFoundError(f"Master file not found: {master_path}")

    output_sheets = process_workbook(
        input_file=input_path,
        master_file=master_path,
        plot_area_ha=plot_area_ha,
        rai_per_hectare=rai_per_hectare,
        sheet_groups=sheet_groups,
    )

    summary_file, detail_file = resolve_output_paths(input_path, str(output_base) if output_base else None)
    write_summary_by_site_workbook(summary_file, output_sheets)
    write_detail_workbook(detail_file, output_sheets)
    return summary_file, detail_file, output_sheets


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")
    args = parse_args()

    input_file = Path(args.input_file)
    master_file = Path(args.master)
    summary_file, detail_file = resolve_output_paths(input_file, args.output)

    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")
    if not master_file.exists():
        raise FileNotFoundError(f"Master file not found: {master_file}")

    output_sheets = process_workbook(
        input_file=input_file,
        master_file=master_file,
        plot_area_ha=args.plot_area_ha,
        rai_per_hectare=args.rai_per_hectare,
    )
    write_summary_by_site_workbook(summary_file, output_sheets)
    write_detail_workbook(detail_file, output_sheets)

    LOG.info("Summary workbook written to: %s", summary_file)
    LOG.info("Detail workbook written to: %s", detail_file)


if __name__ == "__main__":
    main()
