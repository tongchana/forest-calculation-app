from __future__ import annotations

import tempfile
import traceback
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

try:
    from streamlit_sortables import sort_items
except ImportError:  # pragma: no cover - optional UI enhancement
    sort_items = None

import run_forest_calculation as calc


APP_TITLE = "Forest Biomass, Volume, IVI and Shannon Calculator"
APP_SUBTITLE = (
    "Upload field survey data using the provided Excel template and generate calculated results automatically."
)
PLOT_AREA_HA = calc.PLOT_AREA_HA
RAI_PER_HECTARE = calc.RAI_PER_HECTARE
TEMPLATE_FILE = Path("template.xlsx")
MASTER_FILE = Path("species_reference_master_v1.xlsx")
COMPONENT_TEMPLATE_FILE = Path("summary_component.xlsx")
OUTPUT_BASE_FILENAME = "forest_calculation_output.xlsx"
SUMMARY_OUTPUT_FILENAME = "forest_calculation_output_summary_by_site.xlsx"
DETAIL_OUTPUT_FILENAME = "forest_calculation_output_details.xlsx"
COMPONENT_OUTPUT_FILENAME = "forest_component_summary.xlsx"
PREVIEW_SHEETS = [
    "SUMMARY_ALL",
    "SUMMARY_BIOMASS",
    "SUMMARY_VOLUME",
    "SUMMARY_SHANNON",
    "CHECK_UNMATCHED_SPECIES",
]
DEFAULT_GROUP_LABEL = "Component"
MAX_COMPONENTS = 5
SORTABLE_STYLE = """
.sortable-component {
    border: 1px solid rgba(36, 92, 63, 0.12);
    border-radius: 18px;
    padding: 0.75rem;
    background: #f8fcf9;
}
.sortable-container {
    background: white;
    border-radius: 14px;
    border: 1px solid rgba(36, 92, 63, 0.10);
    min-height: 240px;
}
.sortable-container-header {
    background: #e8f2ec;
    color: #1f6b4f;
    font-weight: 700;
    padding: 0.75rem 0.9rem;
    border-bottom: 1px solid rgba(36, 92, 63, 0.08);
}
.sortable-container-body {
    padding: 0.65rem;
}
.sortable-item, .sortable-item:hover {
    background: linear-gradient(135deg, #f4fbf6 0%, #ebf6ef 100%);
    border: 1px solid rgba(46, 125, 90, 0.16);
    color: #183c2c;
    border-radius: 10px;
    margin-bottom: 0.45rem;
    font-weight: 600;
}
"""


def inject_css() -> None:
    st.markdown(
        """
        <style>
        .stApp {
            background: linear-gradient(180deg, #f4fbf6 0%, #eef6f0 100%);
        }
        .main .block-container {
            max-width: 1180px;
            padding-top: 2rem;
            padding-bottom: 3rem;
        }
        .hero-card {
            background: linear-gradient(135deg, #1f6b4f 0%, #3a8b68 100%);
            color: white;
            border-radius: 22px;
            padding: 1.8rem 2rem;
            box-shadow: 0 18px 40px rgba(31, 107, 79, 0.18);
            margin-bottom: 1rem;
        }
        .hero-subtitle {
            font-size: 1rem;
            opacity: 0.92;
            margin-top: 0.45rem;
        }
        .step-card, .info-card, .metric-card {
            background: #ffffff;
            border-radius: 18px;
            border: 1px solid rgba(36, 92, 63, 0.10);
            padding: 1.1rem 1.15rem;
            box-shadow: 0 12px 28px rgba(31, 76, 54, 0.06);
            margin-bottom: 1rem;
        }
        .step-title {
            color: #1f6b4f;
            font-weight: 700;
            font-size: 1.02rem;
            margin-bottom: 0.35rem;
        }
        .metric-label {
            color: #597164;
            font-size: 0.9rem;
            margin-bottom: 0.25rem;
        }
        .metric-value {
            color: #183c2c;
            font-size: 1.6rem;
            font-weight: 700;
            line-height: 1.2;
        }
        .metric-help {
            color: #70877b;
            font-size: 0.82rem;
            margin-top: 0.2rem;
        }
        .warning-card {
            background: #fff8eb;
            border: 1px solid #f3d08b;
            border-radius: 16px;
            padding: 1rem 1.1rem;
            color: #7b5a18;
            margin-bottom: 1rem;
        }
        .success-card {
            background: #edf8f0;
            border: 1px solid #b8e1c2;
            border-radius: 16px;
            padding: 1rem 1.1rem;
            color: #205c37;
            margin-bottom: 1rem;
        }
        .section-label {
            font-size: 0.84rem;
            text-transform: uppercase;
            letter-spacing: 0.08em;
            color: #688173;
            margin-bottom: 0.35rem;
            font-weight: 700;
        }
        div[data-testid="stDownloadButton"] button,
        div[data-testid="stButton"] button {
            border-radius: 12px;
            border: none;
            padding: 0.65rem 1rem;
            font-weight: 600;
        }
        div[data-testid="stFileUploader"] section {
            border-radius: 16px;
            border: 1px dashed #89b39b;
            background: #f8fcf9;
        }
        div[data-testid="stDataFrame"] {
            border-radius: 14px;
            overflow: hidden;
            border: 1px solid rgba(36, 92, 63, 0.08);
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_card(title: str, body: str) -> None:
    st.markdown(
        f"""
        <div class="step-card">
            <div class="step-title">{title}</div>
            <div>{body}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_metric(label: str, value: object, help_text: str = "") -> None:
    display_value = "-" if value is None or (isinstance(value, float) and pd.isna(value)) else value
    st.markdown(
        f"""
        <div class="metric-card">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{display_value}</div>
            <div class="metric-help">{help_text}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def load_binary_file(file_path: Path) -> bytes:
    return file_path.read_bytes()


def get_uploaded_sheet_names(uploaded_file) -> list[str]:
    workbook = load_workbook(filename=BytesIO(uploaded_file.getvalue()), read_only=True, data_only=True)
    try:
        return [sheet_name for sheet_name in workbook.sheetnames if not calc.should_skip_sheet(sheet_name)]
    finally:
        workbook.close()


def make_group_container(index: int) -> dict[str, list[str] | str]:
    return {"header": f"{DEFAULT_GROUP_LABEL} {index}", "items": []}


def order_items_by_reference(items: list[str], reference_order: list[str]) -> list[str]:
    position = {name: idx for idx, name in enumerate(reference_order)}
    unique_items = list(dict.fromkeys(items))
    return sorted(unique_items, key=lambda name: position.get(name, len(reference_order)))


def ensure_sheet_group_state(sheet_names: list[str]) -> None:
    signature = tuple(sheet_names)
    if st.session_state.get("sheet_group_signature") == signature:
        return

    st.session_state.sheet_group_signature = signature
    st.session_state.sheet_group_containers = [{"header": "Available sheets", "items": sheet_names.copy()}]
    st.session_state.sheet_group_name_count = 0


def add_sheet_group() -> None:
    containers = st.session_state.sheet_group_containers
    if len(containers) - 1 >= MAX_COMPONENTS:
        return
    next_index = len(containers)
    containers.append(make_group_container(next_index))
    st.session_state.sheet_group_name_count = next_index


def remove_sheet_group(reference_order: list[str]) -> None:
    containers = st.session_state.sheet_group_containers
    if len(containers) <= 1:
        return

    removed = containers.pop()
    available_items = containers[0]["items"] + removed["items"]
    containers[0]["items"] = order_items_by_reference(available_items, reference_order)
    st.session_state.sheet_group_name_count = len(containers) - 1


def normalize_sortable_containers(
    containers: list[dict[str, list[str] | str]],
    sheet_names: list[str],
) -> list[dict[str, list[str] | str]]:
    allowed = set(sheet_names)
    assigned: set[str] = set()
    normalized: list[dict[str, list[str] | str]] = []

    for idx, container in enumerate(containers):
        items = [item for item in container.get("items", []) if item in allowed]
        deduped: list[str] = []
        for item in items:
            if item in assigned:
                continue
            assigned.add(item)
            deduped.append(item)
        header = "Available sheets" if idx == 0 else container.get("header", f"{DEFAULT_GROUP_LABEL} {idx}")
        normalized.append({"header": header, "items": deduped})

    unassigned = [sheet_name for sheet_name in sheet_names if sheet_name not in assigned]
    if normalized:
        normalized[0]["items"] = order_items_by_reference(normalized[0]["items"] + unassigned, sheet_names)
    return normalized


def render_sheet_group_builder(sheet_names: list[str]) -> list[dict[str, list[str]]]:
    ensure_sheet_group_state(sheet_names)

    st.markdown('<div class="section-label">Optional grouping</div>', unsafe_allow_html=True)
    render_card(
        "Combine Multiple Sheets Into One Calculation Component",
        "Drag worksheets into custom components to calculate combined IVI, biomass, volume, and Shannon results in addition to the normal per-sheet outputs.",
    )

    action_col1, action_col2 = st.columns([1, 1])
    with action_col1:
        st.button(
            "Add component",
            on_click=add_sheet_group,
            disabled=len(st.session_state.sheet_group_containers) - 1 >= MAX_COMPONENTS,
            use_container_width=True,
        )
    with action_col2:
        st.button(
            "Remove last component",
            on_click=remove_sheet_group,
            args=(sheet_names,),
            disabled=len(st.session_state.sheet_group_containers) <= 1,
            use_container_width=True,
        )

    group_count = len(st.session_state.sheet_group_containers) - 1
    st.caption(f"You can create up to {MAX_COMPONENTS} components.")
    for idx in range(1, group_count + 1):
        group_key = f"sheet_group_name_{idx}"
        if group_key not in st.session_state:
            st.session_state[group_key] = f"IVI {DEFAULT_GROUP_LABEL.lower()} {idx}"
        st.text_input(f"Name for {DEFAULT_GROUP_LABEL.lower()} {idx}", key=group_key)

    groups: list[dict[str, list[str]]] = []

    builder_mode = "Drag and drop"
    if sort_items is not None:
        builder_mode = st.radio(
            "Component builder mode",
            options=["Drag and drop", "Simple selection"],
            horizontal=True,
            help="Drag and drop is the default workflow. Switch to Simple selection if your browser or device has trouble with dragging.",
        )

    if builder_mode == "Simple selection":
        assigned: set[str] = set()
        for idx in range(1, group_count + 1):
            selection_key = f"sheet_group_items_{idx}"
            current_selected = [item for item in st.session_state.get(selection_key, []) if item in sheet_names]
            available_options = [item for item in sheet_names if item not in assigned or item in current_selected]
            st.markdown(
                f"""
                <div class="step-card">
                    <div class="step-title">Component {idx}</div>
                    <div>Select one or more worksheets to combine under this component.</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            selected_items = st.multiselect(
                f"Worksheets in component {idx}",
                options=available_options,
                default=current_selected,
                key=selection_key,
                placeholder="Choose worksheets for this component",
            )
            assigned.update(selected_items)
            group_name = (st.session_state.get(f"sheet_group_name_{idx}") or "").strip()
            if selected_items and not group_name:
                st.warning(f"Please enter a name for {DEFAULT_GROUP_LABEL.lower()} {idx} before calculation.")
            elif selected_items:
                groups.append({"name": group_name, "sheet_names": selected_items})

        remaining_sheets = [item for item in sheet_names if item not in assigned]
        if remaining_sheets:
            st.caption(f"Still ungrouped: {', '.join(remaining_sheets)}")
        else:
            st.caption("All worksheets have been assigned to components.")
    else:
        sortable_containers = sort_items(
            st.session_state.sheet_group_containers,
            multi_containers=True,
            custom_style=SORTABLE_STYLE,
            key="sheet_group_sortable",
        )
        st.session_state.sheet_group_containers = normalize_sortable_containers(sortable_containers, sheet_names)

        for idx, container in enumerate(st.session_state.sheet_group_containers[1:], start=1):
            group_name = (st.session_state.get(f"sheet_group_name_{idx}") or "").strip()
            group_items = container["items"]
            if not group_items:
                continue
            if not group_name:
                st.warning(f"Please enter a name for {DEFAULT_GROUP_LABEL.lower()} {idx} before calculation.")
                continue
            groups.append({"name": group_name, "sheet_names": group_items})

    if groups:
        preview_rows = [{"Component name": group["name"], "Sheets": ", ".join(group["sheet_names"])} for group in groups]
        st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)
        st.caption("Each worksheet can belong to one component in this drag-and-drop board.")
    else:
        st.caption("No combined components defined. The app will calculate each worksheet separately, like before.")

    return groups


def render_sidebar() -> None:
    st.sidebar.markdown("## About this tool")
    st.sidebar.write(
        "This web app wraps the existing forest calculation workflow and helps teams move from field-survey Excel data to ready-to-review calculation outputs."
    )

    st.sidebar.markdown("## Calculation scope")
    scope_df = pd.DataFrame(
        [
            ["Tree", "Yes", "Yes", "Yes", "Yes"],
            ["Sapling", "No", "Yes", "No", "Yes"],
            ["Seedling", "No", "No", "No", "Yes"],
            ["Bamboo", "No", "No", "No", "Yes"],
        ],
        columns=["Block", "Biomass", "Volume", "IVI/Shannon", "Count summary"],
    )
    st.sidebar.table(scope_df)

    st.sidebar.markdown("## Template instructions")
    st.sidebar.write(
        "Use the official Excel template, fill in Tree, Sapling, Seedling, and Bamboo survey data offline, then upload the completed workbook here."
    )

    st.sidebar.markdown("## Important calculation note")
    st.sidebar.warning("Biomass is calculated for Tree only. Sapling is not included in biomass.")


def format_metric_value(value: object, decimals: int = 2) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "-"
    if isinstance(value, (int, float)):
        return f"{value:,.{decimals}f}" if isinstance(value, float) or decimals else f"{int(value):,}"
    return str(value)


def build_metrics(summary_all: pd.DataFrame, unmatched: pd.DataFrame) -> list[tuple[str, str, str]]:
    if summary_all.empty:
        return []

    total_tree = pd.to_numeric(summary_all.get("n_tree"), errors="coerce").fillna(0).sum()
    total_sapling = pd.to_numeric(summary_all.get("n_sapling"), errors="coerce").fillna(0).sum()
    total_tree_biomass = pd.to_numeric(summary_all.get("total_tree_biomass"), errors="coerce").fillna(0).sum()
    total_tree_volume = pd.to_numeric(summary_all.get("total_tree_volume_m3"), errors="coerce").fillna(0).sum()
    total_sapling_volume = pd.to_numeric(summary_all.get("total_sapling_volume_m3"), errors="coerce").fillna(0).sum()

    shannon_series = pd.to_numeric(summary_all.get("shannon_index"), errors="coerce")
    shannon_value = shannon_series.mean() if shannon_series is not None and not shannon_series.dropna().empty else None
    unmatched_count = len(unmatched.index) if not unmatched.empty else 0

    return [
        ("Total tree count", format_metric_value(total_tree, 0), "Across all processed worksheets"),
        ("Total sapling count", format_metric_value(total_sapling, 0), "Sapling records included in volume"),
        ("Total tree biomass", format_metric_value(total_tree_biomass, 2), "Tree biomass only"),
        ("Total tree volume", format_metric_value(total_tree_volume, 3), "Tree block volume"),
        ("Total sapling volume", format_metric_value(total_sapling_volume, 3), "Sapling block volume"),
        ("Shannon index", format_metric_value(shannon_value, 6), "Average across available sites"),
        ("Unmatched species", format_metric_value(unmatched_count, 0), "Species needing reference review"),
    ]


def safe_sheet(frame_dict: dict[str, pd.DataFrame], sheet_name: str) -> pd.DataFrame:
    frame = frame_dict.get(sheet_name)
    return frame if isinstance(frame, pd.DataFrame) else pd.DataFrame()


def preview_results(result_sheets: dict[str, pd.DataFrame]) -> None:
    summary_all = safe_sheet(result_sheets, "SUMMARY_ALL")
    summary_biomass = safe_sheet(result_sheets, "SUMMARY_BIOMASS")
    summary_volume = safe_sheet(result_sheets, "SUMMARY_VOLUME")
    summary_shannon = safe_sheet(result_sheets, "SUMMARY_SHANNON")
    unmatched = safe_sheet(result_sheets, "CHECK_UNMATCHED_SPECIES")

    st.markdown('<div class="section-label">Step 4</div>', unsafe_allow_html=True)
    render_card("View Summary Results", "Preview the main summary sheets before downloading the full calculated workbook.")

    metrics = build_metrics(summary_all, unmatched)
    if metrics:
        for start in range(0, len(metrics), 3):
            cols = st.columns(3)
            for col, metric in zip(cols, metrics[start : start + 3]):
                with col:
                    render_metric(*metric)
    else:
        st.info("No summary metrics are available yet.")

    tabs = st.tabs(["Overall Summary", "Biomass", "Volume", "Shannon / IVI", "Unmatched Species"])
    with tabs[0]:
        if summary_all.empty:
            st.info("SUMMARY_ALL is empty.")
        else:
            st.dataframe(summary_all, use_container_width=True)
    with tabs[1]:
        if summary_biomass.empty:
            st.info("SUMMARY_BIOMASS is empty.")
        else:
            st.dataframe(summary_biomass, use_container_width=True)
    with tabs[2]:
        if summary_volume.empty:
            st.info("SUMMARY_VOLUME is empty.")
        else:
            st.dataframe(summary_volume, use_container_width=True)
    with tabs[3]:
        if summary_shannon.empty:
            st.info("SUMMARY_SHANNON is empty.")
        else:
            st.dataframe(summary_shannon, use_container_width=True)
    with tabs[4]:
        if unmatched.empty:
            st.success("No unmatched species were found in the uploaded workbook.")
        else:
            st.markdown(
                '<div class="warning-card">Some species could not be matched with the master reference file. Please review the table below.</div>',
                unsafe_allow_html=True,
            )
            st.dataframe(unmatched, use_container_width=True)


def run_uploaded_workflow(
    uploaded_file,
    plot_area_ha: float,
    rai_per_hectare: float,
    sheet_groups: list[dict[str, list[str]]] | None = None,
) -> tuple[bytes, bytes, bytes | None, dict[str, pd.DataFrame]]:
    with tempfile.TemporaryDirectory() as tmp_dir:
        temp_dir = Path(tmp_dir)
        uploaded_path = temp_dir / uploaded_file.name
        uploaded_path.write_bytes(uploaded_file.getbuffer())

        output_base = temp_dir / OUTPUT_BASE_FILENAME
        split_runner = getattr(calc, "run_calculation_split_outputs", None)
        if split_runner is not None:
            summary_path, detail_path, result_sheets = split_runner(
                input_file=uploaded_path,
                master_file=MASTER_FILE,
                output_base=output_base,
                plot_area_ha=plot_area_ha,
                rai_per_hectare=rai_per_hectare,
                sheet_groups=sheet_groups,
            )
        else:
            result_sheets = calc.process_workbook(
                input_file=uploaded_path,
                master_file=MASTER_FILE,
                plot_area_ha=plot_area_ha,
                rai_per_hectare=rai_per_hectare,
                sheet_groups=sheet_groups,
            )
            summary_path, detail_path = calc.resolve_output_paths(uploaded_path, str(output_base))
            calc.write_summary_by_site_workbook(summary_path, result_sheets)
            calc.write_detail_workbook(detail_path, result_sheets)
        component_bytes = None
        if sheet_groups and COMPONENT_TEMPLATE_FILE.exists():
            component_path = temp_dir / COMPONENT_OUTPUT_FILENAME
            calc.write_component_summary_workbook(component_path, COMPONENT_TEMPLATE_FILE, result_sheets)
            component_bytes = component_path.read_bytes()
        return summary_path.read_bytes(), detail_path.read_bytes(), component_bytes, result_sheets


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide", page_icon="🌿")
    inject_css()
    render_sidebar()

    if "summary_result_bytes" not in st.session_state:
        st.session_state.summary_result_bytes = None
    if "detail_result_bytes" not in st.session_state:
        st.session_state.detail_result_bytes = None
    if "component_result_bytes" not in st.session_state:
        st.session_state.component_result_bytes = None
    if "result_sheets" not in st.session_state:
        st.session_state.result_sheets = None
    if "last_error" not in st.session_state:
        st.session_state.last_error = None
    if "sheet_group_containers" not in st.session_state:
        st.session_state.sheet_group_containers = [{"header": "Available sheets", "items": []}]
    if "sheet_group_signature" not in st.session_state:
        st.session_state.sheet_group_signature = ()
    if "sheet_group_name_count" not in st.session_state:
        st.session_state.sheet_group_name_count = 0

    st.markdown(
        f"""
        <div class="hero-card">
            <h1 style="margin:0;">{APP_TITLE}</h1>
            <div class="hero-subtitle">{APP_SUBTITLE}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="section-label">How to use this tool</div>', unsafe_allow_html=True)
    render_card(
        "Survey workflow",
        """
        1. Download the official template.<br>
        2. Fill in Tree, Sapling, Seedling, and Bamboo data.<br>
        3. Save the Excel file.<br>
        4. Upload the completed file.<br>
        5. Click Calculate.<br>
        6. Preview the summary and download the summary and detail Excel files.
        """,
    )

    if not TEMPLATE_FILE.exists():
        st.error("The official template file 'template.xlsx' is missing from the project directory.")
    if not MASTER_FILE.exists():
        st.error("The species reference file 'species_reference_master_v1.xlsx' is missing from the project directory.")

    col_download, col_upload = st.columns([1, 1.2], gap="large")
    with col_download:
        st.markdown('<div class="section-label">Step 1</div>', unsafe_allow_html=True)
        render_card("Download Excel Template", "Start from the official workbook to keep the expected layout and headers.")
        if TEMPLATE_FILE.exists():
            st.download_button(
                "Download Excel Template",
                data=load_binary_file(TEMPLATE_FILE),
                file_name=TEMPLATE_FILE.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    with col_upload:
        st.markdown('<div class="section-label">Step 2</div>', unsafe_allow_html=True)
        render_card("Upload Completed Template", "Upload the filled workbook after offline field-data entry.")
        uploaded_file = st.file_uploader("Upload completed Excel template", type=["xlsx"])

    selected_sheet_groups: list[dict[str, list[str]]] = []
    if uploaded_file is not None:
        try:
            uploaded_sheet_names = get_uploaded_sheet_names(uploaded_file)
        except Exception:  # noqa: BLE001
            uploaded_sheet_names = []
            st.warning("The workbook was uploaded, but its worksheet list could not be previewed yet.")
        else:
            if uploaded_sheet_names:
                selected_sheet_groups = render_sheet_group_builder(uploaded_sheet_names)
    if selected_sheet_groups and not COMPONENT_TEMPLATE_FILE.exists():
        st.warning("The component summary template file 'summary_component.xlsx' is missing, so the extra component-summary download will not be available.")

    st.markdown('<div class="section-label">Step 3</div>', unsafe_allow_html=True)
    render_card("Calculate", "Run the existing Python calculation workflow using the uploaded workbook and the default species master file.")
    calc_col1, calc_col2, calc_col3 = st.columns([1, 1, 1.4])
    with calc_col1:
        plot_area_ha = st.number_input("Plot area (ha)", min_value=0.0001, value=float(PLOT_AREA_HA), step=0.1, format="%.4f")
    with calc_col2:
        rai_per_hectare = st.number_input(
            "Rai per hectare",
            min_value=0.0001,
            value=float(RAI_PER_HECTARE),
            step=0.25,
            format="%.4f",
        )
    with calc_col3:
        calculate_clicked = st.button("Calculate", type="primary", use_container_width=True)

    if calculate_clicked:
        st.session_state.last_error = None
        st.session_state.summary_result_bytes = None
        st.session_state.detail_result_bytes = None
        st.session_state.component_result_bytes = None
        st.session_state.result_sheets = None

        if uploaded_file is None:
            st.error("Please upload a completed Excel template before running the calculation.")
        elif uploaded_file.type not in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/octet-stream",
        ):
            st.error("Please upload a valid .xlsx Excel file.")
        elif not TEMPLATE_FILE.exists():
            st.error("The official template file 'template.xlsx' is missing from the project directory.")
        elif not MASTER_FILE.exists():
            st.error("The species reference file 'species_reference_master_v1.xlsx' is missing from the project directory.")
        else:
            with st.spinner("Running forest calculation workflow..."):
                try:
                    summary_result_bytes, detail_result_bytes, component_result_bytes, result_sheets = run_uploaded_workflow(
                        uploaded_file=uploaded_file,
                        plot_area_ha=plot_area_ha,
                        rai_per_hectare=rai_per_hectare,
                        sheet_groups=selected_sheet_groups,
                    )
                    st.session_state.summary_result_bytes = summary_result_bytes
                    st.session_state.detail_result_bytes = detail_result_bytes
                    st.session_state.component_result_bytes = component_result_bytes
                    st.session_state.result_sheets = result_sheets
                    st.success("Calculation completed successfully.")
                    st.info("You can preview the results below and download the summary and detail Excel files.")
                except Exception as exc:  # noqa: BLE001
                    st.session_state.last_error = traceback.format_exc()
                    st.error(
                        "The calculation could not be completed. Please check that the uploaded file uses the official template and contains valid Excel data."
                    )
                    if isinstance(exc, FileNotFoundError):
                        st.error(str(exc))
                    with st.expander("Technical error details"):
                        st.code(st.session_state.last_error)

    if st.session_state.result_sheets:
        preview_results(st.session_state.result_sheets)

        st.markdown('<div class="section-label">Step 5</div>', unsafe_allow_html=True)
        render_card(
            "Download Calculated Excel Files",
            "Download the generated summary-by-site workbook, the detailed workbook, and the component-summary workbook when component groups are defined.",
        )
        if st.session_state.component_result_bytes:
            dl_col1, dl_col2, dl_col3 = st.columns(3)
        else:
            dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            st.download_button(
                "Download Summary-by-site Excel",
                data=st.session_state.summary_result_bytes,
                file_name=SUMMARY_OUTPUT_FILENAME,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with dl_col2:
            st.download_button(
                "Download Details Excel",
                data=st.session_state.detail_result_bytes,
                file_name=DETAIL_OUTPUT_FILENAME,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        if st.session_state.component_result_bytes:
            with dl_col3:
                st.download_button(
                    "Download Component Summary Excel",
                    data=st.session_state.component_result_bytes,
                    file_name=COMPONENT_OUTPUT_FILENAME,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )


if __name__ == "__main__":
    main()
