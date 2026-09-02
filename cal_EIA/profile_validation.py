from __future__ import annotations

import math
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_COLUMNS = [
    "no",
    "species",
    "girth_cm",
    "height_m",
    "first_branch_m",
    "x",
    "y",
    "crown_x_plus",
    "crown_x_minus",
    "crown_y_plus",
    "crown_y_minus",
]

PROFILE_HEADER_CELLS = {
    "A1": "No",
    "B1": "Species",
    "C2": "Girth(cm)",
    "D1": "Height(m)",
    "D2": "Total",
    "E2": "1 branch",
    "F1": "Position",
    "F2": "X",
    "G2": "Y",
    "H1": "Crown cover",
    "H2": "X+",
    "I2": "X-",
    "J2": "Y+",
    "K2": "Y-",
}


def normalize_header_value(value: object) -> str:
    if value is None:
        return ""
    return "".join(character.lower() for character in str(value).strip() if character.isalnum())


def parse_finite_number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def validate_profile_sheet(excel_path: Path, sheet_name: str) -> dict[str, object]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return {
                "sheetName": sheet_name,
                "treeCount": 0,
                "valid": False,
                "errors": [f"Worksheet '{sheet_name}' does not exist."],
                "warnings": [],
            }

        worksheet = workbook[sheet_name]
        errors: list[str] = []
        warnings: list[str] = []
        for cell, expected in PROFILE_HEADER_CELLS.items():
            actual = worksheet[cell].value
            if normalize_header_value(actual) != normalize_header_value(expected):
                errors.append(f"{cell} must be '{expected}' (found {actual!r}).")

        tree_rows: list[tuple[int, tuple[object, ...]]] = []
        for excel_row, values in enumerate(
            worksheet.iter_rows(min_row=3, min_col=1, max_col=len(EXPECTED_COLUMNS), values_only=True),
            start=3,
        ):
            if any(value is not None and str(value).strip() for value in values):
                tree_rows.append((excel_row, values))

        if not tree_rows:
            errors.append("No tree records were found below the two-row profile header.")
            return {
                "sheetName": sheet_name,
                "treeCount": 0,
                "valid": False,
                "errors": errors,
                "warnings": warnings,
            }

        tree_numbers: list[float] = []
        for excel_row, values in tree_rows:
            species = "" if values[1] is None else str(values[1]).strip()
            if not species:
                errors.append(f"Row {excel_row}: Species is required.")

            numbers = {
                column: parse_finite_number(values[index])
                for index, column in enumerate(EXPECTED_COLUMNS)
                if column != "species"
            }
            for column, number in numbers.items():
                if number is None:
                    errors.append(f"Row {excel_row}: {column} must be a finite number and cannot be blank.")

            if any(number is None for number in numbers.values()):
                continue

            tree_numbers.append(float(numbers["no"]))
            if numbers["girth_cm"] <= 0:
                errors.append(f"Row {excel_row}: girth_cm must be greater than zero.")
            if numbers["height_m"] <= 0:
                errors.append(f"Row {excel_row}: height_m must be greater than zero.")
            if numbers["first_branch_m"] < 0:
                errors.append(f"Row {excel_row}: first_branch_m cannot be negative.")
            if numbers["first_branch_m"] > numbers["height_m"]:
                errors.append(f"Row {excel_row}: first_branch_m cannot exceed height_m.")
            for column in ["crown_x_plus", "crown_x_minus", "crown_y_plus", "crown_y_minus"]:
                if numbers[column] < 0:
                    errors.append(f"Row {excel_row}: {column} cannot be negative.")

        duplicate_numbers = sorted({number for number in tree_numbers if tree_numbers.count(number) > 1})
        if duplicate_numbers:
            warnings.append(f"Duplicate tree numbers: {duplicate_numbers}.")

        return {
            "sheetName": sheet_name,
            "treeCount": len(tree_rows),
            "valid": not errors,
            "errors": errors,
            "warnings": warnings,
        }
    finally:
        workbook.close()


def inspect_profile_workbook(excel_path: Path) -> dict[str, object]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()

    sheets = [validate_profile_sheet(excel_path, sheet_name) for sheet_name in sheet_names]
    return {
        "sheetNames": sheet_names,
        "validSheetNames": [str(sheet["sheetName"]) for sheet in sheets if sheet["valid"]],
        "invalidSheets": [sheet for sheet in sheets if not sheet["valid"]],
        "sheets": sheets,
    }
