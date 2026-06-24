from __future__ import annotations

from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import font_manager
from matplotlib.lines import Line2D
from matplotlib.patches import PathPatch, Rectangle
from matplotlib.path import Path as MplPath
from openpyxl import load_workbook


EXPECTED_COLUMNS = [
    "no",
    "species",
    "girth_cm",
    "height_m",
    "first_branch_m",
    "x",
    "y",
    "crown_x_plus",
    "crown_x_minus",
    "crown_y_plus",
    "crown_y_minus",
]

# A named tree row must be complete.  Previously incomplete rows were silently
# dropped, which could make a species appear to disappear from a profile.
PROFILE_NUMERIC_COLUMNS = [
    "no",
    "girth_cm",
    "height_m",
    "first_branch_m",
    "x",
    "y",
    "crown_x_plus",
    "crown_x_minus",
    "crown_y_plus",
    "crown_y_minus",
]

PLOT_PALETTE = [
    "#43a047",
    "#1e88e5",
    "#8e24aa",
    "#f4511e",
    "#00acc1",
    "#7cb342",
    "#5e35b1",
    "#d81b60",
    "#546e7a",
    "#00897b",
    "#e53935",
    "#6d4c41",
    "#3949ab",
    "#fb8c00",
    "#8d6e63",
    "#00a0dc",
]

PROFILE_CROWN_WIDTH_SCALE = 1.18
PROFILE_CROWN_HEIGHT_SCALE = 0.7
TRUNK_CROWN_OVERLAP_RATIO = 0.20
FIRST_BRANCH_LENGTH_RATIO = 0.18
FIRST_BRANCH_MIN_LENGTH = 0.45
FIRST_BRANCH_MAX_LENGTH = 1.15
FIRST_BRANCH_BUSH_SCALE = 1.56
SIDE_PADDING_METERS = 4.6
THAI_FONT_FILES = [
    Path(__file__).with_name("Sarabun-Regular.ttf"),
    Path(__file__).with_name("NotoSansThai-Regular.ttf"),
]


def configure_matplotlib() -> None:
    for font_path in THAI_FONT_FILES:
        if font_path.exists():
            font_manager.fontManager.addfont(str(font_path))
    matplotlib.rcParams["axes.unicode_minus"] = False


def get_thai_font_properties(size: float | None = None, weight: str | None = None) -> font_manager.FontProperties | None:
    for font_path in THAI_FONT_FILES:
        if font_path.exists():
            return font_manager.FontProperties(fname=str(font_path), size=size, weight=weight)
    return None


def load_profile_sheet(excel_path: Path, sheet_name: str) -> pd.DataFrame:
    raw = pd.read_excel(excel_path, sheet_name=sheet_name, header=[0, 1])
    raw.columns = EXPECTED_COLUMNS
    df = raw.copy()
    df["species"] = df["species"].fillna("").astype(str).str.strip()
    for column in PROFILE_NUMERIC_COLUMNS:
        df[column] = pd.to_numeric(df[column], errors="coerce")

    named_tree_rows = df["species"].ne("")
    incomplete_rows = named_tree_rows & df[PROFILE_NUMERIC_COLUMNS].isna().any(axis=1)
    if incomplete_rows.any():
        details = []
        for row_index, row in df.loc[incomplete_rows].iterrows():
            missing_columns = [
                column
                for column in PROFILE_NUMERIC_COLUMNS
                if pd.isna(row[column])
            ]
            # Two header rows precede the DataFrame, so index 0 is Excel row 3.
            details.append(
                f"row {row_index + 3} ({row['species']}): {', '.join(missing_columns)}"
            )
        raise ValueError(
            f"Sheet '{sheet_name}' has incomplete tree profile data: "
            + "; ".join(details)
        )

    return df.loc[named_tree_rows].reset_index(drop=True)


def audit_profile_sheet(excel_path: Path, sheet_name: str) -> dict[str, object]:
    """Return the exact tree and species counts that will be rendered."""
    df = load_profile_sheet(excel_path, sheet_name)
    species = sorted(df["species"].unique().tolist())
    return {
        "sheetName": sheet_name,
        "treeCount": int(len(df)),
        "speciesCount": int(len(species)),
        "species": species,
    }


def list_profile_sheets(excel_path: Path) -> list[str]:
    # Explicitly close the workbook so temporary uploaded files can be removed
    # reliably on Windows after profile generation finishes.
    with pd.ExcelFile(excel_path) as workbook:
        return workbook.sheet_names


def build_species_color_map(species_names: list[str]) -> dict[str, str]:
    unique_species = sorted(dict.fromkeys(species_names))
    return {
        species: PLOT_PALETTE[index % len(PLOT_PALETTE)]
        for index, species in enumerate(unique_species)
    }


def make_output_path(excel_path: Path, output_dir: Path, sheet_name: str) -> Path:
    stem = excel_path.stem
    safe_sheet_name = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in sheet_name)
    return output_dir / f"{stem}_{safe_sheet_name}_profile.png"


def build_trunk_widths(df: pd.DataFrame) -> pd.Series:
    girth = df["girth_cm"].fillna(df["girth_cm"].median()).astype(float)
    min_girth = float(girth.min())
    max_girth = float(girth.max())
    if np.isclose(min_girth, max_girth):
        return pd.Series(2.2, index=df.index)

    normalized = (girth - min_girth) / (max_girth - min_girth)
    softened = np.sqrt(normalized)
    return 0.9 + softened * 4.1


def compute_top_view_limits(df: pd.DataFrame) -> tuple[float, float, float, float]:
    left = (df["x"] - df["crown_x_minus"]).min()
    right = (df["x"] + df["crown_x_plus"]).max()
    bottom = (df["y"] - df["crown_y_minus"]).min()
    top = (df["y"] + df["crown_y_plus"]).max()
    return float(left), float(right), float(bottom), float(top)


def compute_profile_limits(df: pd.DataFrame) -> tuple[float, float]:
    left = (df["x"] - df["crown_x_minus"] * PROFILE_CROWN_WIDTH_SCALE).min()
    right = (df["x"] + df["crown_x_plus"] * PROFILE_CROWN_WIDTH_SCALE).max()
    return float(left), float(right)


def pick_first_branch_direction(row: pd.Series | pd.Index | object) -> float:
    seed_value = f"{getattr(row, 'species', '')}|{getattr(row, 'no', '')}|{getattr(row, 'x', '')}|{getattr(row, 'y', '')}"
    return -1.0 if sum(ord(char) for char in seed_value) % 2 == 0 else 1.0


def compute_first_branch_length(crown_width: float) -> float:
    return float(np.clip(crown_width * FIRST_BRANCH_LENGTH_RATIO, FIRST_BRANCH_MIN_LENGTH, FIRST_BRANCH_MAX_LENGTH))


def add_bushy_crown(
    ax: plt.Axes,
    center_x: float,
    center_y: float,
    width: float,
    height: float,
    color: str,
    alpha: float,
    zorder: int,
    edgecolor: str | None = None,
    linewidth: float = 0.0,
) -> None:
    angles = np.linspace(0, 2 * np.pi, 220, endpoint=False)
    width = max(width, 0.18)
    height = max(height, 0.18)
    phase = (center_x * 0.37 + center_y * 0.19 + width * 0.11) % (2 * np.pi)
    modulation = (
        1.0
        + 0.08 * np.sin(5 * angles + phase)
        + 0.05 * np.sin(9 * angles + phase / 2)
        + 0.03 * np.cos(13 * angles - phase)
    )
    modulation = np.clip(modulation, 0.86, 1.16)

    radius_x = (width / 2) * modulation
    radius_y = (height / 2) * modulation
    xs = center_x + radius_x * np.cos(angles)
    ys = center_y + radius_y * np.sin(angles)

    vertices = np.column_stack([xs, ys])
    vertices = np.vstack([vertices, vertices[0]])
    codes = [MplPath.MOVETO] + [MplPath.LINETO] * (len(vertices) - 2) + [MplPath.CLOSEPOLY]
    patch = PathPatch(
        MplPath(vertices, codes),
        facecolor=color,
        edgecolor=edgecolor if edgecolor else "none",
        linewidth=linewidth,
        alpha=alpha,
        zorder=zorder,
        joinstyle="round",
        capstyle="round",
    )
    ax.add_patch(patch)


def draw_top_view(ax: plt.Axes, df: pd.DataFrame, colors: dict[str, str]) -> None:
    x_min = float(np.floor(df["x"].min()))
    x_max = float(np.ceil(df["x"].max()))
    y_min = float(np.floor(df["y"].min()))
    y_max = float(np.ceil(df["y"].max()))
    crown_left, crown_right, crown_bottom, crown_top = compute_top_view_limits(df)

    for row in df.itertuples(index=False):
        crown_width = max(row.crown_x_plus + row.crown_x_minus, 0.2)
        crown_height = max(row.crown_y_plus + row.crown_y_minus, 0.2)
        crown_center_x = row.x + (row.crown_x_plus - row.crown_x_minus) / 2
        crown_center_y = row.y + (row.crown_y_plus - row.crown_y_minus) / 2
        add_bushy_crown(
            ax=ax,
            center_x=crown_center_x,
            center_y=crown_center_y,
            width=crown_width,
            height=crown_height,
            color=colors[row.species],
            alpha=0.5,
            zorder=2,
        )

    ax.scatter(df["x"], df["y"], s=8, color="black", zorder=3)
    plot_width = max(x_max - x_min, 1.0)
    plot_height = max(y_max - y_min, 1.0)
    ax.add_patch(
        Rectangle(
            (x_min, y_min),
            plot_width,
            plot_height,
            fill=False,
            linewidth=1.8,
            edgecolor="black",
            zorder=4,
        )
    )
    ax.set_xlim(np.floor(crown_left - SIDE_PADDING_METERS), np.ceil(crown_right + SIDE_PADDING_METERS))
    ax.set_ylim(np.floor(crown_bottom - 1.0), np.ceil(crown_top + 1.0))
    ax.set_aspect("equal", adjustable="box")
    ax.set_xlabel("Distance (m.)")
    ax.set_ylabel("Distance (m.)")
    ax.grid(False)
    ax.set_xticks(np.arange(0, 41, 5))
    ax.set_yticks(np.arange(0, np.ceil(crown_top + 1.0) + 1, 5))
    ax.spines[["top", "right", "left", "bottom"]].set_visible(False)


def draw_profile_view(ax: plt.Axes, df: pd.DataFrame, colors: dict[str, str]) -> None:
    trunk_widths = build_trunk_widths(df)
    draw_df = df.copy()
    draw_df["trunk_width"] = trunk_widths
    draw_df["crown_width"] = (
        (draw_df["crown_x_plus"] + draw_df["crown_x_minus"]) * PROFILE_CROWN_WIDTH_SCALE
    ).clip(lower=0.6)
    draw_df["crown_depth"] = (
        (draw_df["height_m"] - draw_df["first_branch_m"]) * PROFILE_CROWN_HEIGHT_SCALE
    ).clip(lower=0.8)
    draw_df["crown_area"] = draw_df["crown_width"] * draw_df["crown_depth"]

    for row in draw_df.itertuples(index=False):
        trunk_top_y = min(max(row.height_m - row.crown_depth, 0) + row.crown_depth * TRUNK_CROWN_OVERLAP_RATIO, 19.6)
        ax.plot(
            [row.x, row.x],
            [0, trunk_top_y],
            color="#4e342e",
            linewidth=float(row.trunk_width),
            alpha=0.95,
            zorder=2,
            solid_capstyle="round",
        )

        if pd.notna(row.first_branch_m):
            branch_origin_y = float(np.clip(row.first_branch_m, 0.45, max(row.height_m - 0.4, 0.45)))
            branch_direction = pick_first_branch_direction(row)
            branch_length = compute_first_branch_length(float(row.crown_width))
            branch_dx = branch_length * np.cos(np.deg2rad(45)) * branch_direction
            branch_dy = branch_length * np.sin(np.deg2rad(45))
            branch_end_x = float(row.x + branch_dx)
            branch_end_y = float(min(branch_origin_y + branch_dy, 19.65))
            ax.plot(
                [row.x, branch_end_x],
                [branch_origin_y, branch_end_y],
                color="#4e342e",
                linewidth=max(float(row.trunk_width) * 0.45, 1.0),
                alpha=0.95,
                zorder=2.4,
                solid_capstyle="round",
            )
            add_bushy_crown(
                ax=ax,
                center_x=branch_end_x,
                center_y=branch_end_y,
                width=max(branch_length * FIRST_BRANCH_BUSH_SCALE, 0.2),
                height=max(branch_length * FIRST_BRANCH_BUSH_SCALE, 0.2),
                color=colors[row.species],
                edgecolor=colors[row.species],
                linewidth=0.35,
                alpha=0.8,
                zorder=3.2,
            )

    crown_df = draw_df.sort_values(["crown_area", "height_m"], ascending=[False, False])
    for row in crown_df.itertuples(index=False):
        crown_width = float(row.crown_width)
        crown_base_y = float(max(row.height_m - row.crown_depth, 0))
        crown_depth = float(min(row.crown_depth, 20 - crown_base_y))
        crown_center_x = float(row.x + (row.crown_x_plus - row.crown_x_minus) / 2)
        crown_center_y = float(crown_base_y + crown_depth / 2)
        add_bushy_crown(
            ax=ax,
            center_x=crown_center_x,
            center_y=crown_center_y,
            width=crown_width,
            height=crown_depth,
            color=colors[row.species],
            edgecolor=colors[row.species],
            linewidth=0.6,
            alpha=0.5,
            zorder=3,
        )

    profile_left, profile_right = compute_profile_limits(df)
    ax.set_xlim(np.floor(profile_left - SIDE_PADDING_METERS), np.ceil(profile_right + SIDE_PADDING_METERS))
    ax.set_ylim(0, 20)
    ax.set_xticks(np.arange(0, 41, 5))
    ax.set_yticks(np.arange(0, 21, 5))
    ax.set_xlabel("Distance (m.)")
    ax.set_ylabel("Height (m.)")
    ax.grid(False)
    ax.spines[["top", "right", "left", "bottom"]].set_visible(False)


def render_sheet_profile(excel_path: Path, sheet_name: str, output_dir: Path) -> Path:
    df = load_profile_sheet(excel_path, sheet_name)
    if df.empty:
        raise ValueError(f"Sheet '{sheet_name}' does not contain usable tree profile data.")

    colors = build_species_color_map(df["species"].tolist())
    output_path = make_output_path(excel_path, output_dir, sheet_name)

    figure = plt.figure(figsize=(14.5, 13.2))
    grid = figure.add_gridspec(nrows=3, ncols=1, height_ratios=[1.0, 1.08, 0.4])
    top_ax = figure.add_subplot(grid[0])
    profile_ax = figure.add_subplot(grid[1])
    legend_ax = figure.add_subplot(grid[2])

    draw_top_view(top_ax, df, colors)
    draw_profile_view(profile_ax, df, colors)

    handles = [
        Line2D(
            [0],
            [0],
            marker="s",
            linestyle="None",
            markersize=12,
            markerfacecolor=colors[species],
            markeredgecolor="none",
            label=species,
            alpha=0.85,
        )
        for species in colors
    ]
    handles.append(
        Line2D(
            [0],
            [0],
            marker="o",
            linestyle="None",
            markersize=6,
            markerfacecolor="black",
            markeredgecolor="black",
            label="\u0e15\u0e33\u0e41\u0e2b\u0e19\u0e48\u0e07\u0e25\u0e33\u0e15\u0e49\u0e19",
        )
    )
    legend_ax.axis("off")
    x_left, x_right = profile_ax.get_xlim()
    axis_span = max(x_right - x_left, 1.0)
    legend_left = max((0 - x_left) / axis_span, 0.0)
    legend_width = min(40 / axis_span, 1.0 - legend_left)
    legend_font = get_thai_font_properties(size=9.5)
    legend_title_font = get_thai_font_properties(size=10.5, weight="bold")
    legend_ax.legend(
        handles=handles,
        title="\u0e0a\u0e19\u0e34\u0e14\u0e1e\u0e31\u0e19\u0e18\u0e38\u0e4c\u0e44\u0e21\u0e49",
        loc="center",
        mode="expand",
        ncol=5,
        frameon=True,
        fancybox=True,
        framealpha=0.96,
        edgecolor="#d6d6d6",
        prop=legend_font,
        title_fontproperties=legend_title_font,
        columnspacing=1.0,
        handletextpad=0.5,
        borderpad=0.9,
        labelspacing=0.8,
        bbox_to_anchor=(legend_left, 0.08, legend_width, 0.84),
    )

    figure.subplots_adjust(left=0.09, right=0.91, top=0.96, bottom=0.06, hspace=0.2)
    figure.savefig(output_path, dpi=220, bbox_inches="tight")
    plt.close(figure)
    return output_path


def render_workbook_profiles(excel_path: Path, output_dir: Path) -> list[Path]:
    configure_matplotlib()
    output_dir.mkdir(parents=True, exist_ok=True)
    output_paths: list[Path] = []
    for sheet_name in list_profile_sheets(excel_path):
        try:
            output_paths.append(render_sheet_profile(excel_path, sheet_name, output_dir))
        except ValueError:
            continue
    if not output_paths:
        raise ValueError("No worksheet contained usable profile data.")
    return output_paths


def render_workbook_profile_map(excel_path: Path, output_dir: Path) -> list[tuple[str, Path]]:
    configure_matplotlib()
    output_dir.mkdir(parents=True, exist_ok=True)
    outputs: list[tuple[str, Path]] = []
    for sheet_name in list_profile_sheets(excel_path):
        try:
            outputs.append((sheet_name, render_sheet_profile(excel_path, sheet_name, output_dir)))
        except ValueError:
            continue
    if not outputs:
        raise ValueError("No worksheet contained usable profile data.")
    return outputs


def create_profile_template(source_path: Path, output_path: Path) -> Path:
    workbook = load_workbook(source_path)
    try:
        for ws in workbook.worksheets:
            if ws.max_row > 2:
                ws.delete_rows(3, ws.max_row - 2)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()
    return output_path


def zip_profile_outputs(image_paths: list[Path], zip_path: Path) -> Path:
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    with ZipFile(zip_path, "w", compression=ZIP_DEFLATED) as zip_file:
        for image_path in image_paths:
            zip_file.write(image_path, arcname=image_path.name)
    return zip_path
