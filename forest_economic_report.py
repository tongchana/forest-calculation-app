from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from forest_ecosystem_loss import ECOSYSTEM_UNIT_PRICES, build_ecosystem_loss_detail_rows
from forest_integration import get_component_name_maps


THIN_SIDE = Side(style="thin", color="808080")
THICK_SIDE = Side(style="medium", color="666666")
TITLE_FILL = PatternFill("solid", fgColor="D9E2F3")
HEADER_FILL = PatternFill("solid", fgColor="BFBFBF")
SECTION_FILL = PatternFill("solid", fgColor="D9D9D9")
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")


def safe_sheet_name(name: str, existing_names: set[str]) -> str:
    cleaned = re.sub(r'[:\\/?*\[\]]', "", str(name).strip()) or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    counter = 2
    while candidate in existing_names:
        suffix = f"_{counter}"
        candidate = f"{cleaned[: 31 - len(suffix)]}{suffix}"
        counter += 1
    existing_names.add(candidate)
    return candidate


REPORT_FONT_NAME = "TH Sarabun New"
REPORT_FONT_SIZE = 16
NUMBER_FORMAT = "#,##0.00"
ECOSYSTEM_TOTAL_KEYS = (
    "soil",
    "nitrogen",
    "phosphorus",
    "potassium",
    "water_regulation",
    "warming",
    "co2_absorption",
)


def _apply_cell_style(cell, *, bold: bool = False, fill=None, align: str = "left", border: Side = THIN_SIDE, size: int = REPORT_FONT_SIZE):
    cell.font = Font(name=REPORT_FONT_NAME, size=REPORT_FONT_SIZE, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.fill = fill or WHITE_FILL
    cell.border = Border(left=border, right=border, top=border, bottom=border)


def _apply_number_format(cell) -> None:
    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
        cell.number_format = NUMBER_FORMAT


def _set_table_header(ws, row: int, headers: list[str], fills: list[PatternFill] | None = None) -> None:
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=index, value=header)
        _apply_cell_style(cell, bold=True, fill=(fills[index - 1] if fills else HEADER_FILL), align="center")


def _auto_fit_columns(ws, min_width: int = 12, max_width: int = 32) -> None:
    for column_cells in ws.columns:
        max_len = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = max(min_width, min(max_len + 2, max_width))


def _write_frame(ws, start_row: int, title: str, frame: pd.DataFrame) -> int:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max(1, frame.shape[1]))
    title_cell = ws.cell(start_row, 1, title)
    _apply_cell_style(title_cell, bold=True, fill=SECTION_FILL, align="center", size=13, border=THICK_SIDE)
    start_row += 1
    if frame.empty:
        ws.cell(start_row, 1, "ไม่มีข้อมูล")
        _apply_cell_style(ws.cell(start_row, 1), fill=WARN_FILL)
        return start_row + 2
    headers = [str(column) for column in frame.columns]
    _set_table_header(ws, start_row, headers)
    start_row += 1
    for _, row in frame.iterrows():
        for column_index, value in enumerate(row.tolist(), start=1):
            if isinstance(value, list):
                value = ", ".join(str(item) for item in value)
            elif isinstance(value, dict):
                value = str(value)
            cell = ws.cell(start_row, column_index, value=value)
            align = "right" if isinstance(value, (int, float)) else "left"
            _apply_cell_style(cell, fill=WHITE_FILL, align=align)
            _apply_number_format(cell)
        start_row += 1
    return start_row + 1


def _display_value(value: object) -> object:
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    if isinstance(value, dict):
        return str(value)
    return value


def _sort_tq_values(values: Iterable[object]) -> list[str]:
    def sort_key(text: str) -> tuple[int, float | str]:
        normalized = str(text).strip().upper().replace(" ", "")
        if normalized.startswith("TQ"):
            suffix = normalized[2:]
            try:
                return 0, float(suffix)
            except ValueError:
                return 1, suffix
        return 2, normalized

    cleaned = [str(value).strip() for value in values if str(value).strip()]
    return sorted(dict.fromkeys(cleaned), key=sort_key)


def _write_component_table(
    ws,
    start_row: int,
    title: str,
    headers: list[str],
    rows: list[list[object]],
    total_fill: PatternFill | None = None,
) -> int:
    last_col = len(headers)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=last_col)
    ws.cell(start_row, 1, title)
    _apply_cell_style(ws.cell(start_row, 1), bold=True, fill=SECTION_FILL, align="left", size=13, border=THICK_SIDE)
    start_row += 1
    _set_table_header(ws, start_row, headers)
    start_row += 1
    if not rows:
        ws.cell(start_row, 1, "ไม่มีข้อมูล")
        _apply_cell_style(ws.cell(start_row, 1), fill=WARN_FILL)
        return start_row + 2
    for row in rows:
        is_total = str(row[0]).strip() in {"รวม", "รวมทั้งหมด"}
        for column_index, value in enumerate(row, start=1):
            rendered = _display_value(value)
            ws.cell(start_row, column_index, rendered)
            align = "right" if isinstance(rendered, (int, float)) else "left"
            _apply_cell_style(
                ws.cell(start_row, column_index),
                fill=total_fill if is_total and total_fill else WHITE_FILL,
                align=align,
                bold=is_total,
            )
            _apply_number_format(ws.cell(start_row, column_index))
        start_row += 1
    return start_row + 1


def _component_order(outputs: dict[str, pd.DataFrame]) -> list[tuple[str, str]]:
    internal_to_display, _ = get_component_name_maps(outputs)
    meta = outputs.get("__meta__", {})
    groups = meta.get("sheet_groups", []) if isinstance(meta, dict) else []
    ordered: list[tuple[str, str]] = []
    for group in groups:
        internal_name = str(group.get("internal_name") or "").strip()
        if not internal_name:
            continue
        ordered.append((internal_name, internal_to_display.get(internal_name, internal_name)))
    return ordered


def _component_value_lookup(rows: Iterable[dict[str, object]], key_field: str = "component_id") -> dict[str, dict[str, object]]:
    result: dict[str, dict[str, object]] = {}
    for row in rows:
        key = str(row.get(key_field) or "").strip()
        if key:
            result[key] = row
    return result


def _summary_all_lookup(outputs: dict[str, pd.DataFrame]) -> dict[str, dict[str, object]]:
    frame = outputs.get("SUMMARY_ALL", pd.DataFrame())
    if frame.empty:
        return {}
    result: dict[str, dict[str, object]] = {}
    for _, row in frame.iterrows():
        key = str(row.get("sheet_name") or "").strip()
        if key:
            result[key] = row.to_dict()
    return result


def _warning_rows(bundle: dict[str, object]) -> list[dict[str, object]]:
    warning_rows: list[dict[str, object]] = []
    for module_name in ["forest_economics", "wood_future_value", "regeneration_loss", "ecosystem_loss"]:
        module_result = bundle.get(module_name, {})
        for warning in module_result.get("warnings", []) if isinstance(module_result, dict) else []:
            warning_rows.append({"module": module_name, "component": "", "message": warning})
    for row in bundle.get("forest_economics", {}).get("componentSummaries", []):
        for warning in row.get("warnings", []):
            warning_rows.append({"module": "forest_economics", "component": row.get("component_name", ""), "message": warning})
    for row in bundle.get("ecosystem_loss", {}).get("componentSummaries", []):
        for warning in row.get("warnings", []):
            warning_rows.append({"module": "ecosystem_loss", "component": row.get("component_name", ""), "message": warning})
    return warning_rows


def _build_master_rows(outputs: dict[str, pd.DataFrame], bundle: dict[str, object]) -> list[tuple[str, str, dict[str, float | None] | None]]:
    component_summaries = _component_value_lookup(bundle.get("forest_economics", {}).get("componentSummaries", []))
    regeneration = _component_value_lookup(bundle.get("regeneration_loss", {}).get("componentSummaries", []))
    ecosystem = _component_value_lookup(bundle.get("ecosystem_loss", {}).get("componentSummaries", []))
    summary_all = _summary_all_lookup(outputs)

    def collect(field_getter):
        values: dict[str, float | None] = {}
        for component_id, _display_name in _component_order(outputs):
            values[component_id] = field_getter(component_id)
        return values

    return [
        ("พื้นที่โครงการ", "", None),
        ("ไร่", "", collect(lambda component_id: component_summaries.get(component_id, {}).get("component_area_rai"))),
        ("การสูญเสียต้นไม้", "", None),
        ("ไม้ใหญ่ยืนต้น (ต้น)", "", collect(lambda component_id: summary_all.get(component_id, {}).get("n_tree"))),
        ("ลูกไม้ (ต้น)", "", collect(lambda component_id: summary_all.get(component_id, {}).get("n_sapling"))),
        ("กล้าไม้ (ต้น)", "", collect(lambda component_id: summary_all.get(component_id, {}).get("total_seedling_number"))),
        ("การสูญเสียมูลค่าทางนิเวศวิทยา (ทางตรง)", "", None),
        ("การสูญเสียเนื้อไม้ (ลูกบาศก์เมตร)", "", collect(lambda component_id: component_summaries.get(component_id, {}).get("total_wood_loss_m3"))),
        ("การสูญเสียเนื้อไม้ (บาท)", "", collect(lambda component_id: component_summaries.get(component_id, {}).get("total_wood_value_baht"))),
        ("การสูญเสียมูลค่าไม้หนุ่ม (บาท)", "27 บาท/ต้น/ไร่", collect(lambda component_id: regeneration.get(component_id, {}).get("sapling_loss_baht"))),
        ("การสูญเสียมูลค่ากล้าไม้ (บาท)", "6 บาท/ต้น/ไร่", collect(lambda component_id: regeneration.get(component_id, {}).get("seedling_loss_baht"))),
        ("การประเมินมูลค่าการสูญเสียทางระบบนิเวศ", "", None),
        ("การสูญเสียดินอันเนื่องมาจากการกัดชะพังทลาย", "1,800 บาท/เที่ยว", collect(lambda component_id: _sum_ecosystem_detail(bundle, component_id, "soil"))),
        ("การสูญเสียธาตุไนโตรเจน", "0.035 บาท/กรัม", collect(lambda component_id: _sum_ecosystem_detail(bundle, component_id, "nitrogen"))),
        ("การสูญเสียธาตุฟอสฟอรัส", "0.093 บาท/กรัม", collect(lambda component_id: _sum_ecosystem_detail(bundle, component_id, "phosphorus"))),
        ("การสูญเสียธาตุโพแทสเซียม", "0.88 บาท/กรัม", collect(lambda component_id: _sum_ecosystem_detail(bundle, component_id, "potassium"))),
        ("การสูญเสียระบบควบคุมการดูดซับ-ระบายน้ำ", "1,800 บาท/เที่ยว", collect(lambda component_id: _sum_ecosystem_detail(bundle, component_id, "water_regulation"))),
        ("อากาศที่ร้อนขึ้น", "2.5 บาท/ชั่วโมง", collect(lambda component_id: _sum_ecosystem_detail(bundle, component_id, "warming"))),
        ("การดูดซับก๊าซคาร์บอนไดออกไซด์", "793.5 บาท/ตัน", collect(lambda component_id: _sum_ecosystem_detail(bundle, component_id, "co2_absorption"))),
        ("รวม", "", collect(lambda component_id: _master_loss_total_for_component(bundle, component_summaries, regeneration, component_id))),
    ]


def _sum_ecosystem_detail(bundle: dict[str, object], component_id: str, impact_key: str) -> float | None:
    rows = [row for row in bundle.get("ecosystem_loss", {}).get("groupResults", []) if row.get("component_id") == component_id]
    if not rows:
        return None
    component_area = next(
        (float(row["component_area_rai"]) for row in rows if isinstance(row.get("component_area_rai"), (int, float))),
        None,
    )
    weighted_pairs: list[tuple[float, float]] = []
    for row in rows:
        proxy = _dict_to_ecosystem_proxy(row)
        detail_rows = build_ecosystem_loss_detail_rows(proxy)
        matching = next((item for item in detail_rows if item["impact_key"] == impact_key), None)
        if not matching or not isinstance(matching.get("value_baht_per_rai_per_year"), (int, float)):
            continue
        weight = row.get("representative_area_rai")
        if not isinstance(weight, (int, float)) or weight <= 0:
            weight = row.get("sample_plot_area_rai")
        if not isinstance(weight, (int, float)) or weight <= 0:
            weight = 1.0
        weighted_pairs.append((float(matching["value_baht_per_rai_per_year"]), float(weight)))
    if not weighted_pairs:
        return None
    weighted_per_rai = sum(value * weight for value, weight in weighted_pairs) / sum(weight for _value, weight in weighted_pairs)
    if component_area is None:
        return weighted_per_rai
    return weighted_per_rai * component_area


class _dict_to_ecosystem_proxy:
    def __init__(self, payload: dict[str, object]):
        self.__dict__.update(payload)


def _master_ecosystem_total_for_component(bundle: dict[str, object], component_id: str) -> float | None:
    values = [_sum_ecosystem_detail(bundle, component_id, impact_key) for impact_key in ECOSYSTEM_TOTAL_KEYS]
    numeric = [float(value) for value in values if isinstance(value, (int, float))]
    return sum(numeric) if numeric else None


def _master_loss_total_for_component(
    bundle: dict[str, object],
    component_summaries: dict[str, dict[str, object]],
    regeneration: dict[str, dict[str, object]],
    component_id: str,
) -> float | None:
    values = [
        component_summaries.get(component_id, {}).get("total_wood_value_baht"),
        regeneration.get(component_id, {}).get("sapling_loss_baht"),
        regeneration.get(component_id, {}).get("seedling_loss_baht"),
        _master_ecosystem_total_for_component(bundle, component_id),
    ]
    numeric = [float(value) for value in values if isinstance(value, (int, float))]
    return sum(numeric) if numeric else None


def _write_master_summary(ws, outputs: dict[str, pd.DataFrame], bundle: dict[str, object]) -> None:
    ordered_components = _component_order(outputs)
    headers = ["ประเด็น", "อัตรา"] + [display_name for _component_id, display_name in ordered_components] + ["รวมทั้งหมด"]
    fills = [HEADER_FILL] * len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, "สรุปการประเมินมูลค่าทรัพยากรป่าไม้และระบบนิเวศ")
    _apply_cell_style(ws.cell(1, 1), bold=True, fill=TITLE_FILL, align="center", size=16, border=THICK_SIDE)
    _set_table_header(ws, 2, headers, fills)
    current_row = 3
    for label, rate, values in _build_master_rows(outputs, bundle):
        if values is None:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            ws.cell(current_row, 1, label)
            _apply_cell_style(ws.cell(current_row, 1), bold=True, fill=SECTION_FILL, align="left", size=13, border=THICK_SIDE)
            current_row += 1
            continue
        ws.cell(current_row, 1, label)
        ws.cell(current_row, 2, rate)
        _apply_cell_style(ws.cell(current_row, 1), fill=WHITE_FILL)
        _apply_cell_style(ws.cell(current_row, 2), fill=WHITE_FILL, align="center")
        total = 0.0
        has_total = False
        for offset, (component_id, _display_name) in enumerate(ordered_components, start=3):
            value = values.get(component_id)
            ws.cell(current_row, offset, value)
            _apply_cell_style(ws.cell(current_row, offset), fill=WHITE_FILL, align="right")
            _apply_number_format(ws.cell(current_row, offset))
            if isinstance(value, (int, float)):
                total += float(value)
                has_total = True
        ws.cell(current_row, len(headers), total if has_total else None)
        _apply_cell_style(ws.cell(current_row, len(headers)), fill=TOTAL_FILL, align="right", bold=True)
        _apply_number_format(ws.cell(current_row, len(headers)))
        current_row += 1
    ws.freeze_panes = "C3"
    ws.sheet_view.showGridLines = False
    _auto_fit_columns(ws, min_width=14, max_width=28)


def _component_sheet_frames(outputs: dict[str, pd.DataFrame], bundle: dict[str, object], component_id: str, component_name: str) -> list[tuple[str, pd.DataFrame]]:
    forest_rows = pd.DataFrame(bundle.get("forest_economics", {}).get("detailRows", []))
    species_rows = pd.DataFrame(bundle.get("forest_economics", {}).get("speciesDetailRows", []))
    future_rows = pd.DataFrame()
    for summary in bundle.get("wood_future_value", {}).get("componentSummaries", []):
        if summary.get("component_id") == component_id:
            future_rows = pd.DataFrame(summary.get("period_rows", []))
            break
    regeneration_rows = pd.DataFrame(bundle.get("regeneration_loss", {}).get("componentSummaries", []))
    ecosystem_rows = pd.DataFrame(bundle.get("ecosystem_loss", {}).get("groupResults", []))
    summary_all = outputs.get("SUMMARY_ALL", pd.DataFrame())

    frames = [
        (
            "ข้อมูลทั่วไป",
            pd.DataFrame(
                [
                    {
                        "component_id": component_id,
                        "component_name": component_name,
                        "component_area_rai": next(
                            (row.get("component_area_rai") for row in bundle.get("forest_economics", {}).get("componentSummaries", []) if row.get("component_id") == component_id),
                            None,
                        ),
                        "n_tree": next((row.get("n_tree") for _, row in summary_all[summary_all["sheet_name"] == component_id].iterrows()), None)
                        if not summary_all.empty
                        else None,
                        "n_sapling": next((row.get("n_sapling") for _, row in summary_all[summary_all["sheet_name"] == component_id].iterrows()), None)
                        if not summary_all.empty
                        else None,
                        "total_seedling_number": next((row.get("total_seedling_number") for _, row in summary_all[summary_all["sheet_name"] == component_id].iterrows()), None)
                        if not summary_all.empty
                        else None,
                    }
                ]
            ),
        ),
        (
            "ปริมาตรและมูลค่าตามชั้นคุณภาพไม้",
            forest_rows[forest_rows["component_id"] == component_id].copy() if not forest_rows.empty else pd.DataFrame(),
        ),
        (
            "รายละเอียดตามชนิดไม้",
            species_rows[species_rows["component_id"] == component_id].copy() if not species_rows.empty else pd.DataFrame(),
        ),
        (
            "มูลค่าไม้ในอนาคต",
            future_rows,
        ),
        (
            "ความเสียหายลูกไม้และกล้าไม้",
            regeneration_rows[regeneration_rows["component_id"] == component_id].copy() if not regeneration_rows.empty else pd.DataFrame(),
        ),
        (
            "ผลกระทบระบบนิเวศ",
            ecosystem_rows[ecosystem_rows["component_id"] == component_id].copy() if not ecosystem_rows.empty else pd.DataFrame(),
        ),
    ]
    return frames


def _component_detail_frame(bundle: dict[str, object], component_id: str) -> pd.DataFrame:
    frame = pd.DataFrame(bundle.get("forest_economics", {}).get("detailRows", []))
    if frame.empty:
        return frame
    return frame[frame["component_id"] == component_id].copy()


def _component_species_frame(bundle: dict[str, object], component_id: str) -> pd.DataFrame:
    frame = pd.DataFrame(bundle.get("forest_economics", {}).get("speciesDetailRows", []))
    if frame.empty:
        return frame
    return frame[frame["component_id"] == component_id].copy()


def _component_summary_dict(bundle: dict[str, object], component_id: str) -> dict[str, object]:
    return _component_value_lookup(bundle.get("forest_economics", {}).get("componentSummaries", [])).get(component_id, {})


def _component_regeneration_dict(bundle: dict[str, object], component_id: str) -> dict[str, object]:
    return _component_value_lookup(bundle.get("regeneration_loss", {}).get("componentSummaries", [])).get(component_id, {})


def _component_future_rows(bundle: dict[str, object], component_id: str) -> list[dict[str, object]]:
    for summary in bundle.get("wood_future_value", {}).get("componentSummaries", []):
        if summary.get("component_id") == component_id:
            return list(summary.get("period_rows", []))
    return []


def _build_tq_matrix_rows(
    detail_frame: pd.DataFrame,
    value_field: str,
    tq_columns: list[str],
    component_area_rai: float | None = None,
    include_increment_rate: bool = False,
) -> list[list[object]]:
    if detail_frame.empty:
        return []
    rows: list[list[object]] = []
    for forest_type in sorted(detail_frame["forest_type"].dropna().astype(str).unique().tolist()):
        subset = detail_frame[detail_frame["forest_type"].astype(str) == forest_type]
        row: list[object] = [forest_type]
        if include_increment_rate:
            rate = subset["increment_rate_percent"].iloc[0] if "increment_rate_percent" in subset.columns and not subset.empty else None
            row.append(rate)
        if component_area_rai is not None:
            row.append(component_area_rai)
        total = 0.0
        has_total = False
        for tq in tq_columns:
            tq_subset = subset[subset["tq"].astype(str) == tq]
            value = tq_subset[value_field].sum() if not tq_subset.empty else 0.0
            row.append(value)
            if isinstance(value, (int, float)):
                total += float(value)
                has_total = True
        row.append(total if has_total else None)
        rows.append(row)
    total_row: list[object] = ["รวม"]
    if include_increment_rate:
        total_row.append(None)
    if component_area_rai is not None:
        total_row.append(component_area_rai)
    grand_total = 0.0
    has_grand_total = False
    for tq in tq_columns:
        value = detail_frame[detail_frame["tq"].astype(str) == tq][value_field].sum()
        total_row.append(value)
        if isinstance(value, (int, float)):
            grand_total += float(value)
            has_grand_total = True
    total_row.append(grand_total if has_grand_total else None)
    rows.append(total_row)
    return rows


def _weighted_component_ecosystem_details(bundle: dict[str, object], component_id: str) -> list[dict[str, object]]:
    rows = [row for row in bundle.get("ecosystem_loss", {}).get("groupResults", []) if row.get("component_id") == component_id]
    if not rows:
        return []
    component_area = next(
        (float(row["component_area_rai"]) for row in rows if isinstance(row.get("component_area_rai"), (int, float))),
        None,
    )
    grouped: dict[str, dict[str, object]] = {}
    for row in rows:
        proxy = _dict_to_ecosystem_proxy(row)
        weight = row.get("representative_area_rai")
        if not isinstance(weight, (int, float)) or weight <= 0:
            weight = row.get("sample_plot_area_rai")
        if not isinstance(weight, (int, float)) or weight <= 0:
            weight = 1.0
        for detail in build_ecosystem_loss_detail_rows(proxy):
            impact_key = str(detail["impact_key"])
            accumulator = grouped.setdefault(
                impact_key,
                {
                    "impact_name_th": detail["impact_name_th"],
                    "quantity_unit": detail["quantity_unit"],
                    "unit_price": detail["unit_price"],
                    "unit_price_unit": detail["unit_price_unit"],
                    "weighted_quantity": 0.0,
                    "weighted_value": 0.0,
                    "total_weight": 0.0,
                },
            )
            quantity = detail.get("quantity")
            value = detail.get("value_baht_per_rai_per_year")
            if isinstance(quantity, (int, float)):
                accumulator["weighted_quantity"] += float(quantity) * float(weight)
            if isinstance(value, (int, float)):
                accumulator["weighted_value"] += float(value) * float(weight)
            accumulator["total_weight"] += float(weight)
    result: list[dict[str, object]] = []
    for impact_key, item in grouped.items():
        total_weight = item["total_weight"] or 1.0
        quantity_per_rai = item["weighted_quantity"] / total_weight
        value_per_rai = item["weighted_value"] / total_weight
        result.append(
            {
                "impact_key": impact_key,
                "impact_name_th": item["impact_name_th"],
                "quantity": quantity_per_rai,
                "quantity_unit": item["quantity_unit"],
                "unit_price": item["unit_price"],
                "unit_price_unit": item["unit_price_unit"],
                "value_baht_per_year": value_per_rai * component_area if component_area is not None else value_per_rai,
                "value_unit": "บาท/ปี" if component_area is not None else "บาท/ไร่/ปี",
            }
        )
    return result


def _write_component_sheet(ws, outputs: dict[str, pd.DataFrame], bundle: dict[str, object], component_id: str, component_name: str) -> None:
    detail_frame = _component_detail_frame(bundle, component_id)
    species_frame = _component_species_frame(bundle, component_id)
    summary_dict = _component_summary_dict(bundle, component_id)
    regeneration_dict = _component_regeneration_dict(bundle, component_id)
    future_rows = _component_future_rows(bundle, component_id)
    ecosystem_rows = _weighted_component_ecosystem_details(bundle, component_id)
    tq_columns = _sort_tq_values(detail_frame["tq"].tolist()) if not detail_frame.empty else []
    component_area_rai = summary_dict.get("component_area_rai")

    ws.merge_cells("A1:H1")
    ws["A1"] = f"รายงานองค์ประกอบ: {component_name}"
    _apply_cell_style(ws["A1"], bold=True, fill=TITLE_FILL, align="center", size=16, border=THICK_SIDE)
    ws.merge_cells("A2:H2")
    ws["A2"] = f"พื้นที่โครงการ: {component_area_rai if component_area_rai is not None else '-'} ไร่"
    _apply_cell_style(ws["A2"], fill=HEADER_FILL, align="center", bold=True, border=THICK_SIDE)
    current_row = 4

    general_headers = ["รายการ", "ค่า", "หน่วย"]
    general_rows = [
        ["พื้นที่โครงการ", component_area_rai, "ไร่"],
        ["ประเภทป่าที่พบ", ", ".join(str(item) for item in summary_dict.get("forest_types_detected", [])), ""],
        ["ชั้นคุณภาพไม้ที่พบ", ", ".join(str(item) for item in summary_dict.get("tq_detected", [])), ""],
        ["ไม้ใหญ่ยืนต้น", next((row.get("n_tree") for _, row in outputs.get("SUMMARY_ALL", pd.DataFrame())[outputs.get("SUMMARY_ALL", pd.DataFrame())["sheet_name"] == component_id].iterrows()), None) if not outputs.get("SUMMARY_ALL", pd.DataFrame()).empty else None, "ต้น"],
        ["ลูกไม้", regeneration_dict.get("sapling_density_per_rai"), "ต้น/ไร่"],
        ["กล้าไม้", regeneration_dict.get("seedling_density_per_rai"), "ต้น/ไร่"],
    ]
    current_row = _write_component_table(ws, current_row, "ข้อมูลทั่วไป", general_headers, general_rows)

    volume_headers = ["กลุ่มป่า", "ขนาดพื้นที่โครงการ (ไร่)", *tq_columns, "รวม"]
    volume_rows = _build_tq_matrix_rows(detail_frame, "volume_per_rai_m3", tq_columns, component_area_rai=component_area_rai)
    current_row = _write_component_table(ws, current_row, "ปริมาตรไม้เฉลี่ย (ลูกบาศก์เมตรต่อไร่)", volume_headers, volume_rows, total_fill=TOTAL_FILL)

    loss_headers = ["กลุ่มป่า", *tq_columns, "รวม"]
    loss_rows = _build_tq_matrix_rows(detail_frame, "wood_loss_m3", tq_columns)
    current_row = _write_component_table(ws, current_row, "การสูญเสียรวม (ลูกบาศก์เมตร)", loss_headers, loss_rows, total_fill=TOTAL_FILL)

    increment_headers = ["กลุ่มป่า", "อัตราเพิ่มพูนป่า (%)", "พื้นที่โครงการ", *tq_columns, "รวม"]
    increment_rows = _build_tq_matrix_rows(
        detail_frame,
        "annual_increment_m3_per_year",
        tq_columns,
        component_area_rai=component_area_rai,
        include_increment_rate=True,
    )
    current_row = _write_component_table(ws, current_row, "ความเพิ่มพูนรายปี (ลูกบาศก์เมตร)", increment_headers, increment_rows, total_fill=TOTAL_FILL)

    annual_value_headers = ["กลุ่มป่า", "พื้นที่โครงการ", *tq_columns, "รวม"]
    annual_value_rows = _build_tq_matrix_rows(detail_frame, "annual_wood_value_baht", tq_columns, component_area_rai=component_area_rai)
    current_row = _write_component_table(ws, current_row, "มูลค่าเพิ่มรายปี (บาท)", annual_value_headers, annual_value_rows, total_fill=TOTAL_FILL)

    wood_value_rows = _build_tq_matrix_rows(detail_frame, "wood_value_baht", tq_columns)
    current_row = _write_component_table(ws, current_row, "มูลค่าเนื้อไม้สูญเสีย (บาท)", loss_headers, wood_value_rows, total_fill=TOTAL_FILL)

    future_headers = ["ระยะเวลา (N)", "มูลค่าไม้รายปี (A)", "มูลค่าไม้ในอนาคต (FV)", "มูลค่าไม้ในปัจจุบัน (PV)"]
    future_table_rows = [
        [row.get("period_years"), row.get("annual_wood_value_baht"), row.get("future_value_baht"), row.get("present_value_baht")]
        for row in future_rows
    ]
    current_row = _write_component_table(ws, current_row, "มูลค่าไม้ในอนาคต", future_headers, future_table_rows, total_fill=TOTAL_FILL)

    regeneration_headers = ["รายการ", "ความหนาแน่น", "หน่วย", "อัตรา", "มูลค่า (บาท)"]
    regeneration_rows = [
        ["ลูกไม้", regeneration_dict.get("sapling_density_per_rai"), "ต้น/ไร่", "27 บาท/ต้น/ไร่", regeneration_dict.get("sapling_loss_baht")],
        ["กล้าไม้", regeneration_dict.get("seedling_density_per_rai"), "ต้น/ไร่", "6 บาท/ต้น/ไร่", regeneration_dict.get("seedling_loss_baht")],
        ["รวม", None, "", "", regeneration_dict.get("total_regeneration_loss_baht")],
    ]
    current_row = _write_component_table(ws, current_row, "ความเสียหายลูกไม้และกล้าไม้", regeneration_headers, regeneration_rows, total_fill=TOTAL_FILL)

    ecosystem_headers = ["ผลกระทบ", "ปริมาณ", "หน่วยวัด", "ราคา/หน่วย", "หน่วยราคา", "มูลค่า", "หน่วยมูลค่า"]
    ecosystem_table_rows = [
        [
            row.get("impact_name_th"),
            row.get("quantity"),
            row.get("quantity_unit"),
            row.get("unit_price"),
            row.get("unit_price_unit"),
            row.get("value_baht_per_year"),
            row.get("value_unit"),
        ]
        for row in ecosystem_rows
    ]
    ecosystem_total = sum(float(row.get("value_baht_per_year")) for row in ecosystem_rows if isinstance(row.get("value_baht_per_year"), (int, float)))
    ecosystem_table_rows.append(["รวม", None, "", None, "", ecosystem_total if ecosystem_rows else None, "บาท/ปี"])
    current_row = _write_component_table(ws, current_row, "ผลกระทบระบบนิเวศ", ecosystem_headers, ecosystem_table_rows, total_fill=TOTAL_FILL)

    species_headers = ["forest_type", "tq", "species_name", "species_volume_m3", "wood_price_per_m3", "annual_wood_value_baht", "wood_value_baht", "price_status"]
    species_table_rows = []
    if not species_frame.empty:
        for _, row in species_frame.iterrows():
            species_table_rows.append([row.get(header) for header in species_headers])
    current_row = _write_component_table(ws, current_row, "รายละเอียดชนิดไม้", species_headers, species_table_rows)

    warning_rows = [row for row in _warning_rows(bundle) if row.get("component") == component_name]
    current_row = _write_component_table(
        ws,
        current_row,
        "Warnings",
        ["module", "component", "message"],
        [[row.get("module"), row.get("component"), row.get("message")] for row in warning_rows],
    )
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    _auto_fit_columns(ws, min_width=12, max_width=24)


def _audit_frames(bundle: dict[str, object]) -> dict[str, pd.DataFrame]:
    ecosystem_rows: list[dict[str, object]] = []
    for row in bundle.get("ecosystem_loss", {}).get("groupResults", []):
        ecosystem_rows.extend(build_ecosystem_loss_detail_rows(_dict_to_ecosystem_proxy(row)))
    future_rows: list[dict[str, object]] = []
    for summary in bundle.get("wood_future_value", {}).get("componentSummaries", []):
        for row in summary.get("period_rows", []):
            future_rows.append(
                {
                    "component_id": summary.get("component_id"),
                    "component_name": summary.get("component_name"),
                    **row,
                }
            )
    return {
        "AUDIT_FOREST_ECON": pd.DataFrame(bundle.get("forest_economics", {}).get("detailRows", [])),
        "AUDIT_SPECIES": pd.DataFrame(bundle.get("forest_economics", {}).get("speciesDetailRows", [])),
        "AUDIT_FUTURE_VALUE": pd.DataFrame(future_rows),
        "AUDIT_ECOSYSTEM": pd.DataFrame(ecosystem_rows),
    }


def write_forest_economic_report(
    report_file: str | Path,
    outputs: dict[str, pd.DataFrame],
    bundle: dict[str, object],
) -> Path:
    report_path = Path(report_file)
    workbook = Workbook()
    existing_names: set[str] = set()

    master_ws = workbook.active
    master_ws.title = safe_sheet_name("MASTER_SUMMARY", existing_names)
    _write_master_summary(master_ws, outputs, bundle)

    warnings_ws = workbook.create_sheet(title=safe_sheet_name("MASTER_WARNINGS", existing_names))
    warning_frame = pd.DataFrame(_warning_rows(bundle))
    _write_frame(warnings_ws, 1, "Warnings", warning_frame)
    warnings_ws.sheet_view.showGridLines = False
    _auto_fit_columns(warnings_ws)

    for index, (component_id, component_name) in enumerate(_component_order(outputs), start=1):
        sheet_name = safe_sheet_name(f"COMP_{index:02d}_{component_name}", existing_names)
        component_ws = workbook.create_sheet(title=sheet_name)
        _write_component_sheet(component_ws, outputs, bundle, component_id, component_name)

    for sheet_name, frame in _audit_frames(bundle).items():
        ws = workbook.create_sheet(title=safe_sheet_name(sheet_name, existing_names))
        _write_frame(ws, 1, sheet_name, frame)
        ws.sheet_view.showGridLines = False
        _auto_fit_columns(ws, min_width=12, max_width=24)

    workbook.save(report_path)
    return report_path
