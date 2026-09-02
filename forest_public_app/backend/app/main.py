from __future__ import annotations

import base64
import copy
import gc
import hashlib
import json
import logging
import os
import sys
import shutil
import tempfile
import threading
import time
import uuid
from collections import OrderedDict
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import numpy as np
import pandas as pd
from fastapi.encoders import jsonable_encoder
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from pydantic import BaseModel

ROOT_DIR = Path(__file__).resolve().parents[3]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))
WORKSPACE_DIR = ROOT_DIR if (ROOT_DIR / "cal_EIA").exists() else ROOT_DIR.parent

import run_forest_calculation as calc
from cal_EIA.generate_profile_realistic import render_editable_profile_scene, render_freeform_sprite_experiment
from cal_EIA.profile_diagram_lib import (
    create_profile_template,
    inspect_profile_workbook as inspect_profile_workbook_data,
    list_profile_sheets,
    load_profile_sheet,
    render_workbook_profile_map,
)
from cal_EIA.profile_validation import EXPECTED_COLUMNS
from forest_economic_report import (
    ECOSYSTEM_TOTAL_KEYS,
    _estimated_tree_count_from_density,
    _sum_ecosystem_detail,
    write_forest_economic_report,
)
from forest_ecosystem_loss import build_ecosystem_loss_detail_rows
from forest_integration import EcosystemUserInput, calculate_forest_valuation_bundle_from_outputs

TEMPLATE_FILE = ROOT_DIR / "template.xlsx"
MASTER_FILE = ROOT_DIR / "species_reference_master_v1.xlsx"
COMPONENT_TEMPLATE_FILE = ROOT_DIR / "forest_component_7.xlsx"
PROFILE_TEMPLATE_DIR = WORKSPACE_DIR / "cal_EIA" / "04_templates"
PROFILE_SOURCE_FILE = PROFILE_TEMPLATE_DIR / "profile.xlsx"
PROFILE_TEMPLATE_FILE = PROFILE_TEMPLATE_DIR / "profile_template.xlsx"
if not PROFILE_SOURCE_FILE.exists():
    PROFILE_SOURCE_FILE = WORKSPACE_DIR / "cal_EIA" / "profile.xlsx"
if not PROFILE_TEMPLATE_FILE.exists():
    PROFILE_TEMPLATE_FILE = WORKSPACE_DIR / "cal_EIA" / "profile_template.xlsx"
OUTPUT_BASE_FILENAME = "forest_calculation_output.xlsx"
SUMMARY_OUTPUT_FILENAME = "forest_summary.xlsx"
DETAIL_OUTPUT_FILENAME = "forest_details.xlsx"
COMPONENT_OUTPUT_FILENAME = "forest_components.xlsx"
PROFILE_OUTPUT_FILENAME = "profile_diagram_outputs.zip"
PROFILE_REALISTIC_OUTPUT_FILENAME = "profile_diagram_realistic_outputs.zip"
PROFILE_ASSET_ROOTS = [
    WORKSPACE_DIR / "cal_EIA" / "profile_assets_v2" / "normalized",
    WORKSPACE_DIR / "cal_EIA" / "profile_assets_v3" / "normalized",
]
PROFILE_ASSET_FILES = {"trunk": "trunk.png", "branch": "first_branch.png", "crown": "canopy_side.png"}
PROFILE_EDITOR_SCENES: OrderedDict[str, dict[str, Any]] = OrderedDict()
PROFILE_EDITOR_SCENE_LOCK = threading.Lock()
PROFILE_EDITOR_SCENE_LIMIT = 4
PROFILE_GENERATION_JOBS: OrderedDict[str, dict[str, Any]] = OrderedDict()
PROFILE_GENERATION_JOB_LOCK = threading.Lock()
PROFILE_GENERATION_RENDER_LOCK = threading.Lock()
PROFILE_GENERATION_JOB_LIMIT = 6
ECONOMIC_OUTPUT_FILENAME = "forest_economic_report.xlsx"
ECONOMIC_JSON_FILENAME = "forest_economic_report.json"
WORKFLOW_CACHE_TTL_SECONDS = int(os.getenv("WORKFLOW_CACHE_TTL_SECONDS", "3600"))
WORKFLOW_CACHE_MAX_ENTRIES = int(os.getenv("WORKFLOW_CACHE_MAX_ENTRIES", "8"))
LOG = logging.getLogger(__name__)

WorkflowCacheValue = tuple[bytes, bytes, bytes | None, dict[str, pd.DataFrame]]
WORKFLOW_CACHE: OrderedDict[str, tuple[float, WorkflowCacheValue]] = OrderedDict()
WORKFLOW_CACHE_LOCK = threading.RLock()


def parse_cors_origins() -> list[str]:
    raw_value = os.getenv("CORS_ORIGINS", "*").strip()
    if not raw_value or raw_value == "*":
        return ["*"]
    return [origin.strip() for origin in raw_value.split(",") if origin.strip()]


class MetricCard(BaseModel):
    label: str
    value: str
    help_text: str


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def format_metric_value(value: object, decimals: int = 2) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "-"
    if isinstance(value, (int, float)):
        return f"{value:,.{decimals}f}" if isinstance(value, float) or decimals else f"{int(value):,}"
    return str(value)


def sanitize_for_json(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): sanitize_for_json(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [sanitize_for_json(item) for item in value]
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if isinstance(value, np.generic):
        return sanitize_for_json(value.item())
    return value


def canonical_json(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def build_workflow_cache_key(
    file_bytes: bytes,
    plot_area_ha: float,
    rai_per_hectare: float,
    sheet_groups: list[dict[str, object]] | None,
) -> str:
    fingerprint = {
        "file_sha256": hashlib.sha256(file_bytes).hexdigest(),
        "plot_area_ha": float(plot_area_ha),
        "rai_per_hectare": float(rai_per_hectare),
        "sheet_groups": sheet_groups or [],
    }
    return hashlib.sha256(canonical_json(fingerprint).encode("utf-8")).hexdigest()


def copy_result_sheets(result_sheets: dict[str, Any]) -> dict[str, Any]:
    copied: dict[str, Any] = {}
    for sheet_name, frame in result_sheets.items():
        if isinstance(frame, pd.DataFrame):
            copied[sheet_name] = frame.copy(deep=True)
        else:
            copied[sheet_name] = copy.deepcopy(frame)
    return copied


def copy_workflow_cache_value(value: WorkflowCacheValue) -> WorkflowCacheValue:
    summary_bytes, detail_bytes, component_bytes, result_sheets = value
    return summary_bytes, detail_bytes, component_bytes, copy_result_sheets(result_sheets)


def get_cached_workflow(cache_key: str) -> WorkflowCacheValue | None:
    if WORKFLOW_CACHE_TTL_SECONDS <= 0 or WORKFLOW_CACHE_MAX_ENTRIES <= 0:
        return None
    now = time.monotonic()
    with WORKFLOW_CACHE_LOCK:
        cached = WORKFLOW_CACHE.get(cache_key)
        if cached is None:
            return None
        created_at, value = cached
        if now - created_at > WORKFLOW_CACHE_TTL_SECONDS:
            WORKFLOW_CACHE.pop(cache_key, None)
            return None
        WORKFLOW_CACHE.move_to_end(cache_key)
        return copy_workflow_cache_value(value)


def store_cached_workflow(cache_key: str, value: WorkflowCacheValue) -> None:
    if WORKFLOW_CACHE_TTL_SECONDS <= 0 or WORKFLOW_CACHE_MAX_ENTRIES <= 0:
        return
    with WORKFLOW_CACHE_LOCK:
        WORKFLOW_CACHE[cache_key] = (time.monotonic(), copy_workflow_cache_value(value))
        WORKFLOW_CACHE.move_to_end(cache_key)
        while len(WORKFLOW_CACHE) > WORKFLOW_CACHE_MAX_ENTRIES:
            WORKFLOW_CACHE.popitem(last=False)


def filter_primary_rows(frame: pd.DataFrame, result_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    if frame.empty:
        return frame
    component_names = calc.get_component_sheet_names(result_sheets)
    return calc.filter_out_component_rows(frame, component_names)


def count_unmatched_species(unmatched: pd.DataFrame) -> int:
    if unmatched.empty:
        return 0

    for column_name in ("Species_norm", "Species_raw"):
        if column_name in unmatched.columns:
            series = unmatched[column_name].astype(str).str.strip().replace("", pd.NA).dropna()
            if not series.empty:
                return int(series.nunique())

    return int(len(unmatched.index))


def get_numeric_series(frame: pd.DataFrame, column_name: str) -> pd.Series:
    if frame.empty or column_name not in frame.columns:
        return pd.Series(dtype=float)
    return pd.to_numeric(frame[column_name], errors="coerce").fillna(0)


def build_metrics(
    summary_all: pd.DataFrame,
    unmatched: pd.DataFrame,
    result_sheets: dict[str, pd.DataFrame],
) -> list[MetricCard]:
    if summary_all.empty:
        return []

    filtered_summary = filter_primary_rows(summary_all, result_sheets)
    filtered_unmatched = filter_primary_rows(unmatched, result_sheets)
    if filtered_summary.empty:
        return []

    total_tree = get_numeric_series(filtered_summary, "n_tree").sum()
    total_sapling = get_numeric_series(filtered_summary, "n_sapling").sum()
    total_tree_biomass = get_numeric_series(filtered_summary, "total_tree_biomass").sum()
    total_tree_volume = get_numeric_series(filtered_summary, "total_tree_volume_m3").sum()
    total_sapling_volume = get_numeric_series(filtered_summary, "total_sapling_volume_m3").sum()

    shannon_series = get_numeric_series(filtered_summary, "shannon_index")
    shannon_value = shannon_series.mean() if not shannon_series.empty else None
    unmatched_count = count_unmatched_species(filtered_unmatched)

    return [
        MetricCard(label="Total tree count", value=format_metric_value(total_tree, 0), help_text="Across all processed worksheets"),
        MetricCard(label="Total sapling count", value=format_metric_value(total_sapling, 0), help_text="Sapling records included in volume"),
        MetricCard(label="Total tree biomass", value=format_metric_value(total_tree_biomass, 2), help_text="Tree biomass only"),
        MetricCard(label="Total tree volume", value=format_metric_value(total_tree_volume, 3), help_text="Tree block volume"),
        MetricCard(label="Total sapling volume", value=format_metric_value(total_sapling_volume, 3), help_text="Sapling block volume"),
        MetricCard(label="Shannon index", value=format_metric_value(shannon_value, 6), help_text="Average across available sites"),
        MetricCard(label="Unmatched species", value=format_metric_value(unmatched_count, 0), help_text="Unique unmatched species still reviewed in the QA sheet"),
    ]


def dataframe_records(frame: pd.DataFrame, limit: int = 250) -> list[dict[str, Any]]:
    if frame.empty:
        return []
    sample = frame.head(limit).copy()
    sample = sample.where(pd.notna(sample), None)
    return jsonable_encoder(sanitize_for_json(sample.to_dict(orient="records")))


def build_biomass_payload(
    result_sheets: dict[str, pd.DataFrame],
    plot_area_ha: float,
    rai_per_hectare: float,
    sheet_groups: list[dict[str, object]] | None,
) -> dict[str, Any]:
    summary_all = result_sheets.get("SUMMARY_ALL", pd.DataFrame())
    summary_biomass = result_sheets.get("SUMMARY_BIOMASS", pd.DataFrame())
    summary_volume = result_sheets.get("SUMMARY_VOLUME", pd.DataFrame())
    summary_shannon = result_sheets.get("SUMMARY_SHANNON", pd.DataFrame())
    unmatched = result_sheets.get("CHECK_UNMATCHED_SPECIES", pd.DataFrame())

    try:
        metrics = [metric.model_dump() for metric in build_metrics(summary_all, unmatched, result_sheets)]
    except Exception:  # noqa: BLE001
        LOG.exception("Failed to build biomass metrics.")
        metrics = []

    try:
        component_summaries = build_component_biomass_summary(
            result_sheets=result_sheets,
            sheet_groups=sheet_groups,
            plot_area_ha=plot_area_ha,
            rai_per_hectare=rai_per_hectare,
        )
    except Exception:  # noqa: BLE001
        LOG.exception("Failed to build biomass component summaries.")
        component_summaries = []

    previews: dict[str, list[dict[str, Any]]] = {}
    preview_frames = {
        "summaryAll": summary_all,
        "summaryBiomass": summary_biomass,
        "summaryVolume": summary_volume,
        "summaryShannon": summary_shannon,
        "unmatchedSpecies": unmatched,
    }
    for preview_key, frame in preview_frames.items():
        try:
            previews[preview_key] = dataframe_records(frame)
        except Exception:  # noqa: BLE001
            LOG.exception("Failed to serialize biomass preview '%s'.", preview_key)
            previews[preview_key] = []

    return {
        "metrics": metrics,
        "componentSummaries": component_summaries,
        "previews": previews,
    }


def serialize_download_payload(filename: str, content: bytes) -> dict[str, str]:
    return {
        "filename": filename,
        "contentBase64": base64.b64encode(content).decode("ascii"),
    }


def get_uploaded_sheet_names(file_bytes: bytes) -> list[str]:
    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def ensure_profile_template() -> Path:
    if not PROFILE_SOURCE_FILE.exists():
        raise HTTPException(status_code=500, detail="Profile source workbook is missing.")
    if not PROFILE_TEMPLATE_FILE.exists():
        create_profile_template(PROFILE_SOURCE_FILE, PROFILE_TEMPLATE_FILE)
    return PROFILE_TEMPLATE_FILE


def profile_asset_groups() -> list[dict[str, object]]:
    groups: dict[str, dict[str, object]] = {}
    for root in PROFILE_ASSET_ROOTS:
        if not root.is_dir():
            continue
        for group_dir in sorted(root.iterdir()):
            if not group_dir.is_dir() or group_dir.name in groups:
                continue
            if not all((group_dir / filename).is_file() for filename in PROFILE_ASSET_FILES.values()):
                continue
            groups[group_dir.name] = {
                "id": group_dir.name,
                "files": {
                    part: f"/api/profile/assets/{group_dir.name}/{filename}"
                    for part, filename in PROFILE_ASSET_FILES.items()
                },
            }
    return [groups[group_id] for group_id in sorted(groups)]


def find_profile_asset(group_id: str, asset_name: str) -> Path:
    if asset_name not in PROFILE_ASSET_FILES.values() or Path(asset_name).name != asset_name:
        raise HTTPException(status_code=404, detail="Profile asset is not available.")
    for root in PROFILE_ASSET_ROOTS:
        candidate = root / group_id / asset_name
        if candidate.is_file():
            return candidate
    raise HTTPException(status_code=404, detail="Profile asset group is not available.")


def build_profile_editor_inspection(excel_path: Path) -> dict[str, object]:
    inspection = inspect_profile_workbook_data(excel_path)
    detailed_sheets: list[dict[str, object]] = []
    for sheet in inspection["sheets"]:
        item = dict(sheet)
        if sheet["valid"]:
            dataframe = load_profile_sheet(excel_path, str(sheet["sheetName"]))
            trees: list[dict[str, object]] = []
            for row_index, row in enumerate(dataframe.to_dict(orient="records"), start=3):
                tree = {str(key): value for key, value in row.items()}
                tree["id"] = f"{sheet['sheetName']}:{tree.get('no')}:{row_index}"
                tree["row"] = row_index
                trees.append(sanitize_for_json(tree))
            item["trees"] = trees
        detailed_sheets.append(item)
    return {
        **inspection,
        "sheets": detailed_sheets,
        "assetGroups": profile_asset_groups(),
    }


def apply_profile_editor_rows(file_bytes: bytes, sheet_name: str, rows: list[dict[str, object]]) -> bytes:
    workbook = load_workbook(filename=BytesIO(file_bytes))
    try:
        if sheet_name not in workbook.sheetnames:
            raise HTTPException(status_code=400, detail=f"Profile worksheet '{sheet_name}' was not found.")
        worksheet = workbook[sheet_name]
        for fallback_row, row in enumerate(rows, start=3):
            try:
                excel_row = int(row.get("row", fallback_row))
            except (TypeError, ValueError) as exc:
                raise HTTPException(status_code=400, detail="Each editor tree must have a valid workbook row.") from exc
            if excel_row < 3:
                raise HTTPException(status_code=400, detail="Editor rows must start at workbook row 3.")
            for column_index, column_name in enumerate(EXPECTED_COLUMNS, start=1):
                if column_name in row:
                    worksheet.cell(excel_row, column_index).value = row[column_name]
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
    finally:
        workbook.close()


def normalize_profile_render_mode(value: str | None) -> str:
    mode = normalize_text(value).lower() or "graphic"
    if mode == "illustrate":
        return "realistic"
    if mode not in {"graphic", "realistic"}:
        raise HTTPException(status_code=400, detail="render_mode must be 'graphic', 'realistic', or 'illustrate'.")
    return mode


def parse_profile_editor_json(raw_value: str | None, field_name: str) -> dict[str, Any]:
    if not raw_value:
        return {}
    try:
        parsed = json.loads(raw_value)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid {field_name} JSON: {exc}") from exc
    if not isinstance(parsed, dict):
        raise HTTPException(status_code=400, detail=f"{field_name} must be a JSON object.")
    return parsed


def sanitize_profile_editor_transform(value: object) -> dict[str, float | None]:
    if not isinstance(value, dict):
        return {}
    allowed = {"dx", "dy", "scale", "rotate", "widthScale", "heightScale", "crownWidthScale", "crownHeightScale", "bendYRatio", "bendXRatio"}
    sanitized: dict[str, float | None] = {}
    for key in allowed:
        raw = value.get(key)
        if raw is None and key == "bendYRatio":
            sanitized[key] = None
            continue
        if isinstance(raw, bool) or raw is None:
            continue
        try:
            numeric = float(raw)
        except (TypeError, ValueError):
            continue
        if np.isfinite(numeric):
            sanitized[key] = numeric
    return sanitized


def build_profile_editor_render_overrides(
    sheet_name: str,
    parsed_trees: list[dict[str, object]],
    assignments: dict[str, Any],
    transforms: dict[str, Any],
) -> tuple[dict[int, str], dict[int, dict[str, float | None]], dict[int, dict[str, float | None]]]:
    """Convert V3's tree ids into the row-index maps used by the PNG renderer."""
    valid_group_ids = {str(group["id"]) for group in profile_asset_groups()}
    asset_by_index: dict[int, str] = {}
    transform_by_index: dict[int, dict[str, float | None]] = {}
    top_by_index: dict[int, dict[str, float | None]] = {}

    for fallback_row, tree in enumerate(parsed_trees, start=3):
        try:
            workbook_row = int(tree.get("row", fallback_row))
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail="Each editor tree must have a valid workbook row.") from exc
        row_index = workbook_row - 3
        tree_id = str(tree.get("id") or f"{sheet_name}:{tree.get('no')}:{workbook_row}")

        group_id = assignments.get(tree_id)
        if group_id is not None:
            if not isinstance(group_id, str) or group_id not in valid_group_ids:
                raise HTTPException(status_code=400, detail=f"Unknown profile asset group for tree '{tree_id}'.")
            asset_by_index[row_index] = group_id

        crown = sanitize_profile_editor_transform(transforms.get(f"{tree_id}:crown"))
        trunk = sanitize_profile_editor_transform(transforms.get(f"{tree_id}:trunk"))
        branch = sanitize_profile_editor_transform(transforms.get(f"{tree_id}:branch"))
        merged: dict[str, float | None] = {}
        # Crown scale/position is the tree's primary placement in the export;
        # trunk width/bend and branch rotation are then carried across as V3 edits.
        for key in ("dx", "dy", "scale"):
            if key in crown:
                merged[key] = crown[key]
        if "widthScale" in crown:
            merged["crownWidthScale"] = crown["widthScale"]
        if "heightScale" in crown:
            merged["crownHeightScale"] = crown["heightScale"]
        for key in ("widthScale", "bendXRatio", "bendYRatio"):
            if key in trunk:
                merged[key] = trunk[key]
        if "rotate" in branch:
            merged["rotate"] = branch["rotate"]
        if merged:
            transform_by_index[row_index] = merged

        top = sanitize_profile_editor_transform(transforms.get(f"top:{tree_id}"))
        if top:
            top_by_index[row_index] = top

    return asset_by_index, transform_by_index, top_by_index


def build_profile_outputs(
    uploaded_filename: str,
    file_bytes: bytes,
    render_mode: str = "graphic",
    editor_sheet_name: str | None = None,
    editor_overrides: tuple[dict[int, str], dict[int, dict[str, float | None]], dict[int, dict[str, float | None]]] | None = None,
) -> tuple[list[dict[str, str]], bytes, str]:
    render_mode = normalize_profile_render_mode(render_mode)
    with tempfile.TemporaryDirectory() as tmp_dir:
        temp_dir = Path(tmp_dir)
        uploaded_path = temp_dir / (Path(uploaded_filename).name or "profile.xlsx")
        uploaded_path.write_bytes(file_bytes)

        output_dir = temp_dir / "profile_images"
        if render_mode == "realistic":
            rendered_items = [
                (
                    sheet_name,
                    render_freeform_sprite_experiment(
                        uploaded_path,
                        sheet_name,
                        output_dir,
                        asset_assignments=editor_overrides[0] if editor_sheet_name == sheet_name and editor_overrides else None,
                        tree_transform_map=editor_overrides[1] if editor_sheet_name == sheet_name and editor_overrides else None,
                        top_transform_map=editor_overrides[2] if editor_sheet_name == sheet_name and editor_overrides else None,
                        asset_roots=PROFILE_ASSET_ROOTS if editor_sheet_name == sheet_name and editor_overrides else None,
                    ),
                )
                for sheet_name in list_profile_sheets(uploaded_path)
            ]
            output_filename = PROFILE_REALISTIC_OUTPUT_FILENAME
        else:
            rendered_items = render_workbook_profile_map(uploaded_path, output_dir)
            output_filename = PROFILE_OUTPUT_FILENAME
        payloads: list[dict[str, str]] = []
        image_paths: list[Path] = []
        for sheet_name, image_path in rendered_items:
            image_paths.append(image_path)
            payloads.append(
                {
                    "sheetName": sheet_name,
                    "filename": image_path.name,
                    "contentBase64": base64.b64encode(image_path.read_bytes()).decode("ascii"),
                }
            )

        zip_path = temp_dir / output_filename
        with ZipFile(zip_path, "w", compression=ZIP_DEFLATED) as zip_file:
            for image_path in image_paths:
                zip_file.write(image_path, arcname=image_path.name)
        return payloads, zip_path.read_bytes(), output_filename


def build_profile_output_files(
    uploaded_filename: str,
    file_bytes: bytes,
    render_mode: str,
    job_root: Path,
) -> tuple[dict[str, Any], dict[str, tuple[Path, str]]]:
    """Build the production job assets so the existing editor can consume them."""
    uploaded_path = job_root / (Path(uploaded_filename).name or "profile.xlsx")
    uploaded_path.write_bytes(file_bytes)
    output_dir = job_root / "profile_images"
    normalized_mode = normalize_profile_render_mode(render_mode)
    sheet_names = list_profile_sheets(uploaded_path)
    inspection = inspect_profile_workbook_data(uploaded_path)
    validation_by_name = {str(item["sheetName"]): item for item in inspection["sheets"]}
    validations: list[dict[str, Any]] = []
    for sheet_name in sheet_names:
        dataframe = load_profile_sheet(uploaded_path, sheet_name)
        validations.append({
            "sheetName": sheet_name,
            "treeCount": int(len(dataframe.index)),
            "speciesCount": int(dataframe["species"].nunique()),
            "species": [str(value) for value in dataframe["species"].drop_duplicates().tolist()],
            "valid": bool(validation_by_name.get(sheet_name, {}).get("valid", True)),
        })

    if normalized_mode == "realistic":
        rendered_items = [
            (sheet_name, render_freeform_sprite_experiment(uploaded_path, sheet_name, output_dir))
            for sheet_name in sheet_names
        ]
        output_filename = PROFILE_REALISTIC_OUTPUT_FILENAME
    else:
        rendered_items = render_workbook_profile_map(uploaded_path, output_dir)
        output_filename = PROFILE_OUTPUT_FILENAME

    result_assets: dict[str, tuple[Path, str]] = {}
    image_manifest: list[dict[str, str]] = []
    for image_index, (sheet_name, image_path) in enumerate(rendered_items):
        asset_key = f"profile-image-{image_index}.png"
        result_assets[asset_key] = (image_path, "image/png")
        image_manifest.append({"sheetName": sheet_name, "filename": image_path.name, "file": asset_key})
    zip_path = job_root / output_filename
    with ZipFile(zip_path, "w", compression=ZIP_DEFLATED) as zip_file:
        for _, image_path in rendered_items:
            zip_file.write(image_path, arcname=image_path.name)
    result_assets["profile-output.zip"] = (zip_path, "application/zip")
    uploaded_path.unlink(missing_ok=True)
    return {
        "sheetNames": sheet_names,
        "renderMode": normalized_mode,
        "images": image_manifest,
        "validation": validations,
        "download": {"filename": output_filename, "file": "profile-output.zip"},
    }, result_assets


def build_profile_editor_scene_manifest(file_name: str, file_bytes: bytes) -> dict[str, Any]:
    session_id = uuid.uuid4().hex
    scene_root = Path(tempfile.mkdtemp(prefix=f"profile-editor-{session_id}-"))
    assets: dict[str, Path] = {}
    manifest_sheets: list[dict[str, Any]] = []
    try:
        uploaded_path = scene_root / (Path(file_name).name or "profile.xlsx")
        uploaded_path.write_bytes(file_bytes)
        for sheet_index, sheet_name in enumerate(list_profile_sheets(uploaded_path)):
            rendered = render_editable_profile_scene(uploaded_path, sheet_name, scene_root / "editor_scene")
            slug = f"sheet-{sheet_index + 1}"
            base_key = f"{slug}/base.png"
            assets[base_key] = Path(rendered["basePath"])
            manifest_trees: list[dict[str, Any]] = []
            for tree in rendered["trees"]:
                tree_parts: dict[str, Any] = {}
                for part_name, part in tree["parts"].items():
                    asset_key = f"{slug}/tree-{tree['id']}-{part_name}.png"
                    assets[asset_key] = Path(part["path"])
                    asset_group = tree.get("assetGroup")
                    source_asset_name = PROFILE_ASSET_FILES.get(part_name)
                    tree_parts[part_name] = {
                        # Keep the editable canvas on the original high-resolution
                        # group asset. The captured layer only supplies the exact
                        # measured placement box; it should not become the source
                        # bitmap because that would pixelate when scaled in V3.
                        "file": (
                            f"/api/profile/assets/{asset_group}/{source_asset_name}"
                            if asset_group and source_asset_name
                            else f"/api/profile/editor-scene/{session_id}/asset/{asset_key}"
                        ),
                        "x": part["x"], "y": part["y"], "w": part["w"], "h": part["h"],
                    }
                manifest_trees.append({
                    "id": tree["id"],
                    "species": tree["species"],
                    "assetGroup": tree.get("assetGroup"),
                    "parts": tree_parts,
                })
            manifest_sheets.append({
                "name": sheet_name,
                "slug": slug,
                "base": f"/api/profile/editor-scene/{session_id}/asset/{base_key}",
                "width": rendered["width"],
                "height": rendered["height"],
                "trees": manifest_trees,
            })
            del rendered
            gc.collect()
        with PROFILE_EDITOR_SCENE_LOCK:
            PROFILE_EDITOR_SCENES[session_id] = {
                "root": scene_root,
                "assets": assets,
                "manifest": {
                    "sessionId": session_id,
                    "fileName": file_name,
                    "sheets": manifest_sheets,
                },
            }
            while len(PROFILE_EDITOR_SCENES) > PROFILE_EDITOR_SCENE_LIMIT:
                _, expired_scene = PROFILE_EDITOR_SCENES.popitem(last=False)
                shutil.rmtree(expired_scene["root"], ignore_errors=True)
    except Exception:
        shutil.rmtree(scene_root, ignore_errors=True)
        raise
    return PROFILE_EDITOR_SCENES[session_id]["manifest"]


def run_profile_generation_job(job_id: str, file_name: str, file_bytes: bytes, render_mode: str) -> None:
    job_root = Path(tempfile.mkdtemp(prefix=f"profile-job-{job_id}-"))
    try:
        with PROFILE_GENERATION_RENDER_LOCK:
            with PROFILE_GENERATION_JOB_LOCK:
                PROFILE_GENERATION_JOBS[job_id]["status"] = "rendering"
            profile_manifest, result_assets = build_profile_output_files(file_name, file_bytes, render_mode, job_root)
            for image in profile_manifest["images"]:
                image["file"] = f"/api/profile/generation-job/{job_id}/asset/{image['file']}"
            profile_manifest["download"]["file"] = f"/api/profile/generation-job/{job_id}/asset/{profile_manifest['download']['file']}"
            with PROFILE_GENERATION_JOB_LOCK:
                PROFILE_GENERATION_JOBS[job_id]["status"] = "preparing_editor"
            editor_scene = build_profile_editor_scene_manifest(file_name, file_bytes)
        with PROFILE_GENERATION_JOB_LOCK:
            PROFILE_GENERATION_JOBS[job_id].update({
                "status": "ready",
                "profile": profile_manifest,
                "editorScene": editor_scene,
                "resultAssets": result_assets,
                "jobRoot": job_root,
                "finishedAt": time.time(),
            })
    except Exception as exc:  # noqa: BLE001
        shutil.rmtree(job_root, ignore_errors=True)
        LOG.exception("Profile generation job %s failed.", job_id)
        with PROFILE_GENERATION_JOB_LOCK:
            PROFILE_GENERATION_JOBS[job_id].update({"status": "failed", "detail": str(exc), "finishedAt": time.time()})


def parse_sheet_groups(sheet_groups_raw: str | None) -> list[dict[str, object]] | None:
    if not sheet_groups_raw:
        return None
    try:
        parsed = json.loads(sheet_groups_raw)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid sheet_groups JSON: {exc}") from exc
    if not isinstance(parsed, list):
        raise HTTPException(status_code=400, detail="sheet_groups must be a JSON array.")
    return parsed


def parse_future_periods(raw_value: str | None) -> list[int] | None:
    if not raw_value:
        return None
    try:
        parsed = json.loads(raw_value)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid future_periods JSON: {exc}") from exc
    if not isinstance(parsed, list) or not all(isinstance(item, int) for item in parsed):
        raise HTTPException(status_code=400, detail="future_periods must be a JSON array of integers.")
    return parsed


def parse_economic_inputs(
    economic_inputs_raw: str | None,
    sheet_groups: list[dict[str, object]] | None,
) -> tuple[dict[str, float], list[EcosystemUserInput]]:
    if not economic_inputs_raw:
        return {}, []
    try:
        parsed = json.loads(economic_inputs_raw)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid economic_inputs JSON: {exc}") from exc
    if not isinstance(parsed, list):
        raise HTTPException(status_code=400, detail="economic_inputs must be a JSON array.")

    component_area_inputs: dict[str, float] = {}
    ecosystem_inputs: list[EcosystemUserInput] = []
    group_names = {normalize_text(group.get("name")) for group in sheet_groups or [] if normalize_text(group.get("name"))}
    for item in parsed:
        if not isinstance(item, dict):
            raise HTTPException(status_code=400, detail="Each economic input must be an object.")
        component_name = normalize_text(item.get("component_name"))
        if not component_name:
            raise HTTPException(status_code=400, detail="Each economic input must include component_name.")
        if group_names and component_name not in group_names:
            raise HTTPException(status_code=400, detail=f"Economic input component '{component_name}' is not in grouped components.")
        try:
            component_area_rai = float(item.get("component_area_rai"))
            canopy_cover_percent = float(item.get("canopy_cover_percent"))
            canopy_layer_count = float(item.get("canopy_layer_count"))
            soil_depth_m = float(item.get("soil_depth_m"))
            annual_rainfall_mm = float(item.get("annual_rainfall_mm"))
            topography_score = float(item.get("topography_score"))
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail=f"Invalid numeric economic input for component '{component_name}'.") from exc
        if component_area_rai <= 0 or not 0 <= canopy_cover_percent <= 100 or canopy_layer_count <= 0 or soil_depth_m <= 0 or annual_rainfall_mm < 0 or topography_score <= 0:
            raise HTTPException(status_code=400, detail=f"Economic input values are out of range for component '{component_name}'.")
        component_area_inputs[component_name] = component_area_rai
        ecosystem_inputs.append(
            EcosystemUserInput(
                component_name=component_name,
                component_area_rai=component_area_rai,
                canopy_cover_percent=canopy_cover_percent,
                canopy_layer_count=canopy_layer_count,
                soil_depth_m=soil_depth_m,
                annual_rainfall_mm=annual_rainfall_mm,
                topography_score=topography_score,
            )
        )

    missing_groups = sorted(group_names - set(component_area_inputs.keys()))
    if missing_groups:
        raise HTTPException(
            status_code=400,
            detail=f"Missing economic inputs for grouped component(s): {', '.join(missing_groups)}",
        )
    return component_area_inputs, ecosystem_inputs


def build_component_biomass_summary(
    result_sheets: dict[str, pd.DataFrame],
    sheet_groups: list[dict[str, object]] | None,
    plot_area_ha: float,
    rai_per_hectare: float,
) -> list[dict[str, Any]]:
    summary_all = result_sheets.get("SUMMARY_ALL", pd.DataFrame())
    detail_tree = result_sheets.get("DETAIL_TREE_BIOMASS", pd.DataFrame())
    if summary_all.empty or "sheet_name" not in summary_all.columns:
        return []
    component_display_map = calc.get_component_display_name_map(result_sheets)
    component_names = calc.get_component_group_names_in_order(result_sheets)
    if sheet_groups:
        ordered_names = component_names
    else:
        ordered_names = filter_primary_rows(summary_all, result_sheets).get("sheet_name", pd.Series(dtype=object)).dropna().astype(str).tolist()

    rows: list[dict[str, Any]] = []
    for component_name in ordered_names:
        matching_summary = summary_all[summary_all["sheet_name"].astype(str) == component_name]
        if matching_summary.empty:
            continue
        summary_row = matching_summary.iloc[0]
        matching_tree = detail_tree[detail_tree["sheet_name"].astype(str) == component_name] if not detail_tree.empty else pd.DataFrame()
        forest_types = sorted(
            {
                normalize_text(value)
                for value in matching_tree.get("forest_type_clean", pd.Series(dtype=object)).dropna().tolist()
                if normalize_text(value)
            }
        )
        plot_count = (
            matching_tree["Plot"].astype(str).str.strip().replace("", pd.NA).dropna().nunique()
            if not matching_tree.empty and "Plot" in matching_tree.columns
            else 0
        )
        rows.append(
            {
                "componentName": component_display_map.get(component_name, component_name),
                "internalName": component_name,
                "includedSheets": next(
                    (
                        [normalize_text(sheet_name) for sheet_name in group.get("sheet_names", []) if normalize_text(sheet_name)]
                        for group in sheet_groups or []
                        if normalize_text(group.get("internal_name")) == component_name or normalize_text(group.get("name")) == component_display_map.get(component_name, component_name)
                    ),
                    [component_name],
                ),
                "forestTypes": forest_types,
                "plotCount": plot_count,
                "sampleAreaRai": plot_count * plot_area_ha * rai_per_hectare if plot_count else None,
                "totalBiomass": summary_row.get("total_tree_biomass"),
                "totalWoodVolume": summary_row.get("total_tree_volume_m3"),
                "treeCount": summary_row.get("n_tree"),
                "saplingCount": summary_row.get("n_sapling"),
                "shannonIndex": summary_row.get("shannon_index"),
            }
        )
    return rows


def build_economic_preview(bundle: dict[str, object], outputs: dict[str, pd.DataFrame]) -> dict[str, Any]:
    component_rows: list[dict[str, Any]] = []
    ecosystem_component_lookup = {
        normalize_text(row.get("component_id")): row
        for row in bundle.get("ecosystem_loss", {}).get("componentSummaries", [])
        if normalize_text(row.get("component_id"))
    }
    regeneration_lookup = {
        normalize_text(row.get("component_id")): row
        for row in bundle.get("regeneration_loss", {}).get("componentSummaries", [])
        if normalize_text(row.get("component_id"))
    }
    for row in bundle.get("forest_economics", {}).get("componentSummaries", []):
        component_id = normalize_text(row.get("component_id"))
        eco_row = ecosystem_component_lookup.get(component_id, {})
        regen_row = regeneration_lookup.get(component_id, {})
        report_ecosystem_total = sum(
            float(value)
            for value in (_sum_ecosystem_detail(bundle, component_id, impact_key) for impact_key in ECOSYSTEM_TOTAL_KEYS)
            if isinstance(value, (int, float))
        )
        report_total_loss_values = [
            row.get("total_wood_value_baht"),
            regen_row.get("sapling_loss_baht"),
            regen_row.get("seedling_loss_baht"),
            report_ecosystem_total,
        ]
        report_total_loss = sum(float(value) for value in report_total_loss_values if isinstance(value, (int, float)))
        component_rows.append(
            {
                "componentId": component_id,
                "componentName": row.get("component_name"),
                "componentAreaRai": row.get("component_area_rai"),
                "estimatedTreeCount": _estimated_tree_count_from_density(outputs, component_id, row.get("component_area_rai")),
                "estimatedSaplingCount": regen_row.get("sapling_estimated_count"),
                "estimatedSeedlingCount": regen_row.get("seedling_estimated_count"),
                "forestTypes": row.get("forest_types_detected", []),
                "tqs": row.get("tq_detected", []),
                "totalWoodLossM3": row.get("total_wood_loss_m3"),
                "totalAnnualIncrementM3PerYear": row.get("total_annual_increment_m3_per_year"),
                "totalAnnualWoodValueBaht": row.get("total_annual_wood_value_baht"),
                "totalWoodValueBaht": row.get("total_wood_value_baht"),
                "totalRegenerationLossBaht": regen_row.get("total_regeneration_loss_baht"),
                "totalEcosystemLossBahtPerYear": report_ecosystem_total,
                "moduleEcosystemLossBahtPerYear": eco_row.get("total_ecosystem_loss_baht_per_year"),
                "totalReportLossBaht": report_total_loss,
                "warnings": row.get("warnings", []),
            }
        )

    impact_rows: list[dict[str, Any]] = []
    for row in bundle.get("ecosystem_loss", {}).get("groupResults", []):
        proxy = type("EcosystemProxy", (), row)
        for detail in build_ecosystem_loss_detail_rows(proxy):
            impact_rows.append(
                {
                    "componentId": row.get("component_id"),
                    "componentName": row.get("component_name"),
                    "forestType": row.get("forest_type"),
                    "impactKey": detail.get("impact_key"),
                    "impactNameTh": detail.get("impact_name_th"),
                    "quantity": detail.get("quantity"),
                    "quantityUnit": detail.get("quantity_unit"),
                    "unitPrice": detail.get("unit_price"),
                    "unitPriceUnit": detail.get("unit_price_unit"),
                    "valueBahtPerRaiPerYear": detail.get("value_baht_per_rai_per_year"),
                }
            )

    grand_total = bundle.get("forest_economics", {}).get("grandTotal", {})
    economic_metrics = [
        MetricCard(label="Economic components", value=format_metric_value(len(component_rows), 0), help_text="Grouped components included in the economic run"),
        MetricCard(label="Total wood loss", value=format_metric_value(grand_total.get("total_wood_loss_m3"), 3), help_text="Combined wood stock loss across grouped components"),
        MetricCard(
            label="Total loss in report",
            value=format_metric_value(
                sum(
                    float(row.get("totalReportLossBaht"))
                    for row in component_rows
                    if isinstance(row.get("totalReportLossBaht"), (int, float))
                ),
                2,
            ),
            help_text="Matches the MASTER_SUMMARY total row after scaling survey density and TQ volume per rai to project area",
        ),
        MetricCard(
            label="Ecosystem loss / year",
            value=format_metric_value(
                sum(
                    float(row.get("totalEcosystemLossBahtPerYear"))
                    for row in component_rows
                    if isinstance(row.get("totalEcosystemLossBahtPerYear"), (int, float))
                ),
                2,
            ),
            help_text="Sum of ecosystem loss totals from all grouped components",
        ),
    ]

    return {
        "metrics": [metric.model_dump() for metric in economic_metrics],
        "componentSummaries": component_rows,
        "woodDetails": bundle.get("forest_economics", {}).get("detailRows", []),
        "ecosystemSummaries": bundle.get("ecosystem_loss", {}).get("componentSummaries", []),
        "ecosystemImpactDetails": impact_rows,
        "futureValueRows": bundle.get("wood_future_value", {}).get("periodRows", []),
        "warnings": bundle.get("warnings", []),
    }


def run_uploaded_workflow(
    uploaded_filename: str,
    file_bytes: bytes,
    plot_area_ha: float,
    rai_per_hectare: float,
    sheet_groups: list[dict[str, object]] | None = None,
) -> tuple[bytes, bytes, bytes | None, dict[str, pd.DataFrame]]:
    with tempfile.TemporaryDirectory() as tmp_dir:
        temp_dir = Path(tmp_dir)
        # Keep the uploaded workbook path deterministic. The browser-provided
        # filename is metadata only and must not control Excel engine detection.
        uploaded_path = temp_dir / "uploaded_workbook.xlsx"
        uploaded_path.write_bytes(file_bytes)

        output_base = temp_dir / OUTPUT_BASE_FILENAME
        split_runner = getattr(calc, "run_calculation_split_outputs", None)
        if split_runner is not None:
            summary_path, detail_path, result_sheets = split_runner(
                input_file=uploaded_path,
                master_file=MASTER_FILE,
                output_base=output_base,
                plot_area_ha=plot_area_ha,
                rai_per_hectare=rai_per_hectare,
                sheet_groups=sheet_groups,
            )
        else:
            result_sheets = calc.process_workbook(
                input_file=uploaded_path,
                master_file=MASTER_FILE,
                plot_area_ha=plot_area_ha,
                rai_per_hectare=rai_per_hectare,
                sheet_groups=sheet_groups,
            )
            summary_path, detail_path = calc.resolve_output_paths(uploaded_path, str(output_base))
            calc.write_summary_by_site_workbook(summary_path, result_sheets)
            calc.write_detail_workbook(detail_path, result_sheets)

        component_bytes = None
        if sheet_groups and COMPONENT_TEMPLATE_FILE.exists():
            component_path = temp_dir / COMPONENT_OUTPUT_FILENAME
            calc.write_component_summary_workbook(
                component_path,
                COMPONENT_TEMPLATE_FILE,
                result_sheets,
                summary_file=summary_path,
            )
            component_bytes = component_path.read_bytes()

        return summary_path.read_bytes(), detail_path.read_bytes(), component_bytes, result_sheets


app = FastAPI(title="Forest Public App API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=parse_cors_origins(),
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception) -> JSONResponse:
    LOG.exception("Unhandled exception for %s %s", request.method, request.url.path)
    return JSONResponse(
        status_code=500,
        content={
            "detail": str(exc) or exc.__class__.__name__,
            "errorType": exc.__class__.__name__,
            "path": request.url.path,
        },
    )


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/config")
def config() -> dict[str, Any]:
    return {
        "plotAreaHa": calc.PLOT_AREA_HA,
        "raiPerHectare": calc.RAI_PER_HECTARE,
        "templateAvailable": TEMPLATE_FILE.exists(),
        "profileTemplateAvailable": PROFILE_SOURCE_FILE.exists(),
        "masterAvailable": MASTER_FILE.exists(),
        "componentTemplateAvailable": COMPONENT_TEMPLATE_FILE.exists(),
    }


@app.get("/api/template")
def template_download() -> FileResponse:
    if not TEMPLATE_FILE.exists():
        raise HTTPException(status_code=404, detail="Template file is missing.")
    return FileResponse(TEMPLATE_FILE, filename=TEMPLATE_FILE.name)


@app.get("/api/profile/template")
def profile_template_download() -> FileResponse:
    template_path = ensure_profile_template()
    return FileResponse(template_path, filename=template_path.name)


@app.get("/api/profile/assets/{group_id}/{asset_name}")
def profile_asset_download(group_id: str, asset_name: str) -> FileResponse:
    asset_path = find_profile_asset(group_id, asset_name)
    return FileResponse(asset_path, media_type="image/png", filename=asset_path.name)


@app.post("/api/inspect")
async def inspect_workbook(file: UploadFile = File(...)) -> dict[str, Any]:
    file_bytes = await file.read()
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload a valid .xlsx file.")
    return {
        "fileName": file.filename,
        "sheetNames": get_uploaded_sheet_names(file_bytes),
    }


@app.post("/api/profile/inspect")
async def inspect_profile_workbook(file: UploadFile = File(...)) -> dict[str, Any]:
    file_bytes = await file.read()
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload a valid .xlsx file.")
    try:
        with tempfile.TemporaryDirectory() as tmp_dir:
            uploaded_path = Path(tmp_dir) / (Path(file.filename).name or "profile.xlsx")
            uploaded_path.write_bytes(file_bytes)
            inspection = build_profile_editor_inspection(uploaded_path)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Could not inspect profile workbook: {exc}") from exc
    return {"fileName": file.filename, **inspection}


@app.post("/api/calculate")
async def calculate(
    file: UploadFile = File(...),
    plot_area_ha: float = Form(...),
    rai_per_hectare: float = Form(...),
    sheet_groups: str | None = Form(default=None),
    calculation_scope: str = Form(default="biomass_only"),
    economic_inputs: str | None = Form(default=None),
    future_interest_rate: float = Form(default=0.01),
    future_periods: str | None = Form(default=None),
) -> dict[str, Any]:
    if not MASTER_FILE.exists():
        raise HTTPException(status_code=500, detail="species_reference_master_v1.xlsx is missing.")

    file_bytes = await file.read()
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload a valid .xlsx file.")

    parsed_sheet_groups = parse_sheet_groups(sheet_groups)
    parsed_future_periods = parse_future_periods(future_periods)
    component_area_inputs, ecosystem_inputs = parse_economic_inputs(economic_inputs, parsed_sheet_groups)
    scope = normalize_text(calculation_scope).lower() or "biomass_only"
    if scope not in {"biomass_only", "economic_only", "biomass_and_economic"}:
        raise HTTPException(status_code=400, detail="Invalid calculation_scope.")
    if scope != "biomass_only" and not parsed_sheet_groups:
        raise HTTPException(status_code=400, detail="Economic calculation requires grouped components from Step 4.")
    if scope != "biomass_only" and not ecosystem_inputs:
        raise HTTPException(status_code=400, detail="Economic calculation requires per-component economic inputs.")

    workflow_cache_key = build_workflow_cache_key(
        file_bytes=file_bytes,
        plot_area_ha=plot_area_ha,
        rai_per_hectare=rai_per_hectare,
        sheet_groups=parsed_sheet_groups,
    )
    cached_workflow = get_cached_workflow(workflow_cache_key)
    if cached_workflow is not None:
        LOG.info("Using cached biomass workflow result.")
        summary_bytes, detail_bytes, component_bytes, result_sheets = cached_workflow
    else:
        try:
            summary_bytes, detail_bytes, component_bytes, result_sheets = run_uploaded_workflow(
                uploaded_filename=file.filename,
                file_bytes=file_bytes,
                plot_area_ha=plot_area_ha,
                rai_per_hectare=rai_per_hectare,
                sheet_groups=parsed_sheet_groups,
            )
            store_cached_workflow(
                workflow_cache_key,
                (summary_bytes, detail_bytes, component_bytes, result_sheets),
            )
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=500, detail=str(exc)) from exc

    try:
        biomass_payload = None
        if scope in {"biomass_only", "biomass_and_economic"}:
            biomass_payload = build_biomass_payload(
                result_sheets=result_sheets,
                sheet_groups=parsed_sheet_groups,
                plot_area_ha=plot_area_ha,
                rai_per_hectare=rai_per_hectare,
            )

        economic_payload = None
        economic_report_download = None
        economic_json_download = None
        if scope in {"economic_only", "biomass_and_economic"}:
            bundle = calculate_forest_valuation_bundle_from_outputs(
                outputs=result_sheets,
                component_area_inputs=component_area_inputs,
                ecosystem_user_inputs=ecosystem_inputs,
                future_interest_rate=future_interest_rate,
                future_periods_years=parsed_future_periods,
            )
            economic_payload = build_economic_preview(bundle, result_sheets)
            with tempfile.TemporaryDirectory() as tmp_dir:
                temp_dir = Path(tmp_dir)
                report_path = temp_dir / ECONOMIC_OUTPUT_FILENAME
                write_forest_economic_report(report_path, result_sheets, bundle)
                economic_report_download = serialize_download_payload(ECONOMIC_OUTPUT_FILENAME, report_path.read_bytes())
            economic_json_download = serialize_download_payload(
                ECONOMIC_JSON_FILENAME,
                json.dumps(bundle, ensure_ascii=False, indent=2).encode("utf-8"),
            )

        return jsonable_encoder(
            sanitize_for_json(
                {
                    "calculationScope": scope,
                    "biomass": biomass_payload,
                    "economic": economic_payload,
                    "downloads": {
                        "biomassSummary": serialize_download_payload(SUMMARY_OUTPUT_FILENAME, summary_bytes)
                        if scope in {"biomass_only", "biomass_and_economic"}
                        else None,
                        "biomassDetail": serialize_download_payload(DETAIL_OUTPUT_FILENAME, detail_bytes)
                        if scope in {"biomass_only", "biomass_and_economic"}
                        else None,
                        "biomassComponent": serialize_download_payload(COMPONENT_OUTPUT_FILENAME, component_bytes)
                        if scope in {"biomass_only", "biomass_and_economic"} and component_bytes is not None
                        else None,
                        "economicReport": economic_report_download,
                        "economicJson": economic_json_download,
                    },
                }
            )
        )
    except Exception as exc:  # noqa: BLE001
        LOG.exception("Failed to assemble calculate response.")
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/profile/calculate")
async def calculate_profile(
    file: UploadFile = File(...),
    render_mode: str = Form(default="graphic"),
) -> dict[str, Any]:
    file_bytes = await file.read()
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload a valid .xlsx file.")

    normalized_mode = normalize_profile_render_mode(render_mode)
    try:
        images, zip_bytes, output_filename = build_profile_outputs(file.filename, file_bytes, normalized_mode)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return {
        "sheetNames": [item["sheetName"] for item in images],
        "renderMode": normalized_mode,
        "images": images,
        "download": {
            "filename": output_filename,
            "contentBase64": base64.b64encode(zip_bytes).decode("ascii"),
        },
    }


@app.post("/api/profile/generation-job")
async def create_profile_generation_job(
    file: UploadFile = File(...),
    render_mode: str = Form(default="graphic"),
) -> dict[str, str]:
    file_bytes = await file.read()
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload a valid .xlsx file.")
    normalized_mode = normalize_profile_render_mode(render_mode)
    job_id = uuid.uuid4().hex
    with PROFILE_GENERATION_JOB_LOCK:
        PROFILE_GENERATION_JOBS[job_id] = {"status": "queued", "createdAt": time.time()}
        while len(PROFILE_GENERATION_JOBS) > PROFILE_GENERATION_JOB_LIMIT:
            _, expired_job = PROFILE_GENERATION_JOBS.popitem(last=False)
            expired_root = expired_job.get("jobRoot")
            if expired_root:
                shutil.rmtree(expired_root, ignore_errors=True)
    threading.Thread(
        target=run_profile_generation_job,
        args=(job_id, file.filename, file_bytes, normalized_mode),
        daemon=True,
    ).start()
    return {"jobId": job_id, "status": "queued"}


@app.get("/api/profile/generation-job/{job_id}")
def get_profile_generation_job(job_id: str) -> dict[str, Any]:
    with PROFILE_GENERATION_JOB_LOCK:
        job = PROFILE_GENERATION_JOBS.get(job_id)
        if job is None:
            raise HTTPException(status_code=404, detail="Profile generation job expired or was not found.")
        return {key: value for key, value in job.items() if key not in {"resultAssets", "jobRoot"}}


@app.get("/api/profile/generation-job/{job_id}/asset/{asset_key:path}")
def get_profile_generation_asset(job_id: str, asset_key: str) -> FileResponse:
    with PROFILE_GENERATION_JOB_LOCK:
        job = PROFILE_GENERATION_JOBS.get(job_id)
        asset = job.get("resultAssets", {}).get(asset_key) if job else None
    if asset is None:
        raise HTTPException(status_code=404, detail="Profile generation asset expired or was not found.")
    asset_file, media_type = asset
    if not asset_file.exists():
        raise HTTPException(status_code=404, detail="Profile generation asset expired or was not found.")
    return FileResponse(asset_file, media_type=media_type, headers={"Cache-Control": "private, max-age=3600"})


@app.get("/api/profile/generation-job/{job_id}/editor-bundle")
def get_profile_generation_editor_bundle(job_id: str) -> Response:
    with PROFILE_GENERATION_JOB_LOCK:
        job = PROFILE_GENERATION_JOBS.get(job_id)
        editor_scene = job.get("editorScene") if job else None
    if not editor_scene:
        raise HTTPException(status_code=404, detail="Editable profile layers are not ready.")
    session_id = editor_scene["sessionId"]
    with PROFILE_EDITOR_SCENE_LOCK:
        scene = PROFILE_EDITOR_SCENES.get(session_id)
        assets = dict(scene["assets"]) if scene else None
    if assets is None:
        raise HTTPException(status_code=404, detail="Editable profile layers expired or were not found.")

    url_prefix = f"/api/profile/editor-scene/{session_id}/asset/"
    offset = 0
    asset_records: list[dict[str, Any]] = []
    payload = bytearray()
    for asset_key, asset_file in assets.items():
        asset_bytes = asset_file.read_bytes()
        asset_records.append({"url": f"{url_prefix}{asset_key}", "offset": offset, "length": len(asset_bytes)})
        payload.extend(asset_bytes)
        offset += len(asset_bytes)
    manifest_bytes = json.dumps({"assets": asset_records}, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    content = len(manifest_bytes).to_bytes(4, "big") + manifest_bytes + bytes(payload)
    return Response(content=content, media_type="application/octet-stream")


@app.get("/api/profile/editor-scene/{session_id}")
def get_profile_editor_scene(session_id: str) -> dict[str, Any]:
    with PROFILE_EDITOR_SCENE_LOCK:
        scene = PROFILE_EDITOR_SCENES.get(session_id)
        manifest = scene.get("manifest") if scene else None
    if manifest is None:
        raise HTTPException(status_code=404, detail="Editable profile scene expired or was not found.")
    return manifest


@app.get("/api/profile/editor-scene/{session_id}/asset/{asset_key:path}")
def get_profile_editor_scene_asset(session_id: str, asset_key: str) -> FileResponse:
    with PROFILE_EDITOR_SCENE_LOCK:
        scene = PROFILE_EDITOR_SCENES.get(session_id)
        asset_file = scene.get("assets", {}).get(asset_key) if scene else None
    if asset_file is None or not asset_file.exists():
        raise HTTPException(status_code=404, detail="Editable profile layer expired or was not found.")
    return FileResponse(asset_file, media_type="image/png", headers={"Cache-Control": "private, max-age=3600"})


@app.post("/api/profile/editor/calculate")
async def calculate_profile_editor(
    file: UploadFile = File(...),
    sheet_name: str = Form(...),
    trees: str = Form(...),
    assignments: str | None = Form(default=None),
    transforms: str | None = Form(default=None),
    render_mode: str = Form(default="graphic"),
) -> dict[str, Any]:
    """Render edited profile data through the established production renderer.

    The selected render mode preserves the production graphic or realistic/illustrate
    output format. In realistic mode, V3's coherent asset group and tree transforms
    are applied to the selected worksheet before the production-format PNG is built.
    """
    file_bytes = await file.read()
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload a valid .xlsx file.")
    try:
        parsed_trees = json.loads(trees)
        if not isinstance(parsed_trees, list) or not all(isinstance(item, dict) for item in parsed_trees):
            raise ValueError("trees must be a JSON array of objects.")
        parsed_assignments = parse_profile_editor_json(assignments, "assignments")
        parsed_transforms = parse_profile_editor_json(transforms, "transforms")
        editor_file_bytes = apply_profile_editor_rows(file_bytes, sheet_name, parsed_trees)
        with tempfile.TemporaryDirectory() as tmp_dir:
            edited_path = Path(tmp_dir) / (Path(file.filename).name or "profile.xlsx")
            edited_path.write_bytes(editor_file_bytes)
            inspection = inspect_profile_workbook_data(edited_path)
        if sheet_name not in inspection["validSheetNames"]:
            issue = next(
                (
                    "; ".join(str(error) for error in sheet.get("errors", []))
                    for sheet in inspection["invalidSheets"]
                    if sheet.get("sheetName") == sheet_name
                ),
                "The edited worksheet failed profile validation.",
            )
            raise HTTPException(status_code=400, detail=issue)
        normalized_mode = normalize_profile_render_mode(render_mode)
        editor_overrides = build_profile_editor_render_overrides(
            sheet_name=sheet_name,
            parsed_trees=parsed_trees,
            assignments=parsed_assignments,
            transforms=parsed_transforms,
        )
        images, zip_bytes, output_filename = build_profile_outputs(
            file.filename,
            editor_file_bytes,
            normalized_mode,
            editor_sheet_name=sheet_name,
            editor_overrides=editor_overrides,
        )
    except HTTPException:
        raise
    except (TypeError, ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=400, detail=f"Invalid editor data: {exc}") from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return {
        "sheetNames": [item["sheetName"] for item in images],
        "renderMode": normalized_mode,
        "images": images,
        "download": {
            "filename": output_filename,
            "contentBase64": base64.b64encode(zip_bytes).decode("ascii"),
        },
        "editorApplied": True,
        "assetAssignmentsAccepted": assignments is not None,
        "transformsAccepted": transforms is not None,
        "assetAssignmentsApplied": normalized_mode == "realistic" and bool(editor_overrides[0]),
        "transformsApplied": normalized_mode == "realistic" and bool(editor_overrides[1] or editor_overrides[2]),
    }
