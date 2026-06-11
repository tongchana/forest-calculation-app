from __future__ import annotations

import os
import subprocess
import sys
import tempfile
import traceback
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

try:
    from streamlit_sortables import sort_items
except ImportError:  # pragma: no cover - optional UI enhancement
    sort_items = None

ROOT_DIR = Path(__file__).resolve().parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

import run_forest_calculation as calc


APP_TITLE = "Forest Field Studio"
APP_SUBTITLE = (
    "A calmer, more guided way to turn field survey workbooks into biomass, volume, IVI, and Shannon outputs."
)
PLOT_AREA_HA = calc.PLOT_AREA_HA
RAI_PER_HECTARE = calc.RAI_PER_HECTARE
TEMPLATE_FILE = ROOT_DIR / "template.xlsx"
MASTER_FILE = ROOT_DIR / "species_reference_master_v1.xlsx"
COMPONENT_TEMPLATE_FILE = ROOT_DIR / "forest_component_7.xlsx"
OUTPUT_BASE_FILENAME = "forest_calculation_output.xlsx"
SUMMARY_OUTPUT_FILENAME = "forest_calculation_output_summary_by_site.xlsx"
DETAIL_OUTPUT_FILENAME = "forest_calculation_output_details.xlsx"
COMPONENT_OUTPUT_FILENAME = "forest_component_summary.xlsx"
APP_ACCENT = "#2b7a5d"
APP_DEEP = "#163729"
APP_SOFT = "#eef5ef"
PREVIEW_SHEETS = [
    "SUMMARY_ALL",
    "SUMMARY_BIOMASS",
    "SUMMARY_VOLUME",
    "SUMMARY_SHANNON",
    "CHECK_UNMATCHED_SPECIES",
]
DEFAULT_GROUP_LABEL = "Component"
MAX_COMPONENTS = 7
SORTABLE_STYLE = """
.sortable-component {
    border: none;
    border-radius: 18px;
    padding: 0;
    background: transparent;
}
.sortable-container {
    background: #e8f2ec;
    border-radius: 14px;
    border: 1px solid rgba(36, 92, 63, 0.10);
    min-height: 86px;
    box-shadow: none;
    overflow: hidden;
}
.sortable-container-header {
    background: #e8f2ec;
    color: #1f6b4f;
    font-weight: 700;
    padding: 0.75rem 0.9rem;
    border-bottom: 1px solid rgba(36, 92, 63, 0.08);
}
.sortable-container-body {
    padding: 0.65rem;
    min-height: 56px;
    background: #e8f2ec;
}
.sortable-item, .sortable-item:hover {
    background: linear-gradient(135deg, #f4fbf6 0%, #ebf6ef 100%);
    border: 1px solid rgba(46, 125, 90, 0.16);
    color: #183c2c;
    border-radius: 10px;
    margin-bottom: 0.45rem;
    font-weight: 600;
}
"""


@st.cache_data(show_spinner=False)
def get_build_version() -> str:
    for env_name in ("STREAMLIT_GIT_COMMIT", "GITHUB_SHA", "COMMIT_SHA", "RENDER_GIT_COMMIT"):
        env_value = os.getenv(env_name, "").strip()
        if env_value:
            return env_value[:7]

    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=ROOT_DIR,
            check=True,
            capture_output=True,
            text=True,
        )
    except Exception:  # noqa: BLE001
        return "unknown"

    build = result.stdout.strip()
    return build or "unknown"


def inject_css() -> None:
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&family=DM+Serif+Display:ital@0;1&display=swap');
        :root {
            --forest-ink: #163729;
            --forest-green: #2b7a5d;
            --forest-green-soft: #eaf4ed;
            --forest-mist: #f5fbf6;
            --forest-gold: #cfb56b;
            --forest-line: rgba(22, 55, 41, 0.10);
        }
        .stApp {
            background:
                radial-gradient(circle at top left, rgba(111, 170, 133, 0.18), transparent 28%),
                radial-gradient(circle at 80% 12%, rgba(207, 181, 107, 0.16), transparent 20%),
                linear-gradient(180deg, #f7fbf7 0%, #eef5ef 52%, #f7fbf7 100%);
            color: var(--forest-ink);
            font-family: 'Manrope', sans-serif;
        }
        .main .block-container {
            max-width: 1220px;
            padding-top: 1.35rem;
            padding-bottom: 4rem;
        }
        h1, h2, h3 {
            font-family: 'DM Serif Display', serif !important;
            letter-spacing: -0.02em;
        }
        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, rgba(240, 247, 241, 0.98), rgba(230, 241, 233, 0.98));
            border-right: 1px solid var(--forest-line);
        }
        [data-testid="stSidebar"] * {
            color: var(--forest-ink);
        }
        .top-nav {
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 1rem;
            padding: 0.4rem 0 1rem;
        }
        .brand-pill {
            display: inline-flex;
            align-items: center;
            gap: 0.65rem;
            border-radius: 999px;
            padding: 0.65rem 1rem;
            background: rgba(255,255,255,0.72);
            border: 1px solid rgba(22, 55, 41, 0.08);
            box-shadow: 0 16px 40px rgba(22, 55, 41, 0.06);
            backdrop-filter: blur(8px);
        }
        .brand-seed {
            width: 0.95rem;
            height: 0.95rem;
            border-radius: 999px 0 999px 999px;
            background: linear-gradient(135deg, #8ed29d, #2b7a5d);
            transform: rotate(-22deg);
            box-shadow: 0 0 0 6px rgba(43, 122, 93, 0.08);
        }
        .brand-name {
            font-size: 0.92rem;
            font-weight: 800;
            letter-spacing: 0.08em;
            text-transform: uppercase;
        }
        .hero-shell {
            position: relative;
            overflow: hidden;
            border-radius: 36px;
            padding: 2.2rem 2.3rem;
            min-height: 360px;
            margin-bottom: 1.2rem;
            background:
                radial-gradient(circle at 0% 0%, rgba(255,255,255,0.22), transparent 25%),
                linear-gradient(135deg, #174131 0%, #2b7a5d 42%, #77a45f 100%);
            color: white;
            box-shadow: 0 28px 70px rgba(30, 84, 61, 0.18);
        }
        .hero-shell::before,
        .hero-shell::after {
            content: "";
            position: absolute;
            border-radius: 999px;
            filter: blur(3px);
            opacity: 0.35;
            animation: drift 10s ease-in-out infinite;
        }
        .hero-shell::before {
            width: 220px;
            height: 220px;
            top: -80px;
            right: 10%;
            background: rgba(255,255,255,0.18);
        }
        .hero-shell::after {
            width: 160px;
            height: 160px;
            bottom: -40px;
            left: 48%;
            background: rgba(233, 250, 233, 0.20);
            animation-delay: 1.4s;
        }
        .hero-grid {
            display: grid;
            grid-template-columns: minmax(0, 1.25fr) minmax(280px, 0.75fr);
            gap: 1.4rem;
            align-items: end;
        }
        .hero-eyebrow {
            display: inline-flex;
            align-items: center;
            gap: 0.6rem;
            border-radius: 999px;
            padding: 0.35rem 0.8rem;
            background: rgba(245, 255, 247, 0.14);
            border: 1px solid rgba(255,255,255,0.18);
            font-size: 0.8rem;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.08em;
        }
        .hero-title {
            font-size: clamp(2.8rem, 4.8vw, 4.6rem);
            line-height: 0.95;
            margin: 1rem 0 0.9rem;
            max-width: 8.7ch;
        }
        .hero-copy {
            font-size: 1.02rem;
            line-height: 1.75;
            max-width: 58ch;
            color: rgba(247, 255, 249, 0.92);
        }
        .hero-meta {
            display: flex;
            gap: 1rem;
            flex-wrap: wrap;
            margin-top: 1.2rem;
        }
        .hero-chip {
            border-radius: 999px;
            padding: 0.55rem 0.9rem;
            background: rgba(255,255,255,0.12);
            border: 1px solid rgba(255,255,255,0.12);
            font-size: 0.9rem;
            font-weight: 600;
        }
        .hero-panel {
            border-radius: 28px;
            padding: 1.1rem;
            background: rgba(250, 255, 251, 0.10);
            border: 1px solid rgba(255,255,255,0.16);
            backdrop-filter: blur(12px);
        }
        .hero-panel-title {
            text-transform: uppercase;
            letter-spacing: 0.08em;
            font-size: 0.75rem;
            opacity: 0.8;
            font-weight: 700;
        }
        .hero-panel-steps {
            display: grid;
            gap: 0.75rem;
            margin-top: 1rem;
        }
        .hero-step {
            display: grid;
            grid-template-columns: 30px 1fr;
            gap: 0.75rem;
            align-items: start;
            background: rgba(255,255,255,0.08);
            border-radius: 18px;
            padding: 0.75rem;
            animation: rise 0.7s ease both;
        }
        .hero-step-index {
            width: 30px;
            height: 30px;
            border-radius: 999px;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            background: rgba(255,255,255,0.2);
            font-weight: 700;
        }
        .story-card, .step-card, .metric-card, .result-shell {
            background: rgba(255,255,255,0.80);
            border-radius: 28px;
            border: 1px solid var(--forest-line);
            padding: 1.2rem 1.25rem;
            box-shadow: 0 22px 55px rgba(22, 55, 41, 0.06);
            backdrop-filter: blur(10px);
            margin-bottom: 1rem;
        }
        .story-card {
            padding: 1.25rem 1.35rem;
        }
        .step-title {
            color: var(--forest-green);
            font-weight: 800;
            font-size: 1.05rem;
            margin-bottom: 0.4rem;
        }
        .step-card {
            transition: transform 180ms ease, box-shadow 180ms ease;
        }
        .step-card:hover, .metric-card:hover {
            transform: translateY(-2px);
            box-shadow: 0 28px 60px rgba(22, 55, 41, 0.09);
        }
        .section-label {
            font-size: 0.78rem;
            text-transform: uppercase;
            letter-spacing: 0.14em;
            color: #6a8374;
            margin-bottom: 0.5rem;
            font-weight: 800;
        }
        .story-grid {
            display: grid;
            grid-template-columns: repeat(3, minmax(0, 1fr));
            gap: 1rem;
            margin: 1rem 0 1.35rem;
        }
        .story-kicker {
            color: #6a8374;
            font-size: 0.9rem;
            margin-bottom: 0.3rem;
        }
        .story-value {
            color: var(--forest-ink);
            font-size: 1.6rem;
            font-weight: 800;
            line-height: 1.1;
        }
        .metric-label {
            color: #5f7769;
            font-size: 0.86rem;
            margin-bottom: 0.3rem;
        }
        .metric-value {
            color: var(--forest-ink);
            font-size: 1.6rem;
            font-weight: 800;
            line-height: 1.2;
        }
        .metric-help {
            color: #6f8679;
            font-size: 0.82rem;
            margin-top: 0.28rem;
        }
        .warning-card {
            background: #fff7e8;
            border: 1px solid #ecd58f;
            border-radius: 20px;
            padding: 1rem 1.1rem;
            color: #7b5a18;
            margin-bottom: 1rem;
        }
        .success-card {
            background: #eef8f1;
            border: 1px solid #b8e1c2;
            border-radius: 20px;
            padding: 1rem 1.1rem;
            color: #205c37;
            margin-bottom: 1rem;
        }
        .download-ribbon {
            display: inline-flex;
            align-items: center;
            gap: 0.6rem;
            border-radius: 999px;
            padding: 0.45rem 0.85rem;
            background: rgba(43, 122, 93, 0.09);
            color: var(--forest-green);
            font-weight: 700;
            font-size: 0.84rem;
            margin-bottom: 0.9rem;
        }
        .result-shell {
            padding: 1.35rem;
        }
        .result-note {
            color: #688173;
            line-height: 1.7;
            margin-bottom: 0.75rem;
        }
        .stTabs [data-baseweb="tab-list"] {
            gap: 0.5rem;
            background: rgba(233, 243, 236, 0.9);
            padding: 0.35rem;
            border-radius: 999px;
        }
        .stTabs [data-baseweb="tab"] {
            border-radius: 999px;
            font-weight: 700;
            height: 42px;
        }
        div[data-testid="stDownloadButton"] button,
        div[data-testid="stButton"] button {
            border-radius: 999px;
            border: none;
            padding: 0.78rem 1.1rem;
            font-weight: 700;
            box-shadow: 0 16px 30px rgba(43, 122, 93, 0.10);
        }
        div[data-testid="stButton"] button[kind="primary"],
        div[data-testid="stDownloadButton"] button {
            background: linear-gradient(135deg, #245d46, #2f7f5f);
            color: white;
        }
        div[data-testid="stFileUploader"] section {
            border-radius: 24px;
            border: 1px dashed #89b39b;
            background: linear-gradient(180deg, #fbfefb, #f1f8f2);
            min-height: 160px;
        }
        div[data-testid="stDataFrame"] {
            border-radius: 18px;
            overflow: hidden;
            border: 1px solid rgba(36, 92, 63, 0.08);
        }
        .floating-note {
            color: #5d7568;
            font-size: 0.94rem;
            line-height: 1.7;
        }
        .stream-panel {
            padding: 1.15rem;
            border-radius: 24px;
            border: 1px solid var(--forest-line);
            background: rgba(255,255,255,0.74);
            box-shadow: 0 20px 48px rgba(22, 55, 41, 0.05);
        }
        .stream-panel h4 {
            margin: 0 0 0.45rem 0;
            font-family: 'DM Serif Display', serif !important;
            font-size: 1.25rem;
        }
        @keyframes drift {
            0%, 100% { transform: translateY(0px); }
            50% { transform: translateY(10px); }
        }
        @keyframes rise {
            from { opacity: 0; transform: translateY(14px); }
            to { opacity: 1; transform: translateY(0); }
        }
        @media (max-width: 980px) {
            .hero-grid, .story-grid {
                grid-template-columns: 1fr;
            }
            .hero-title {
                max-width: none;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_card(title: str, body: str) -> None:
    st.markdown(
        f"""
        <div class="step-card">
            <div class="step-title">{title}</div>
            <div>{body}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_metric(label: str, value: object, help_text: str = "") -> None:
    display_value = "-" if value is None or (isinstance(value, float) and pd.isna(value)) else value
    st.markdown(
        f"""
        <div class="metric-card">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{display_value}</div>
            <div class="metric-help">{help_text}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_top_nav() -> None:
    st.markdown(
        """
        <div class="top-nav">
            <div class="brand-pill">
                <div class="brand-seed"></div>
                <div class="brand-name">Forest Field Studio</div>
            </div>
            <div class="download-ribbon">Guided survey-to-summary experience</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_hero() -> None:
    st.markdown(
        f"""
        <div class="hero-shell">
            <div class="hero-grid">
                <div>
                    <div class="hero-eyebrow">Forest-inspired workflow</div>
                    <h1 class="hero-title">{APP_TITLE}</h1>
                    <div class="hero-copy">{APP_SUBTITLE}</div>
                    <div class="hero-meta">
                        <div class="hero-chip">Biomass, Volume, IVI, Shannon</div>
                        <div class="hero-chip">Worksheet grouping included</div>
                        <div class="hero-chip">Build {get_build_version()}</div>
                    </div>
                </div>
                <div class="hero-panel">
                    <div class="hero-panel-title">What happens here</div>
                    <div class="hero-panel-steps">
                        <div class="hero-step">
                            <div class="hero-step-index">1</div>
                            <div>Bring in a finished workbook and review every worksheet before calculation.</div>
                        </div>
                        <div class="hero-step">
                            <div class="hero-step-index">2</div>
                            <div>Optionally merge worksheets into custom components for combined outputs.</div>
                        </div>
                        <div class="hero-step">
                            <div class="hero-step-index">3</div>
                            <div>Download polished summary, detail, and component workbooks when processing completes.</div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_story_band() -> None:
    st.markdown(
        """
        <div class="story-grid">
            <div class="story-card">
                <div class="story-kicker">Soft guidance</div>
                <div class="story-value">3-step</div>
                <div class="floating-note">A calmer path through download, upload, and calculation without exposing raw backend complexity.</div>
            </div>
            <div class="story-card">
                <div class="story-kicker">Team-ready</div>
                <div class="story-value">7 components</div>
                <div class="floating-note">Group multiple worksheets into one named component while keeping original sheet outputs in the same run.</div>
            </div>
            <div class="story-card">
                <div class="story-kicker">Reference safety</div>
                <div class="story-value">Master-linked</div>
                <div class="floating-note">Uses the same calculation engine and species master as the production Streamlit app you already have today.</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def load_binary_file(file_path: Path) -> bytes:
    return file_path.read_bytes()


def get_uploaded_sheet_names(uploaded_file) -> list[str]:
    workbook = load_workbook(filename=BytesIO(uploaded_file.getvalue()), read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def get_calculation_sheet_names(sheet_names: list[str]) -> list[str]:
    return list(sheet_names)


def make_group_container(index: int) -> dict[str, list[str] | str]:
    return {"header": f"{DEFAULT_GROUP_LABEL} {index}", "items": []}


def order_items_by_reference(items: list[str], reference_order: list[str]) -> list[str]:
    position = {name: idx for idx, name in enumerate(reference_order)}
    unique_items = list(dict.fromkeys(items))
    return sorted(unique_items, key=lambda name: position.get(name, len(reference_order)))


def ensure_sheet_group_state(sheet_names: list[str]) -> None:
    signature = tuple(sheet_names)
    if st.session_state.get("sheet_group_signature") == signature:
        return

    st.session_state.sheet_group_signature = signature
    st.session_state.sheet_group_containers = [{"header": "Available sheets", "items": sheet_names.copy()}]
    st.session_state.sheet_group_name_count = 0
    st.session_state.sheet_group_sortable_version = 0


def bump_sheet_group_sortable_version() -> None:
    st.session_state.sheet_group_sortable_version = st.session_state.get("sheet_group_sortable_version", 0) + 1


def bump_batch_move_version() -> None:
    st.session_state.batch_move_version = st.session_state.get("batch_move_version", 0) + 1


def add_sheet_group() -> None:
    containers = st.session_state.sheet_group_containers
    if len(containers) - 1 >= MAX_COMPONENTS:
        return
    next_index = len(containers)
    containers.append(make_group_container(next_index))
    st.session_state.sheet_group_name_count = next_index
    bump_sheet_group_sortable_version()
    bump_batch_move_version()


def remove_sheet_group(reference_order: list[str]) -> None:
    containers = st.session_state.sheet_group_containers
    if len(containers) <= 1:
        return

    removed = containers.pop()
    available_items = containers[0]["items"] + removed["items"]
    containers[0]["items"] = order_items_by_reference(available_items, reference_order)
    st.session_state.sheet_group_name_count = len(containers) - 1
    bump_sheet_group_sortable_version()
    bump_batch_move_version()


def move_selected_sheets_to_group(selected_sheet_names: list[str], target_group_index: int, reference_order: list[str]) -> None:
    selected = [sheet_name for sheet_name in selected_sheet_names if sheet_name in reference_order]
    containers = st.session_state.sheet_group_containers
    if not selected or target_group_index <= 0 or target_group_index >= len(containers):
        return

    selected_set = set(selected)
    for container in containers:
        container["items"] = [item for item in container["items"] if item not in selected_set]

    target_items = containers[target_group_index]["items"] + selected
    containers[target_group_index]["items"] = order_items_by_reference(target_items, reference_order)
    containers[0]["items"] = order_items_by_reference(containers[0]["items"], reference_order)
    bump_sheet_group_sortable_version()
    bump_batch_move_version()


def move_selected_sheets_between_containers(
    selected_sheet_names: list[str],
    source_container_index: int,
    target_container_index: int,
    reference_order: list[str],
) -> None:
    containers = st.session_state.sheet_group_containers
    if (
        source_container_index < 0
        or target_container_index < 0
        or source_container_index >= len(containers)
        or target_container_index >= len(containers)
        or source_container_index == target_container_index
    ):
        return

    source_items = containers[source_container_index]["items"]
    selected = [sheet_name for sheet_name in selected_sheet_names if sheet_name in source_items and sheet_name in reference_order]
    if not selected:
        return

    selected_set = set(selected)
    containers[source_container_index]["items"] = [item for item in source_items if item not in selected_set]
    target_items = containers[target_container_index]["items"] + selected
    containers[target_container_index]["items"] = order_items_by_reference(target_items, reference_order)
    containers[0]["items"] = order_items_by_reference(containers[0]["items"], reference_order)
    bump_sheet_group_sortable_version()
    bump_batch_move_version()


def normalize_sortable_containers(
    containers: list[dict[str, list[str] | str]],
    sheet_names: list[str],
) -> list[dict[str, list[str] | str]]:
    allowed = set(sheet_names)
    assigned: set[str] = set()
    normalized: list[dict[str, list[str] | str]] = []

    for idx, container in enumerate(containers):
        items = [item for item in container.get("items", []) if item in allowed]
        deduped: list[str] = []
        for item in items:
            if item in assigned:
                continue
            assigned.add(item)
            deduped.append(item)
        header = "Available sheets" if idx == 0 else container.get("header", f"{DEFAULT_GROUP_LABEL} {idx}")
        normalized.append({"header": header, "items": deduped})

    unassigned = [sheet_name for sheet_name in sheet_names if sheet_name not in assigned]
    if normalized:
        normalized[0]["items"] = order_items_by_reference(normalized[0]["items"] + unassigned, sheet_names)
    return normalized


def render_sheet_group_builder(sheet_names: list[str]) -> list[dict[str, list[str]]]:
    ensure_sheet_group_state(sheet_names)

    st.markdown('<div class="section-label">Optional grouping</div>', unsafe_allow_html=True)
    render_card(
        "Combine Multiple Sheets Into One Calculation Component",
        "Drag worksheets into custom components to calculate combined IVI, biomass, volume, and Shannon results in addition to the normal per-sheet outputs.",
    )

    action_col1, action_col2 = st.columns([1, 1])
    with action_col1:
        st.button(
            "Add component",
            on_click=add_sheet_group,
            disabled=len(st.session_state.sheet_group_containers) - 1 >= MAX_COMPONENTS,
            use_container_width=True,
        )
    with action_col2:
        st.button(
            "Remove last component",
            on_click=remove_sheet_group,
            args=(sheet_names,),
            disabled=len(st.session_state.sheet_group_containers) <= 1,
            use_container_width=True,
        )

    group_count = len(st.session_state.sheet_group_containers) - 1
    st.caption(f"You can create up to {MAX_COMPONENTS} components.")
    for idx in range(1, group_count + 1):
        group_key = f"sheet_group_name_{idx}"
        if group_key not in st.session_state:
            st.session_state[group_key] = f"IVI {DEFAULT_GROUP_LABEL.lower()} {idx}"
        st.text_input(f"Name for {DEFAULT_GROUP_LABEL.lower()} {idx}", key=group_key)

    groups: list[dict[str, list[str]]] = []

    builder_mode = "Batch move"
    if sort_items is not None:
        builder_mode = st.radio(
            "Component builder mode",
            options=["Batch move", "Drag and drop", "Simple selection"],
            index=0,
            horizontal=True,
            help="Batch move is the fastest way to assign many worksheets at once. Drag and drop and Simple selection are still available if you prefer them.",
        )

    if builder_mode == "Simple selection":
        assigned: set[str] = set()
        for idx in range(1, group_count + 1):
            selection_key = f"sheet_group_items_{idx}"
            current_selected = [item for item in st.session_state.get(selection_key, []) if item in sheet_names]
            available_options = [item for item in sheet_names if item not in assigned or item in current_selected]
            st.markdown(
                f"""
                <div class="step-card">
                    <div class="step-title">Component {idx}</div>
                    <div>Select one or more worksheets to combine under this component.</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            selected_items = st.multiselect(
                f"Worksheets in component {idx}",
                options=available_options,
                default=current_selected,
                key=selection_key,
                placeholder="Choose worksheets for this component",
            )
            assigned.update(selected_items)
            group_name = (st.session_state.get(f"sheet_group_name_{idx}") or "").strip()
            if selected_items and not group_name:
                st.warning(f"Please enter a name for {DEFAULT_GROUP_LABEL.lower()} {idx} before calculation.")
            elif selected_items:
                groups.append({"name": group_name, "sheet_names": selected_items})

        remaining_sheets = [item for item in sheet_names if item not in assigned]
        if remaining_sheets:
            st.caption(f"Still ungrouped: {', '.join(remaining_sheets)}")
        else:
            st.caption("All worksheets have been assigned to components.")
    elif builder_mode == "Batch move":
        st.markdown(
            """
            <div class="step-card">
                <div class="step-title">Batch move worksheets</div>
                <div>
                    Select one source list, choose multiple worksheets, then move them to another list in one click.
                    This is the recommended workflow when you want something like Ctrl/Shift multi-select.
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        container_labels = []
        for idx, container in enumerate(st.session_state.sheet_group_containers):
            if idx == 0:
                container_labels.append("Available sheets")
            else:
                group_name = (st.session_state.get(f"sheet_group_name_{idx}") or container["header"]).strip()
                container_labels.append(f"Component {idx}: {group_name}")

        batch_col1, batch_col2, batch_col3 = st.columns([1.2, 2.2, 1.2])
        with batch_col1:
            source_label = st.selectbox(
                "Source list",
                options=container_labels,
                key="batch_move_source_label",
            )
        source_index = container_labels.index(source_label)
        source_items = st.session_state.sheet_group_containers[source_index]["items"]
        batch_move_key_suffix = st.session_state.get("batch_move_version", 0)

        with batch_col2:
            batch_selected = st.multiselect(
                "Select worksheets",
                options=source_items,
                key=f"batch_move_sheet_names_{batch_move_key_suffix}",
                placeholder="Choose one or more worksheets",
            )

        with batch_col3:
            target_candidates = [label for idx, label in enumerate(container_labels) if idx != source_index]
            if target_candidates:
                target_label = st.selectbox(
                    "Move to",
                    options=target_candidates,
                    key=f"batch_move_target_label_{batch_move_key_suffix}",
                )
            else:
                target_label = None
                st.selectbox(
                    "Move to",
                    options=["No destination available"],
                    key="batch_move_target_disabled",
                    disabled=True,
                )

        if st.button(
            "Move selected worksheets",
            use_container_width=True,
            disabled=not batch_selected or not target_candidates,
        ):
            move_selected_sheets_between_containers(
                batch_selected,
                source_index,
                container_labels.index(target_label),
                sheet_names,
            )
            st.rerun()

        preview_cols = st.columns(min(max(len(st.session_state.sheet_group_containers), 1), 4))
        for idx, container in enumerate(st.session_state.sheet_group_containers):
            col = preview_cols[idx % len(preview_cols)]
            title = container_labels[idx]
            items = container["items"]
            body = "<br>".join(items) if items else "<span style='color:#70877b;'>No worksheets</span>"
            with col:
                st.markdown(
                    f"""
                    <div class="step-card">
                        <div class="step-title">{title}</div>
                        <div>{body}</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

        for idx, container in enumerate(st.session_state.sheet_group_containers[1:], start=1):
            group_name = (st.session_state.get(f"sheet_group_name_{idx}") or "").strip()
            group_items = container["items"]
            if not group_items:
                continue
            if not group_name:
                st.warning(f"Please enter a name for {DEFAULT_GROUP_LABEL.lower()} {idx} before calculation.")
                continue
            groups.append({"name": group_name, "sheet_names": group_items})
    else:
        sortable_key = f"sheet_group_sortable_{st.session_state.get('sheet_group_sortable_version', 0)}_{group_count}"
        sortable_containers = sort_items(
            st.session_state.sheet_group_containers,
            multi_containers=True,
            custom_style=SORTABLE_STYLE,
            key=sortable_key,
        )
        st.session_state.sheet_group_containers = normalize_sortable_containers(sortable_containers, sheet_names)

        for idx, container in enumerate(st.session_state.sheet_group_containers[1:], start=1):
            group_name = (st.session_state.get(f"sheet_group_name_{idx}") or "").strip()
            group_items = container["items"]
            if not group_items:
                continue
            if not group_name:
                st.warning(f"Please enter a name for {DEFAULT_GROUP_LABEL.lower()} {idx} before calculation.")
                continue
            groups.append({"name": group_name, "sheet_names": group_items})

    if groups:
        preview_rows = [{"Component name": group["name"], "Sheets": ", ".join(group["sheet_names"])} for group in groups]
        st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)
        st.caption("Each worksheet can belong to one component in this drag-and-drop board.")
    else:
        st.caption("No combined components defined. The app will calculate each worksheet separately, like before.")

    return groups


def render_sidebar() -> None:
    st.sidebar.markdown("## Forest Field Studio")
    st.sidebar.write(
        "A Forest-inspired Streamlit surface for the same calculation engine. It keeps the workflow familiar while making the experience calmer, softer, and easier to scan."
    )

    st.sidebar.markdown("## What It Calculates")
    scope_df = pd.DataFrame(
        [
            ["Tree", "Yes", "Yes", "Yes", "Yes"],
            ["Sapling", "No", "Yes", "No", "Yes"],
            ["Seedling", "No", "No", "No", "Yes"],
            ["Bamboo", "No", "No", "No", "Yes"],
        ],
        columns=["Block", "Biomass", "Volume", "IVI/Shannon", "Count summary"],
    )
    st.sidebar.table(scope_df)

    st.sidebar.markdown("## Recommended Flow")
    st.sidebar.write(
        "Download the official template, fill Tree, Sapling, Seedling, and Bamboo data offline, then upload the workbook here for a guided processing run."
    )

    st.sidebar.markdown("## Important Note")
    st.sidebar.warning("Biomass is calculated for Tree only. Sapling is excluded from biomass outputs.")


def format_metric_value(value: object, decimals: int = 2) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "-"
    if isinstance(value, (int, float)):
        return f"{value:,.{decimals}f}" if isinstance(value, float) or decimals else f"{int(value):,}"
    return str(value)


def build_metrics(summary_all: pd.DataFrame, unmatched: pd.DataFrame) -> list[tuple[str, str, str]]:
    if summary_all.empty:
        return []

    total_tree = pd.to_numeric(summary_all.get("n_tree"), errors="coerce").fillna(0).sum()
    total_sapling = pd.to_numeric(summary_all.get("n_sapling"), errors="coerce").fillna(0).sum()
    total_tree_biomass = pd.to_numeric(summary_all.get("total_tree_biomass"), errors="coerce").fillna(0).sum()
    total_tree_volume = pd.to_numeric(summary_all.get("total_tree_volume_m3"), errors="coerce").fillna(0).sum()
    total_sapling_volume = pd.to_numeric(summary_all.get("total_sapling_volume_m3"), errors="coerce").fillna(0).sum()

    shannon_series = pd.to_numeric(summary_all.get("shannon_index"), errors="coerce")
    shannon_value = shannon_series.mean() if shannon_series is not None and not shannon_series.dropna().empty else None
    unmatched_count = len(unmatched.index) if not unmatched.empty else 0

    return [
        ("Total tree count", format_metric_value(total_tree, 0), "Across all processed worksheets"),
        ("Total sapling count", format_metric_value(total_sapling, 0), "Sapling records included in volume"),
        ("Total tree biomass", format_metric_value(total_tree_biomass, 2), "Tree biomass only"),
        ("Total tree volume", format_metric_value(total_tree_volume, 3), "Tree block volume"),
        ("Total sapling volume", format_metric_value(total_sapling_volume, 3), "Sapling block volume"),
        ("Shannon index", format_metric_value(shannon_value, 6), "Average across available sites"),
        ("Unmatched species", format_metric_value(unmatched_count, 0), "Species needing reference review"),
    ]


def safe_sheet(frame_dict: dict[str, pd.DataFrame], sheet_name: str) -> pd.DataFrame:
    frame = frame_dict.get(sheet_name)
    return frame if isinstance(frame, pd.DataFrame) else pd.DataFrame()


def preview_results(result_sheets: dict[str, pd.DataFrame]) -> None:
    summary_all = safe_sheet(result_sheets, "SUMMARY_ALL")
    summary_biomass = safe_sheet(result_sheets, "SUMMARY_BIOMASS")
    summary_volume = safe_sheet(result_sheets, "SUMMARY_VOLUME")
    summary_shannon = safe_sheet(result_sheets, "SUMMARY_SHANNON")
    unmatched = safe_sheet(result_sheets, "CHECK_UNMATCHED_SPECIES")

    st.markdown('<div class="section-label">Step 4 · Review the canopy</div>', unsafe_allow_html=True)
    st.markdown(
        """
        <div class="result-shell">
            <div class="step-title">Preview Summary Results</div>
            <div class="result-note">Review the main summary sheets before downloading the full calculated workbooks. The metrics below are pulled directly from the current run.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    metrics = build_metrics(summary_all, unmatched)
    if metrics:
        for start in range(0, len(metrics), 3):
            cols = st.columns(3)
            for col, metric in zip(cols, metrics[start : start + 3]):
                with col:
                    render_metric(*metric)
    else:
        st.info("No summary metrics are available yet.")

    tabs = st.tabs(["Overall Summary", "Biomass", "Volume", "Shannon / IVI", "Unmatched Species"])
    with tabs[0]:
        if summary_all.empty:
            st.info("SUMMARY_ALL is empty.")
        else:
            st.dataframe(summary_all, use_container_width=True)
    with tabs[1]:
        if summary_biomass.empty:
            st.info("SUMMARY_BIOMASS is empty.")
        else:
            st.dataframe(summary_biomass, use_container_width=True)
    with tabs[2]:
        if summary_volume.empty:
            st.info("SUMMARY_VOLUME is empty.")
        else:
            st.dataframe(summary_volume, use_container_width=True)
    with tabs[3]:
        if summary_shannon.empty:
            st.info("SUMMARY_SHANNON is empty.")
        else:
            st.dataframe(summary_shannon, use_container_width=True)
    with tabs[4]:
        if unmatched.empty:
            st.success("No unmatched species were found in the uploaded workbook.")
        else:
            st.markdown(
                '<div class="warning-card">Some species could not be matched with the master reference file. Please review the table below.</div>',
                unsafe_allow_html=True,
            )
            st.dataframe(unmatched, use_container_width=True)


def run_uploaded_workflow(
    uploaded_file,
    plot_area_ha: float,
    rai_per_hectare: float,
    sheet_groups: list[dict[str, list[str]]] | None = None,
) -> tuple[bytes, bytes, bytes | None, dict[str, pd.DataFrame]]:
    with tempfile.TemporaryDirectory() as tmp_dir:
        temp_dir = Path(tmp_dir)
        uploaded_path = temp_dir / uploaded_file.name
        uploaded_path.write_bytes(uploaded_file.getbuffer())

        output_base = temp_dir / OUTPUT_BASE_FILENAME
        split_runner = getattr(calc, "run_calculation_split_outputs", None)
        if split_runner is not None:
            summary_path, detail_path, result_sheets = split_runner(
                input_file=uploaded_path,
                master_file=MASTER_FILE,
                output_base=output_base,
                plot_area_ha=plot_area_ha,
                rai_per_hectare=rai_per_hectare,
                sheet_groups=sheet_groups,
            )
        else:
            result_sheets = calc.process_workbook(
                input_file=uploaded_path,
                master_file=MASTER_FILE,
                plot_area_ha=plot_area_ha,
                rai_per_hectare=rai_per_hectare,
                sheet_groups=sheet_groups,
            )
            summary_path, detail_path = calc.resolve_output_paths(uploaded_path, str(output_base))
            calc.write_summary_by_site_workbook(summary_path, result_sheets)
            calc.write_detail_workbook(detail_path, result_sheets)
        component_bytes = None
        if sheet_groups and COMPONENT_TEMPLATE_FILE.exists():
            component_path = temp_dir / COMPONENT_OUTPUT_FILENAME
            calc.write_component_summary_workbook(
                component_path,
                COMPONENT_TEMPLATE_FILE,
                result_sheets,
                summary_file=summary_path,
            )
            component_bytes = component_path.read_bytes()
        return summary_path.read_bytes(), detail_path.read_bytes(), component_bytes, result_sheets


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide", page_icon="🌿")
    inject_css()
    render_sidebar()

    if "summary_result_bytes" not in st.session_state:
        st.session_state.summary_result_bytes = None
    if "detail_result_bytes" not in st.session_state:
        st.session_state.detail_result_bytes = None
    if "component_result_bytes" not in st.session_state:
        st.session_state.component_result_bytes = None
    if "result_sheets" not in st.session_state:
        st.session_state.result_sheets = None
    if "last_error" not in st.session_state:
        st.session_state.last_error = None
    if "sheet_group_containers" not in st.session_state:
        st.session_state.sheet_group_containers = [{"header": "Available sheets", "items": []}]
    if "sheet_group_signature" not in st.session_state:
        st.session_state.sheet_group_signature = ()
    if "sheet_group_name_count" not in st.session_state:
        st.session_state.sheet_group_name_count = 0
    if "batch_move_version" not in st.session_state:
        st.session_state.batch_move_version = 0

    render_top_nav()
    render_hero()
    render_story_band()

    st.markdown('<div class="section-label">Experience overview</div>', unsafe_allow_html=True)
    st.markdown(
        """
        <div class="stream-panel">
            <h4>From field workbook to clean forest outputs</h4>
            <div class="floating-note">
                This alternate app keeps the original Python calculation workflow, but wraps it in a more guided, polished experience.
                Download the template, upload your finished workbook, optionally build components, then calculate and review the generated summaries.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if not TEMPLATE_FILE.exists():
        st.error("The official template file 'template.xlsx' is missing from the project directory.")
    if not MASTER_FILE.exists():
        st.error("The species reference file 'species_reference_master_v1.xlsx' is missing from the project directory.")

    col_download, col_upload = st.columns([1, 1.2], gap="large")
    with col_download:
        st.markdown('<div class="section-label">Step 1 · Prepare the workbook</div>', unsafe_allow_html=True)
        render_card("Download The Official Template", "Start from the master workbook so the layout, header rows, and downstream calculation blocks stay aligned.")
        if TEMPLATE_FILE.exists():
            st.download_button(
                "Download Template Workbook",
                data=load_binary_file(TEMPLATE_FILE),
                file_name=TEMPLATE_FILE.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    with col_upload:
        st.markdown('<div class="section-label">Step 2 · Bring your survey in</div>', unsafe_allow_html=True)
        render_card("Upload Completed Survey Workbook", "Drop in the filled Excel file and the app will detect every worksheet before calculation starts.")
        uploaded_file = st.file_uploader("Upload completed Excel template", type=["xlsx"], label_visibility="collapsed")

    selected_sheet_groups: list[dict[str, list[str]]] = []
    if uploaded_file is not None:
        try:
            uploaded_sheet_names = get_uploaded_sheet_names(uploaded_file)
            calculation_sheet_names = get_calculation_sheet_names(uploaded_sheet_names)
        except Exception:  # noqa: BLE001
            uploaded_sheet_names = []
            calculation_sheet_names = []
            st.warning("The workbook was uploaded, but its worksheet list could not be previewed yet.")
        else:
            if uploaded_sheet_names:
                st.caption(f"Detected worksheets: {', '.join(uploaded_sheet_names)}")
            if calculation_sheet_names:
                selected_sheet_groups = render_sheet_group_builder(calculation_sheet_names)
    if selected_sheet_groups and not COMPONENT_TEMPLATE_FILE.exists():
        st.warning("The component summary template file 'forest_component_7.xlsx' is missing, so the extra component-summary download will not be available.")

    st.markdown('<div class="section-label">Step 3 · Grow the output</div>', unsafe_allow_html=True)
    render_card("Run The Calculation Workflow", "Use the same production calculation engine and species master, with adjustable plot area and rai conversion before processing.")
    calc_col1, calc_col2, calc_col3 = st.columns([1, 1, 1.4])
    with calc_col1:
        plot_area_ha = st.number_input("Plot area (ha)", min_value=0.0001, value=float(PLOT_AREA_HA), step=0.1, format="%.4f")
    with calc_col2:
        rai_per_hectare = st.number_input(
            "Rai per hectare",
            min_value=0.0001,
            value=float(RAI_PER_HECTARE),
            step=0.25,
            format="%.4f",
        )
    with calc_col3:
        calculate_clicked = st.button("Calculate", type="primary", use_container_width=True)

    if calculate_clicked:
        st.session_state.last_error = None
        st.session_state.summary_result_bytes = None
        st.session_state.detail_result_bytes = None
        st.session_state.component_result_bytes = None
        st.session_state.result_sheets = None

        if uploaded_file is None:
            st.error("Please upload a completed Excel template before running the calculation.")
        elif uploaded_file.type not in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/octet-stream",
        ):
            st.error("Please upload a valid .xlsx Excel file.")
        elif not TEMPLATE_FILE.exists():
            st.error("The official template file 'template.xlsx' is missing from the project directory.")
        elif not MASTER_FILE.exists():
            st.error("The species reference file 'species_reference_master_v1.xlsx' is missing from the project directory.")
        else:
            with st.spinner("Running forest calculation workflow..."):
                try:
                    summary_result_bytes, detail_result_bytes, component_result_bytes, result_sheets = run_uploaded_workflow(
                        uploaded_file=uploaded_file,
                        plot_area_ha=plot_area_ha,
                        rai_per_hectare=rai_per_hectare,
                        sheet_groups=selected_sheet_groups,
                    )
                    st.session_state.summary_result_bytes = summary_result_bytes
                    st.session_state.detail_result_bytes = detail_result_bytes
                    st.session_state.component_result_bytes = component_result_bytes
                    st.session_state.result_sheets = result_sheets
                    st.success("Calculation completed successfully.")
                    st.info("The new canopy is ready. Preview the results below, then download the generated workbooks.")
                except Exception as exc:  # noqa: BLE001
                    st.session_state.last_error = traceback.format_exc()
                    st.error(
                        "The calculation could not be completed. Please check that the uploaded file uses the official template and contains valid Excel data."
                    )
                    if isinstance(exc, FileNotFoundError):
                        st.error(str(exc))
                    with st.expander("Technical error details"):
                        st.code(st.session_state.last_error)

    if st.session_state.result_sheets:
        preview_results(st.session_state.result_sheets)

        st.markdown('<div class="section-label">Step 5 · Take the files with you</div>', unsafe_allow_html=True)
        render_card(
            "Download Calculated Workbooks",
            "Export the summary-by-site workbook, detailed workbook, and the component-summary workbook when grouped components are present.",
        )
        if st.session_state.component_result_bytes:
            dl_col1, dl_col2, dl_col3 = st.columns(3)
        else:
            dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            st.download_button(
                "Download Summary-by-site Excel",
                data=st.session_state.summary_result_bytes,
                file_name=SUMMARY_OUTPUT_FILENAME,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with dl_col2:
            st.download_button(
                "Download Details Excel",
                data=st.session_state.detail_result_bytes,
                file_name=DETAIL_OUTPUT_FILENAME,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        if st.session_state.component_result_bytes:
            with dl_col3:
                st.download_button(
                    "Download Component Summary Excel",
                    data=st.session_state.component_result_bytes,
                    file_name=COMPONENT_OUTPUT_FILENAME,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )


if __name__ == "__main__":
    main()
